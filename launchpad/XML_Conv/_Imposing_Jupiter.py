from functools import partial
import re
import shutil
import sys
import traceback
from os.path import join
from pathlib import Path

from openpyxl import load_workbook

try:
    from launchpad.functions import ConversionClient
except ImportError:
    pass

months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May',
          'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# IPL patterns
pattern_partspec = re.compile(
    r"(<partSpec>.*?<partIdent manufacturerCodeValue=\"([^\"]*)\"\r?\s?"
    r"partNumberValue=\"([^\"]*)\"/>.*?</partSpec>)",
    re.S)
pattern_partkw = re.compile(r"<partKeyword>([^<]*)<")
pattern_shortname = re.compile(r"<shortName>([^<]*)<")
pattern_eref = re.compile(r"enterpriseRef\smanufacturerCodeValue=\"([^\"]+)\"")
pattern_dpl2 = re.compile(
    r"(pmEntryType=\"pmt90\".*?</pmEntry>\n?)<pmEntry.*?(?=</content>)", re.S)
pattern_booktitle = re.compile(
    r"<partinfo[^<]+<(?:title|cmpnom)>([^<]+)</(?:title|cmpnom)>")
pattern_unit = re.compile(
    r"<unit([^<]+)<title>([^<]+)</title>(.*?)</unit>", re.S)
pattern_sect = re.compile(
    r"<section([^<]+)<title>([^<]+)</title>(.*?)</section>", re.S)
pattern_chapt = re.compile(
    r"<chapter([^<]+)<title>([^<]+)</title>(.*?)</chapter>", re.S)
pattern_fig = re.compile(r"<figure([^>]+)>(.*?)</figure>", re.S)
pattern_chapnbr = re.compile(r"chapnbr=\"(\d{2})\"")
pattern_sectnbr = re.compile(r"sectnbr=\"(\d{2})\"")
pattern_unitnbr = re.compile(r"unitnbr=\"(\d{2})\"")
pattern_fignbr = re.compile(r"fignbr=\"([^\"]+)\"")
pattern_graphic = re.compile(r"<graphic(.*?)</graphic>", re.S)
pattern_key = re.compile(r"key=\"([^\"]+)\"")
pattern_title = re.compile(r"<title>([^<]+)</title>")
pattern_sheet = re.compile(r"<sheet([^<]+)</sheet>")
pattern_gnbr = re.compile(r"gnbr=\"([^\"]+)\"")
pattern_imgarea = re.compile(r"imgarea=\"(\w{2})\"")
pattern_itemdata = re.compile(r"<item([^>]+)>(.*?)</item>", re.S)
pattern_itemnbr = re.compile(r"itemnbr=\"([^\"]+)\"")
pattern_attach = re.compile(r"attach=\"([^\"]+)\"")
pattern_illusind = re.compile(r"illusind=\"([^\"]+)\"")
pattern_indent = re.compile(r"indent=\"([^\"]+)\"")
pattern_chg = re.compile(r"chg=\"([^\"]+)\"")
pattern_tags = re.compile(r"(<(?!nom)(\w+)[^>]*>(?:(.*?)</\2>)?)", re.S)
pattern_sb = re.compile(r"sbtype=\"([^\"]+)\">([^<]+)")
pattern_mfr = re.compile(r"<mfr>(\w{5})</mfr>")
pattern_reffig = re.compile(
    r"FIG(?:URE)?\s?(\d+\w?)((?:,?\s?(?:AND)? \d+\w?(?=\s|<|,))*)")
pattern_refsect = re.compile(r"\d\d-\d\d-\d\d")
pattern_refitem = re.compile(r"ITEM\s?(\w+)")

# SBL, RTR, VCL
pattern_vendata = re.compile(r"(<vendata>.*?</vendata>)", re.S)
pattern_v = re.compile(r"row rowsep=\"0\">(\n)?<entry><para>([^V])")
pattern_mfrcont = re.compile(r"</?mfrcont>")
pattern_sblist = re.compile(r"<sblist.*?</sblist>", re.S)
pattern_sbdata = re.compile(r"(<sbdata>.*?</sbdata>)", re.S)
pattern_sbdatatag = re.compile(r"</?sbdata>")
pattern_trlist = re.compile(r"<trlist.*?</trlist>", re.S)
pattern_trdata = re.compile(r"<trdata>.*?</trdata>", re.S)

# Table data/links
pattern_trans = re.compile(
    r"(?<=<title>TRANSMITTAL INFORMATION</title>)\n?(<para>(.*?)</para>)",
    re.S)
pattern_table = re.compile(r"<table.*?</table>", re.S)
pattern_rev = re.compile(r"<\?Pub\s?\n?/?_rev\n?\?>")
pattern_thead = re.compile(r"<thead.*?</thead>", re.S)
pattern_tbody = re.compile(r"<tbody.*?</tbody>", re.S)
pattern_para = re.compile(r"<para>(.*?)</para>", re.S)
pattern_iref = re.compile(
    r"<internalRef.*?</internalRef>(?!\sthrough|\sand)", re.S)
pattern_rht = re.compile(r'<\?PubTbl row rht[^>]+>\n?')
# Copyright
pattern_cpyrght = re.compile(
    r"(?s)<cpyrght>.*?(\d{4},\s\d{4}).*?</cpyrght>", re.S)
pattern_years = re.compile(r"\d{4},\s\d{4}")

# 00PA
pattern_partinfo = re.compile(r"<partinfo.*?</partinfo>", re.S)
pattern_mfrpnr = re.compile(r"<mfrpnr>.*?</mfrpnr>", re.S)
pattern_pnr = re.compile(r"<pnr>([^<]+)</pnr>", re.S)
pattern_model = re.compile(r'model=([^>]+)')
pattern_prodxref = re.compile(
    r"<productCrossRefTable.*?</productCrossRefTable>", re.S)

# PMC
pattern_shrtpmtitle = re.compile(r'model="([^\"]+)')
pattern_eipc = re.compile(r'<(?:eipc|cmm)([^>]+)>')
pattern_oidate = re.compile(r'oidate="([^"]+)"')
pattern_docnbr = re.compile(r'docnbr="([^"]+)"')
pattern_tsn = re.compile(r'tsn="([^"]+)"')
pattern_revdate = re.compile(r'revdate="([^"]+)"')
pattern_pmtitle = re.compile(r'<pmTitle>[^<]+</pmTitle>')
pattern_shorttitle = re.compile(r'<shortPmTitle>[^<]+</shortPmTitle>')
pattern_inidate = re.compile(
    r'<externalPubCode pubCodingScheme="initialDate">.*?</externalPubCode>',
    re.S)
pattern_int = re.compile(
    r'<externalPubCode pubCodingScheme="INT">.*?</externalPubCode>', re.S)
pattern_issnum = re.compile(r'issueNumber="[^"]+"')
pattern_issdate = re.compile(r'<issueDate[^/]*/>')
pattern_eccn = re.compile(r'ECCN:\s?[^<]+')
pattern_authdoc = re.compile(r'authorityDocument="[^"]+" pmEntryType="pmt56"')

pattern_ftnote = re.compile(r'<ftnote.*?</ftnote>', re.S)
pattern_ftnotesub = re.compile(
    r'</tbody>.*?</ftnote>(?=\s?</table>)', re.S)
pattern_techname = re.compile(r'<techName>[^<]+</techName>')
pattern_note = re.compile(r'<note>.*?</note>', re.S)
pattern_refint = re.compile(r'<refint\srefid="([^"]+)"></refint>')
pattern_entity = re.compile(r'&(.+?);')
pattern_howto = re.compile(
    r'<title>INTRODUCTION</title>\n?.*?(<list2>.*?</list2>)', re.S)

pattern_fixsteps_table = re.compile(
    r'((?:</levelledPara>\n?)*</levelledPara>)\n*(<table.*?</table>)', re.S)
pattern_fixsteps = re.compile(
    r'((?:</levelledPara>\n?)*)<levelledPara>\n*'
    r'(<(?:table|para><(?:random|sequential)List>|figure).*?'
    r'</(?:table|(?:random|sequential)List></para|figure)>)\n*</levelledPara>',
    re.S)
pattern_fixsteps2 = re.compile(
    r'((?:</levelledPara>\n?)*</levelledPara>)\n*'
    r'(<(?:para><(?:random|sequential)List>|figure).*?'
    r'</(?:(?:random|sequential)List></para|figure)>)\n*(</levelledPara>)?',
    re.S)


def exit_handler_partial(logText, status=0, filename="log.txt", cc=None):
    if cc is not None:
        cc.exit_handler(status)
    elif len(logText) > 0:
        try:
            log = ""
            for l in logText:
                log += l + '\n'
            Path(filename).write_text(log.strip('\n'))
        except Exception:
            print(traceback.format_exc())
            if status == 0:
                status = 1
        finally:
            input("Press any key to continue")
            sys.exit(status)
    else:
        input("Press any key to continue")
        sys.exit(status)


def log_print_partial(logText, message, printToLogOnly=False, cc=None):
    if cc is not None:
        cc.log_print(message, printToLogOnly)
    else:
        if(not printToLogOnly):
            print(message)
        logText.append(str(message))

# Gets input using Launchpad or generic method


def get_input_partial(message, default=None, cc=None):
    if not cc:
        return input(message)
    else:
        return cc.get_user_input(message, default)


def AddV(match):
    return "row rowsep=\"0\">%s<entry><para>V%s" % (
        match.group(1) if match.group(1) is not None else '', match.group(2))


def ReplaceRow(match):
    if '/' in match.group(0):
        return "</row>"
    else:
        return "<row rowsep=\"0\">"


def ReplaceEntry(match):
    if '/' in match.group(0):
        return "</para></entry>"
    else:
        return "<entry><para>"


def check_match(match, group_num=1):
    if match is not None:
        return(match.group(group_num))
    else:
        return("")


def get_source():
    source = list(Path('.').glob('*.sgm'))
    if len(source) == 1:
        print("Found source file.")
        return Path(source[0])
    else:
        print(">>>Please input the Source file:")
        source = Path(input())
        while not source.is_file():
            print(">>>File does not exist. Please try again:")
            source = Path(input())
        return source


def EIPCConvert(file, directory, cc):
    log_file_name = "eipc_converter_log.txt"
    if cc is None:
        print("Warning: Launchpad Not Found")
    else:
        cc.log_name = log_file_name

    logText = []

    # Define partials. Makes function calls a little cleaner so cc doesn't need to be passed in each call.
    log_print = partial(log_print_partial, logText, cc=cc)
    # get_input = partial(get_input_partial, cc=cc)
    exit_handler = partial(exit_handler_partial, logText,
                           filename=log_file_name, cc=cc)
    path_name = Path(directory)
    source_file = Path(file)
    skip_figure_items = {}

    def GenerateReferTos(ref):
        nonlocal section
        nonlocal gpd_list
        nonlocal gpd_tags
        nonlocal skip_figure_items
        # If 'ATA' is in the reference, or FIG isn't, referTo's won't support it.
        if "ATA" in ref or "FIG" not in ref or "THRU" in ref:
            gpd_list.append(ref)
            gpd_tags.append("see")
            return ""

        # Get the refer type
        if "NHA" in ref:
            if "ITEM" not in ref:
                ref_type = "rft01"
            else:
                ref_type = "rft03"
        elif "DETAILS" in ref:
            ref_type = "rft02"
        elif "FOR BKDN" in ref:
            ref_type = "rft10"
        elif "REMOVAL" in ref:
            ref_type = "rft06"
        elif "FIELD MAINT" in ref:
            ref_type = "rft51"
        elif "FOR INSTALLATION" in ref:
            ref_type = "rft52"
        elif "PLMB INST" in ref:
            ref_type = "rft53"
        elif "FURTHER BKDN" in ref:
            ref_type = "rft54"
        else:
            gpd_list.append(ref)
            gpd_tags.append("see")
            return ""

        if "THIS SECT" in ref:
            ref_section = section
        else:
            refsect = check_match(pattern_refsect.search(ref), 0)

            # Get the referenced section. If there isn't one, use the current section
            ref_section = refsect if refsect != "" else section

            if len(ref_section) != 8 or ref_section[2] != '-' or ref_section[5] != '-':
                gpd_list.append(ref)
                gpd_tags.append("see")
                return ""

        # Break down the section into components
        ref_section = re.split('-', ref_section)
        ref_sys = ref_section[0]
        ref_subsys = ref_section[1][0]
        ref_subsubsys = ref_section[1][-1]
        ref_assy = ref_section[2]

        # Get the figure and figure variant, if applicable
        ref_figure = pattern_reffig.findall(ref)
        if len(ref_figure) != 1:
            gpd_list.append(ref)
            gpd_tags.append("see")
            return ""

        ref_figure = ref_figure[0]
        if ref_figure is not None:
            figures = []
            if ref_figure[1] != "":
                figures = re.findall(r'\d+\w?', ref_figure[1])
            figures.append(ref_figure[0])
            fig_var_pairs = []
            for figure in figures:
                if figure[-1].isalpha():
                    ref_figvar = figure[-1]
                    figure = figure[0:-1]
                else:
                    ref_figvar = ""
                fig_var_pairs.append((figure, ref_figvar))

        # Get the item and item variant, if applicable
        if "ITEM" in ref:
            ref_item = check_match(pattern_refitem.search(ref))
            if ref_item[-1].isalpha():
                ref_itemvar = ref_item[-1]
                if ref_figure in skip_figure_items:
                    if ref_item[0:-1] in skip_figure_items[ref_figure]:
                        if ord(ref_itemvar) >= ord('I'):
                            ref_itemvar = chr(ord(ref_itemvar) + 1)
                        if ord(ref_itemvar) >= ord('O'):
                            ref_itemvar = chr(ord(ref_itemvar) + 1)
                ref_item = ref_item[0:-1]
            else:
                ref_itemvar = ""
        else:
            ref_item = "001"
            ref_itemvar = ""

        refer_string = ""
        for f in fig_var_pairs:
            refer_string += """\n<referTo refType="%s"><catalogSeqNumberRef assyCode="%s"
    figureNumber="%s%s" %sitem="%s%s" %ssubSubSystemCode="%s" subSystemCode="%s"
    systemCode="%s"></catalogSeqNumberRef></referTo>""" % (ref_type, ref_assy, "0" * (2 - len(f[0])), f[0], "figureNumberVariant=\"%s\" " % f[1] if f[1] != "" else "", "0" * (3 - len(ref_item)), ref_item, "itemVariant=\"%s\" " % ref_itemvar if ref_itemvar != "" else "", ref_subsubsys, ref_subsys, ref_sys)
        return refer_string

    # This could be a generic function in a separate module?
    def GetDmref(name):
        components = re.split('-', name)

        comp_dict = {
            "model": components[1],
            "diff": components[2],
            "sys": components[3],
            "subsys": components[4][0],
            "subsub": components[4][1],
            "assy": components[5],
            "disassy": components[6][0:2],
            "disvar": components[6][2],
            "info": components[7][0:3],
            "infovar": components[7][3],
            "loc": components[8].split('_')[0][0]
        }

        return "\n<dmRef><dmRefIdent><dmCode assyCode=\"%s\" disassyCode=\"%s\" disassyCodeVariant=\"%s\" infoCode=\"%s\" infoCodeVariant=\"%s\" itemLocationCode=\"%s\" modelIdentCode=\"%s\" subSubSystemCode=\"%s\" subSystemCode=\"%s\" systemCode=\"%s\" systemDiffCode=\"%s\"/></dmRefIdent></dmRef>" % (comp_dict["assy"], comp_dict["disassy"], comp_dict["disvar"], comp_dict["info"], comp_dict["infovar"], comp_dict["loc"], comp_dict["model"], comp_dict["subsub"], comp_dict["subsys"], comp_dict["sys"], comp_dict["diff"])

    def GetVendorCodes():
        nonlocal source_text
        file = "DMC-HONAERO-EAB-00-00-00-00A-018Z-C_sx-US.XML"

        text = source_text

        vendata = re.findall(pattern_vendata, text)
        if len(vendata) == 0:
            print("Could Not Locate Data In The Source.")
            return
        table_body = ""
        for v in vendata:
            table_body += v + '\n'
        table_body = re.sub(pattern_mfrcont, "", table_body)
        for r in (
            ("<vendata>", "<row rowsep=\"0\">"),
            ("</vendata>", "</row>"),
            ("<mad>", "<entry><para>"),
            ("<mfr>", "<entry><para>"),
            ("</mad>", "</para></entry>"),
                ("</mfr>", "</para></entry>")):
            table_body = table_body.replace(*r)

        table_body = re.subn(pattern_v, AddV, table_body)[0]

        table = """<table frame="topbot"><tgroup cols="2">
        <?PubTbl tgroup dispwid="700.00px"?>
        <colspec colname="col1" colsep="0" colwidth="0.28*"/><colspec
        colname="col2" colwidth="1.72*"/><thead><row rowsep="0">
        <entry rowsep="1">
        <para>Code</para>
        </entry>
        <entry rowsep="1">
        <para>Vendor</para>
        </entry>
        </row></thead><tbody>
        %s
        </tbody></tgroup></table>""" % (table_body)

        text = (path_name / file).read_text(encoding="utf-8")

        text, replaced = pattern_table.subn(table, text)
        if replaced == 0:
            if "</description>" in text:
                text = text.replace("</description>", table + "</description>")
            else:
                text = text.replace("</mainProcedure>",
                                    table + "</mainProcedure>")

        (path_name / file).write_text(text, encoding="utf-8")
        log_print("\nGenerated Vendor Code Table.")

    def GetServiceBulletin():
        file = "DMC-HONAERO-EAB-00-00-00-00A-008A-D_sx-US.XML"
        nonlocal source_text

        # Read content into a string
        text = source_text

        sblist = re.search(pattern_sblist, text)
        if sblist is None:
            log_print("Could Not Locate Service Bulletin List In Source")
            return

        table_body = ""
        sbdata = pattern_sbdata.findall(sblist.group(0))

        for s in sbdata:
            s = re.sub(r"<(?!\?)/?(?!sbdata)\w+>", ReplaceEntry, s)
            table_body += s + '\n'

        table_body = re.sub(pattern_sbdatatag, ReplaceRow, table_body)

        # for r in (("<sbdata>", "<row rowsep=\"0\">"), ("</sbdata>", "</row>"), ("<sbnbr>","<entry><para>"), ("<sbtitle>","<entry><para>"),  ("</sbnbr>","</para></entry>"), ("</sbtitle>","</para></entry>"), ("<issdate>","<entry><para>"), ("</issdate>","</para></entry>"), ("<ics>","<entry><para>"), ("</ics>","</para></entry>")):
        # table_body = table_body.replace(*r)

        table = """<table id="SB-1" frame="topbot"><tgroup cols="4">
    <?PubTbl tgroup dispwid="700.00px"?>
    <colspec colname="col1" colsep="0" colwidth="25.18*" /><colspec
    colname="col2" colsep="0" colwidth="43.81*" /><colspec colname="col3"
    colsep="0" colwidth="14.56*" /><colspec colname="COLSPEC0"
    colwidth="16.45*" /><thead><row rowsep="1">
    <entry>
    <para>Service Bulletin /</para>
    <para>Revision Number</para>
    </entry>
    <entry>
    <para>Title</para>
    </entry>
    <entry>
    <para>Date Put In</para>
    <para>Manual</para>
    </entry>
    <entry>
    <para>Status</para>
    </entry>
    </row></thead><tbody>
    %s
    </tbody></tgroup></table>""" % (table_body)

        file_path = join(path_name, file)
        f = open(file_path, encoding="utf-8")
        text = f.read()
        f.close()

        text, replaced = pattern_table.subn(table, text)
        if replaced == 0:
            if "</description>" in text:
                text = text.replace("</description>", table + "</description>")
            else:
                text = text.replace("</mainProcedure>",
                                    table + "</mainProcedure>")

        f = open(file_path, "w", encoding="utf-8")
        f.write(text)
        f.close()
        log_print("\nGenerated Service Bulletin Table.")

    def GetTemporaryRevision():
        file = "DMC-HONAERO-EAB-00-00-00-00A-003C-D_sx-US.XML"
        nonlocal source_text

        # Read content into a string
        text = source_text

        trlist = re.search(pattern_trlist, text)
        if trlist is None:
            log_print("Could Not Locate Temporary Revision List In Source")
            return

        table_body = ""
        trdata = pattern_trdata.findall(trlist.group(0))

        for t in trdata:
            table_body += t + '\n'

        for r in (
            ("<trdata>", "<row rowsep=\"0\">"),
            ("</trdata>",
             "\n<entry></entry>\n<entry></entry>\n<entry></entry>\n<entry></entry>\n<entry></entry></row>"),
            ("<trnbr>", "<entry><para>"),
            ("<trstatus>", "<entry><para>"),
            ("</trnbr>", "</para></entry>"),
            ("</trstatus>", "</para></entry>"),
            ("<trloc>", "<entry><para>"),
                ("</trloc>", "</para></entry>")):
            table_body = table_body.replace(*r)

        table = """<table id="TR-1" frame="topbot"><tgroup cols="8">
    <?PubTbl tgroup dispwid="700.00px"?>
    <colspec align="center" colname="col1" colsep="0" colwidth="1.21*"/><colspec colname="col2" colsep="0" colwidth="1.48*"/>
    <colspec align="center" colname="col3" colsep="0" colwidth="0.98*"/><colspec align="center" colname="col4" colsep="0" colwidth="1.30*"/>
    <colspec align="center" colname="col5" colsep="0" colwidth="1.30*"/><colspec colname="col6" colsep="0" colwidth="0.87*"/>
    <colspec colname="col7" colsep="0" colwidth="1.15*"/><colspec colname="col8" colsep="0" colwidth="0.70*"/><thead><row rowsep="1">
    <entry colsep="0" valign="bottom">
    <?PubTbl cell border-left-style="none"?>
    <para>Temporary</para>
    <para>Revision</para>
    <para>Number</para>
    </entry>
    <entry align="center" colsep="0" valign="bottom">
    <para>Status</para>
    </entry>
    <entry colsep="0" valign="bottom">
    <para>Page Number</para>
    </entry>
    <entry colsep="0" valign="bottom">
    <para>Issue Date</para>
    </entry>
    <entry colsep="0" valign="bottom">
    <para>Date Put In</para>
    <para>Manual</para>
    </entry>
    <entry align="center" colsep="0" valign="bottom">
    <para>By</para>
    </entry>
    <entry align="center" colsep="0" valign="bottom">
    <para>Date</para>
    <para>Removed</para>
    <para>From</para>
    <para>Manual</para>
    </entry>
    <entry align="center" valign="bottom">
    <?PubTbl cell border-right-style="none"?>
    <para>By</para>
    </entry>
    </row></thead><tbody>
    %s
    </tbody></tgroup></table>""" % (table_body)

        file_path = join(path_name, file)
        f = open(file_path, encoding="utf-8")
        text = f.read()
        f.close()

        text, replaced = pattern_table.subn(table, text)
        if replaced == 0:
            if "</description>" in text:
                text = text.replace("</description>", table + "</description>")
            else:
                text = text.replace("</mainProcedure>",
                                    table + "</mainProcedure>")

        f = open(file_path, "w", encoding="utf-8")
        f.write(text)
        f.close()
        log_print("\nGenerated Temporary Revision Table.")

    def SubTable(match):
        nonlocal template_file
        nonlocal tables
        nonlocal tableIter
        table = match.group(0)
        if "DMC-HONAERO-EAA-00-00-00-00A-018A-C_sx-US.XML" not in str(template_file):
            # Sub in the header and body so that the table layout/dimensions remain the same
            thead = check_match(pattern_thead.search(tables[tableIter]), 0)
            tbody = check_match(pattern_tbody.search(tables[tableIter]), 0)
            table = re.sub(pattern_thead, thead, table)
            table = re.sub(pattern_tbody, tbody, table)
            tableIter += 1
        else:
            table = tables[tableIter]
            tableIter += 1
        return table.strip('\n')

    def NotePara(match):
        return re.sub('para>', 'notePara>', match.group(0))

    def FootnoteRefRepl(match):
        nonlocal note_ids
        for i, refid in enumerate(note_ids):
            if match.group(1) == refid:
                return "<superScript>%s</superScript>" % str(i+1)
        else:
            return match.group(0)

    def ReadProcessedBluesheet():
        bs = sorted(path_name.glob("20*.xlsx"))
        if len(bs) == 1:
            bs = bs[0]
        elif len(bs) == 0:
            log_print("\nCould not locate processed Bluesheet. Skipping...")
            return
        else:
            log_print("\nLocated multiple processed Bluesheets. Skipping...")
            return

        wb = load_workbook(str(bs), data_only=True)
        ws = wb.active
        list_of_ids = {}
        for row in ws:
            val = row[0].value
            if val is not None:
                if "ICN-" in val and row[7].value is not None:
                    try:
                        old_id = re.search(r'\d{6}', row[7].value).group(0)
                        new_id = re.search(
                            r"0000(\d{6})", row[0].value).group(1)
                        list_of_ids[old_id] = new_id
                    except Exception:
                        log_print(
                            r"Warning: Error occurred while fixing Dup&Mod issue for {}".format(val), True)
                        continue
        if len(list_of_ids) > 0:
            log_print("\nSuccesfully read processed Bluesheet")
            return list_of_ids
        else:
            log_print("\nNo changed entries found in processed Bluesheet")
            return

    def WriteToBluesheet(data, source):
        try:
            shutil.copyfile(str(Path('./launchpad/XML_Conv/BlueSheet Template.xlsx').resolve()), str(path_name / 'BlueSheet.xlsx'))
            wb = load_workbook(str(path_name / 'BlueSheet.xlsx'))
        except FileNotFoundError:
            log_print("\nBluesheet not found... skipping.")
            return
        ws = wb['Blue Sheet']

        for row in data:
            ws.append(row)

        ws = wb.worksheets[0]
        if cc is not None:
            user_info = cc.get_user_info()
            if user_info is not None:
                ws['B1'] = user_info.fname + " " + user_info.lname
                ws['B2'] = user_info.email

        ws['B4'] = check_match(re.search(r'(?s)<eipc[^>]+?docnbr="([^"]+)"', source))
        ws['B5'] = check_match(re.search(r'\b\d{2}-\d{2}-\d{2}\b', source), 0)
        ws['B6'] = re.sub(r'(?s)<acronym>.*?<acronymTerm>([^<]+).*?</acronym>', r'\1', check_match(re.search(r'<partinfo[^>]*>\s+<title>(.*?)</title>', source)))
        ws['B7'] = check_match(re.search(r'(?s)<eipc[^>]+?spl="([^"]+)"', source))
        ws['B8'] = check_match(re.search(r'<!ENTITY ECCN\w+ "([^"]+)">', source)) or check_match(re.search(r'ECCN:?\s?(\w+)', source))
        ws['B9'] = check_match(re.search(r'(?s)<eipc[^>]+?type="([^"]+)"', source)).upper() or "EIPC"

        wb.save(filename=str(path_name / "BlueSheet.xlsx"))

    def GetEntityText(match):
        return check_match(re.search(
            '<!ENTITY %s ([^>]+)>' % (match.group(1)),
            source_text)).strip('"')

    def SubAcro(match):
        content = match.group(0)
        content = re.sub('(acro|abbr)', 'acronym', content)
        content = re.sub('acronymterm', 'acronymTerm', content)
        content = re.sub('acronymname', 'acronymDefinition', content)
        # <abbr><abbrterm>No.</abbrterm> <abbrname>Number</abbrname></abbr>
        return content
    # String where we log all our log_print() messages so we can later write them to log file if desired.

    # Get the path for this script
    template_file = None
    tableIter = 0
    tables = None
    section = None
    gpd_list = []
    gpd_tags = []

    for f in list(Path('./EIPC').glob('*.xml')):
        shutil.copy(f, (path_name / f.name))
    # shutil.copy(Path('./EIPC/BlueSheet.xlsx'), (path_name / 'BlueSheet.xlsx'))
    try:  # Locate PMC
        pmc = sorted(path_name.glob('PMC*.xml'))

        if len(pmc) != 1:
            log_print(
                "Please ensure exactly one 'PMC*.xml' file is present in the directory")
            exit_handler(1)

        pmc = pmc[0]

        pm_text = ""
    except Exception:
        log_print("Failed while locating PMC:")
        log_print(traceback.format_exc())
        exit_handler(1)

    try:  # Get source data for Intro
        source_text = source_file.read_text(encoding='utf-8')
        source_text = re.sub(r'[\r\n]+>', '>', source_text)
        source_text = re.sub(r'<\?Pub Dtl\?>', '', source_text)
        source_text = re.sub(pattern_rev, "", source_text)
        source_text = re.sub(pattern_note, NotePara, source_text)
        source_text = re.sub(
            r'<unlist[^>]*>', r'<para><randomList>', source_text)
        source_text = re.sub(
            r'</unlist>', r'</randomList></para>', source_text)
        source_text = re.sub(r'unlitem>', r'listItem>', source_text)
        source_text = re.sub(
            re.compile(r'<(acro|abbr)>.*?</\1>', re.S),
            SubAcro, source_text)
        cpyrght = check_match(pattern_cpyrght.search(source_text))
        para_data = check_match(pattern_trans.search(source_text))
        tables = pattern_table.findall(source_text)
        partinfo = pattern_partinfo.findall(source_text)
        eipc_info = check_match(pattern_eipc.search(source_text))
        eccn = check_match(pattern_eccn.search(source_text), 0)
        cage = check_match(pattern_mfr.search(source_text)).strip()

        howto = check_match(pattern_howto.search(source_text))
        howto = re.sub(r'<(/)?list\d>', r'', howto)
        howto = re.sub(r'<(l2item)>\n?<para>', r'<\1>\n<title>', howto)
        howto = re.sub(r'(<title>[^<]*)</para>', r'\1</title>', howto)
        howto = re.sub(r'<(/)?l\ditem>\n?', r'<\1levelledPara>\n', howto)

        eccn = re.sub(pattern_entity, GetEntityText, eccn)

        # If we didn't find the copyright years, it might be an entity.
        if cpyrght == "":
            cpyrght_entity = check_match(
                re.search(
                    re.compile(
                        r'<cpyrght>.*?(&.+?;).*?</cpyrght>', re.S),
                    source_text))
            if cpyrght_entity != "":
                cpyrght = re.sub(pattern_entity, GetEntityText, cpyrght_entity)
    except Exception:
        log_print("Fatal Error: Failed to get Intro data from Source:")
        log_print(traceback.format_exc())
        exit_handler(1)

    tables = tables[0:-1]

    for i, table in enumerate(tables):
        tables[i] = re.sub(pattern_rht, "", tables[i])
        # Deal with footnotes in tables. Will add each footnote to a attention sequential list item in the last row of the table, and replace the refids with superscript numbers.
        ftnotes = pattern_ftnote.findall(tables[i])
        if len(ftnotes) > 0:
            cols = re.findall(r'colname="([^"]+)"', table)
            end_col = cols[-1]
            st_col = cols[0]
            note_ids = []
            note_content = "\n<row>\n<entry colsep=\"0\" nameend=\"%s\" namest=\"%s\">\n<?PubTbl cell border-left-style=\"none\" border-right-style=\"none\"?>\n<note>\n<attentionSequentialList>\n" % (
                end_col, st_col)
            for fn in ftnotes:
                note_ids.append(check_match(
                    re.search('ftnoteid="([^"]+)"', fn)))
                note_content += "<attentionSequentialListItem>"
                note_paras = pattern_para.findall(fn)
                for p in note_paras:
                    note_content += "<attentionListItemPara>%s</attentionListItemPara>\n" % p
                note_content += "</attentionSequentialListItem>"
            note_content += "\n</attentionSequentialList></note>\n</entry>\n</row></tbody></tgroup>"
            tables[i] = re.sub(pattern_ftnotesub, note_content, tables[i])
            tables[i] = re.sub(pattern_refint, FootnoteRefRepl, tables[i])
    tableIter = 0

    try:  # 003A
        template_file = path_name / "DMC-HONAERO-EAB-00-00-00-00A-003A-D_sx-US.XML"
        text = template_file.read_text(encoding='utf-8')

        if para_data != "":
            text = re.sub(
                pattern_para, re.sub(
                    pattern_entity, GetEntityText, para_data),
                text, 1)
        else:
            log_print(
                "Could not locate ATA and Revision Date To Fill 003A file. Please fill this in manually.")

        text = re.sub(pattern_table, SubTable, text)

        template_file.write_text(text, encoding="utf-8")
        log_print("\nPopulated 003A.")
    except Exception:
        log_print("Failed to populate the 003A file:")
        log_print(traceback.format_exc())

    try:  # 018A
        template_file = path_name / "DMC-HONAERO-EAA-00-00-00-00A-018A-C_sx-US.XML"
        text = template_file.read_text(encoding='utf-8')

        howto = re.sub(
            pattern_refint,
            r'<internalRef internalRefId="\1" internalRefTargetType="irtt02"></internalRef>',
            howto)
        howto = re.sub(pattern_table, SubTable, howto)

        text = re.sub(
            re.compile(r'<description>.*?</description>', re.S),
            r'<description>%s</description>' % howto, text)
        while True:
            fix1 = re.subn(pattern_fixsteps, r'\2\n\1', text)
            text = fix1[0]
            fix2 = re.subn(pattern_fixsteps2, r'\2\n\1\3', text)
            text = fix2[0]
            fix3 = re.subn(pattern_fixsteps_table, r'\2\n\1', text)
            text = fix3[0]
            if fix1[1] + fix2[1] + fix3[1] == 0:
                break

        template_file.write_text(text, encoding="utf-8")
        log_print("\nPopulated 018A.")
    except Exception:
        log_print("Failed to populate the 018A file:")
        traceback.print_exc()

    try:  # 018D
        template_file = path_name / "DMC-HONAERO-EAA-00-00-00-00A-018D-C_sx-US.XML"
        text = template_file.read_text(encoding='utf-8')

        text = re.sub(pattern_table, SubTable, text)

        template_file.write_text(text, encoding="utf-8")
        log_print("\nPopulated 018D.")
    except Exception:
        log_print("Failed to populate the 018D file:")
        log_print(traceback.format_exc())

    try:  # 018F
        template_file = path_name / "DMC-HONAERO-EAA-00-00-00-00A-018F-C_sx-US.XML"
        text = template_file.read_text(encoding='utf-8')

        text = re.sub(pattern_table, SubTable, text)

        template_file.write_text(text, encoding="utf-8")
        log_print("\nPopulated 018F.")
    except Exception:
        log_print("Failed to populate the 018F file:")
        log_print(traceback.format_exc())

    try:  # 021A
        if cpyrght == "":
            log_print(
                "No copyright data found. Please fill out the copyright years in 021A manually.")
        else:
            template_file = path_name / "DMC-HONAERO-EAB-00-00-00-00A-021A-D_sx-US.XML"
            text = template_file.read_text(encoding='utf-8')

            text = re.sub(pattern_years, cpyrght, text)

            template_file.write_text(text, encoding="utf-8")
            log_print("\nPopulated 021A.")
    except Exception:
        log_print("Failed to populate the 021A file:")
        log_print(traceback.format_exc())

    pmTitle = None

    try:  # 00PA
        template_file = path_name / "DMC-HONAERO-EAB-00-00-00-01A-00PA-C_sx-US.XML"
        text = template_file.read_text(encoding='utf-8')
        partinfo_text = "<productCrossRefTable>"
        for part in partinfo:
            # pmTitle is in the first partinfo. Need to make sure we don't overwrite it when the second match comes up empty
            title = re.sub(
                pattern_entity, GetEntityText,
                check_match(pattern_booktitle.search(part)))
            if title != "":
                pmTitle = title

            model = check_match(pattern_model.search(part))
            mfrpnr = pattern_mfrpnr.findall(part)
            for p in mfrpnr:
                mfr = check_match(pattern_mfr.search(p))
                pnr = check_match(pattern_pnr.search(p))
                partinfo_text += """\n<product>\n
    <assign applicPropertyIdent="PN" applicPropertyType="prodattr" applicPropertyValue="%s"/>\n
    <assign applicPropertyIdent="Model" applicPropertyType="prodattr" applicPropertyValue=%s/>\n
    <assign applicPropertyIdent="cage" applicPropertyType="prodattr" applicPropertyValue="%s"/>\n</product>""" % (pnr, model, mfr)
        partinfo_text += "\n</productCrossRefTable>"
        # iterate through parts
        text = re.sub(pattern_prodxref, partinfo_text, text)

        template_file.write_text(text, encoding="utf-8")
        log_print("\nPopulated 00PA.")
    except Exception:
        log_print("Failed to populate the 00PA file:")
        log_print(traceback.format_exc())

    issueDate = None
    issueNumber = None

    try:  # PMC
        if eipc_info != "":
            text = pmc.read_text(encoding='utf-8')

            pmTitle = '<pmTitle>%s</pmTitle>' % pmTitle.title()
            text = re.sub(pattern_pmtitle, pmTitle, text)

            shortPmTitle = check_match(pattern_shrtpmtitle.search(eipc_info))
            shortPmTitle = '<shortPmTitle>%s </shortPmTitle>' % shortPmTitle
            text = re.sub(pattern_shorttitle, shortPmTitle, text)

            initialDate = check_match(pattern_oidate.search(eipc_info))
            initialDate = '<externalPubCode pubCodingScheme="initialDate">%s %s %s</externalPubCode>' % (
                initialDate[6:], months[int(initialDate[4:6])], initialDate[0:4])
            text = re.sub(pattern_inidate, initialDate, text)

            docnbr = check_match(pattern_docnbr.search(eipc_info))
            docnbr_xml = '<externalPubCode pubCodingScheme="INT">%s</externalPubCode>' % docnbr
            text = re.sub(pattern_int, docnbr_xml, text)

            issueNumber = check_match(pattern_tsn.search(eipc_info))
            issueNumber = 'issueNumber="%s"' % (
                "0" * (3 - len(str(issueNumber))) + str(issueNumber))
            text = re.sub(pattern_issnum, issueNumber, text)

            text = re.sub(r'pmIssuer="\d{5}"', 'pmIssuer="%s"' % cage, text)

            issueDate = check_match(pattern_revdate.search(eipc_info))
            if len(issueDate) >= 8:
                issueDate = '<issueDate day="%s" month="%s" year="%s"/>' % (
                    issueDate[6:], issueDate[4:6], issueDate[0:4])
                text = re.sub(pattern_issdate, issueDate, text)

            if eccn != "":
                text = re.sub(pattern_eccn, eccn, text)

            text = re.sub(
                pattern_authdoc, 'authorityDocument="%s" pmEntryType="pmt56"' %
                docnbr, text)

            pmc.write_text(text, encoding="utf-8")
            log_print("\nPopulated PMC Header Data.")
        else:
            log_print("Failed to get PMC header data.")
    except Exception:
        log_print("Failed while filling out PMC data:")
        log_print(traceback.format_exc())

    try:  # Record of Temporary Revisions
        GetTemporaryRevision()
    except Exception:
        log_print("Failed to convert Record Of Temporary Revisions table:")
        log_print(traceback.format_exc())

    try:  # Service Bulletin List
        GetServiceBulletin()
    except Exception:
        log_print("Failed to convert Service Bulletin List:")
        log_print(traceback.format_exc())

    try:  # Vendor Code List
        GetVendorCodes()
    except Exception:
        log_print("Failed to convert Vendor Code List:")
        log_print(traceback.format_exc())

    changed_ids = {}
    try:
        changed_ids = ReadProcessedBluesheet()
    except Exception:
        log_print("Failed to read processed Bluesheet:")
        log_print(traceback.format_exc())

    try:  # Generate IPL
        bluesheet_data = []
        text = re.sub(r'<\?Pub[^>]*>', '', source_text)
        book_title = re.sub(
            pattern_entity, GetEntityText,
            check_match(pattern_booktitle.search(text))).upper()

        listOfChapters = pattern_chapt.findall(text)
        # list_of_sections = {}
        if len(listOfChapters) == 0:
            raise Exception('No Chapters Found!')
            exit_handler(1)

        for cpt in listOfChapters:
            listOfSections = pattern_sect.findall(cpt[2])
            chapt = pattern_chapnbr.search(cpt[0]).group(1)
            pm_text += "\n<pmEntry authorityDocument=\"%s\" pmEntryType=\"pmt75\">\n<pmEntryTitle>%s</pmEntryTitle>" % (
                chapt, cpt[1])
            for sct in listOfSections:
                listOfUnits = pattern_unit.findall(sct[2])
                sect = pattern_sectnbr.search(sct[0]).group(1)
                pm_text += "\n<pmEntry authorityDocument=\"%s-%s\">\n<pmEntryTitle>%s</pmEntryTitle>" % (
                    chapt, sect, sct[1])
                for unt in listOfUnits:
                    listOfFigures = pattern_fig.findall(unt[2])
                    sect = pattern_sectnbr.search(unt[0]).group(1)
                    unit = pattern_unitnbr.search(unt[0]).group(1)
                    pm_text += "\n<pmEntry authorityDocument=\"%s-%s-%s\">\n<pmEntryTitle>%s</pmEntryTitle>\n<pmEntry>" % (
                        chapt, sect, unit, unt[1])
                    disassyCode = 0
                    firstEntry = True
                    for k, fig in enumerate(listOfFigures):
                        figtext = ""
                        entities = []

                        # Get fignbr
                        fignbr = pattern_fignbr.search(fig[0]).group(1)

                        # Data for each graphic, including key, title and the ICN and page size for each sheet
                        graphic = check_match(pattern_graphic.search(fig[1]))
                        if graphic != "":
                            key = pattern_key.search(fig[0]).group(1)
                            title = pattern_title.search(fig[1]).group(1)
                            figure_text = "<figure id=\"%s\">\n<title>%s</title>\n" % (
                                key, title)
                            sheets = pattern_sheet.findall(graphic)

                            # Generate the graphic ICN, and add foldout dimensions if required
                            for i, sheet in enumerate(sheets):
                                gnbr = pattern_gnbr.search(sheet).group(1)
                                if "ID" in gnbr:
                                    gnbr = re.split('-', gnbr)[1]
                                    if changed_ids is not None:
                                        if gnbr in changed_ids:
                                            gnbr = changed_ids[gnbr]
                                    icn = "ICN-%s-0000%s-001-01" % (cage, gnbr)
                                elif "ICN" in gnbr:
                                    icn = gnbr
                                else:
                                    log_print(
                                        "Warning: %s not a recognized image ID format." % gnbr)

                                for b in bluesheet_data:
                                    if gnbr == b[1]:
                                        break
                                else:
                                    bluesheet_data.append(
                                        ["I", gnbr, "", "", fignbr +
                                         ((",SH%s" % (i + 1))
                                          if len(sheets) > 1 else ""),
                                         title, "A"])

                                entities.append(icn)
                                foldout = True if check_match(
                                    pattern_imgarea.search(sheet)).lower() == "hl" else False
                                figure_text += "<graphic infoEntityIdent=\"%s\"%s></graphic>\n" % (
                                    icn, " reproductionHeight=\"209.55 mm\" reproductionWidth=\"355.6 mm\"" if foldout else "")
                            figure_text += "</figure>"

                            # Generate entity list to add to file header
                            ent_text = ""
                            for e in entities:
                                ent_text += "\n<!ENTITY %s SYSTEM \"Graphics/%s.cgm\" NDATA cgm>" % (
                                    e, e)
                        else:
                            figure_text = '<figure changeType="delete">\n<title>Deleted</title>\n<graphic></graphic>\n</figure>'
                        # Header data
                        disassyCode += 1
                        if disassyCode == 100:
                            disassyCode = 1

                        disassyCodeVariant = chr(ord('A') + ((k + 1) // 100))
                        figtext += """<?xml version="1.0" encoding="UTF-8"?>
    <!--Arbortext, Inc., 1988-2014, v.4002-->
    <!DOCTYPE dmodule [%s
    <!ENTITY %% ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
    %%ISOEntities;
    <!NOTATION cgm SYSTEM "cgm">
    ]>
    <?Pub Inc?>
    <dmodule xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/ipd.xsd">
    <?MENTORPATH ?>
    <identAndStatusSection>
    <dmAddress>
    <dmIdent>
    <dmCode assyCode="%s" disassyCode="%s%s" disassyCodeVariant="%s"
    infoCode="941" infoCodeVariant="A" itemLocationCode="C"
    modelIdentCode="HON%s" subSubSystemCode="%s" subSystemCode="%s"
    systemCode="%s" systemDiffCode="EAA"/>
    <language countryIsoCode="US" languageIsoCode="sx"/>
    <issueInfo inWork="00" issueNumber="001"/></dmIdent>
    <dmAddressItems>
    <issueDate day="01" month="01" year="1979"/>
    <dmTitle><techName>%s</techName>
    <infoName>%s</infoName>
    </dmTitle>
    </dmAddressItems></dmAddress>
    <dmStatus issueType="new">
    <security securityClassification="01"/>
    <responsiblePartnerCompany enterpriseCode="%s">
    </responsiblePartnerCompany>
    <originator enterpriseCode="%s"></originator>
    <applic>
    <displayText>
    <simplePara>ALL</simplePara>
    </displayText>
    </applic>
    <brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00"
    disassyCodeVariant="AA" infoCode="022" infoCodeVariant="A"
    itemLocationCode="D" modelIdentCode="S1000DBIKE" subSubSystemCode="0"
    subSystemCode="0" systemCode="00" systemDiffCode="AAA"/></dmRefIdent>
    </dmRef></brexDmRef>
    <qualityAssurance>
    <unverified/></qualityAssurance>
    </dmStatus>
    </identAndStatusSection>
    <content>
    <illustratedPartsCatalog>""" % (ent_text, unit, "0" if disassyCode < 10 else "", disassyCode, disassyCodeVariant, cage, sect[-1], sect[0], chapt, unt[1], book_title, cage, cage)
                        figtext += figure_text

                        section = "%s-%s-%s" % (chapt, sect, unit)
                        module_name = 'DMC-HON%s-EAA-%s-%s%s-%s-%s%s%s-941A-C_sx-US.xml' % (
                            cage, chapt, sect[0], sect[-1], unit, "0" if disassyCode < 10 else "", disassyCode, disassyCodeVariant)
                        dmref = GetDmref(module_name)

                        listOfItemData = pattern_itemdata.findall(fig[1])
                        itemvar_skip = 0
                        skip_item = None
                        for i in listOfItemData:
                            illusind = check_match(
                                pattern_illusind.search(i[0]))
                            illustrated = " - " if illusind == "0" else "   "

                            attach = check_match(pattern_attach.search(i[0]))

                            # tags: dd, kwd, adt, upa, pnr, effcode, csdmfr, optmfr, sbs, rp, rps
                            tags = pattern_tags.findall(i[1])
                            use_list = []
                            dd = ""
                            refto = []
                            partNumberValue = ""
                            partKeyword = ""
                            quantity = ""
                            shortName = ""
                            gpd_tags = []
                            gpd_list = []
                            mfr = ""
                            deleted = False
                            for tag in tags:
                                if tag[1] == "part":
                                    mp = check_match(
                                        re.search(r'mp="([^"]+)"', tag[0]))
                                    if(mp != ""):
                                        gpd_list.append(mp)
                                        gpd_tags.append("mp")
                                    psTags = pattern_tags.findall(tag[2])
                                    for ptag in psTags:
                                        if ptag[1] == "kwd":
                                            partKeyword = ptag[2].replace(
                                                '\n', ' ')
                                            partKeyword = partKeyword.strip()
                                        elif ptag[1] == "pnr":
                                            partNumberValue = ptag[2]
                                        elif ptag[1] == "adt":
                                            shortName = ptag[2]
                                            if shortName != "" and partKeyword[-1] != '-':
                                                partKeyword += '-'
                                        elif ptag[1] == "dd":
                                            dd = ptag[2]
                                        elif ptag[1] == "csd":
                                            gpd_list.append(
                                                "(CSD %s)" % ptag[2])
                                            gpd_tags.append("csdmfr")
                                        elif ptag[1] == "opt":
                                            gpd_list.append(
                                                "(OPT MFR %s)" % ptag[2])
                                            gpd_tags.append(ptag[1])
                                        elif ptag[1] == "mfr":
                                            mfr = ptag[2]
                                        elif ptag[1] == "venptnbr":
                                            gpd_list.append(ptag[2])
                                            gpd_tags.append(ptag[1])
                                        else:
                                            gpd_list.append(ptag[2])
                                            gpd_tags.append(ptag[1])
                                elif tag[1] == "itemspec":
                                    isTags = pattern_tags.findall(tag[2])
                                    for itag in isTags:
                                        if itag[1] == "effcode":
                                            use_list.append(itag[2])
                                        elif itag[1] == "upa":
                                            quantity = itag[2]
                                        elif itag[1] == "msc":
                                            para = pattern_tags.search(
                                                itag[2]).group(3)
                                            if "SEE" in para:
                                                refto.append(
                                                    para.replace('.', ''))
                                            elif "MISSING NOTE" in para:
                                                continue
                                            else:
                                                gpd_list.append(para)
                                                if "REPLACES" in para:
                                                    gpd_tags.append("rps")
                                                else:
                                                    gpd_tags.append(itag[1])
                                        elif itag[1] == "rp":
                                            gpd_list.append(
                                                "REPLACED BY ITEM %s" % itag[2])
                                            gpd_tags.append("replacedBy")
                                        elif itag[1] == "rps":
                                            gpd_list.append(
                                                "REPLACES ITEM %s" % itag[2])
                                            gpd_tags.append("replaces")
                                        elif itag[1] == "sb" or itag[1] == "sbnbr":
                                            sb_match = pattern_sb.search(
                                                itag
                                                [0])
                                            if sb_match is None:
                                                continue
                                            sb_type = sb_match.group(1)
                                            sb_nbr = sb_match.group(2)
                                            sb = "(%s SB %s)" % (
                                                sb_type, sb_nbr)
                                            gpd_list.append(sb)
                                            gpd_tags.append("serviceBulletin")
                                        elif itag[1] == "fnha" or itag[1] == "fds":
                                            refto.append(
                                                itag[2].replace('.', ''))
                                        elif itag[1] == "apn":
                                            apn = "ALTN PART FOR ITEM %s" % itag[2]
                                            gpd_list.append(apn)
                                            gpd_tags.append("msc")
                                        elif itag[1] == "uwp":
                                            gpd_list.append(
                                                "USED WITH PART %s" % itag[2])
                                            gpd_tags.append("usedWithPart")
                                        elif itag[1] == "uoa":
                                            gpd_list.append(
                                                "USED ON ASSY: %s" % itag[2])
                                            gpd_tags.append("usedOnAssy")
                                        elif itag[1] == "sym":
                                            side = check_match(
                                                re.search(
                                                    r'side="([^"]+)"', itag
                                                    [0]))
                                            if side == "LR":
                                                gpd_list.append(
                                                    "LEFT AND RIGHT-HAND ENGINES")
                                            elif side == "LH":
                                                gpd_list.append(
                                                    "LEFT-HAND ENGINE")
                                            elif side == "RH":
                                                gpd_list.append(
                                                    "RIGHT-HAND ENGINE")
                                            else:
                                                log_print(
                                                    "Unrecognized value %s in attribute 'side' in <sym/> tag" % side)
                                                continue
                                            gpd_tags.append(itag[1])
                                        else:
                                            if "SEE" in itag[2]:
                                                refto.append(
                                                    itag[2].replace('.', ''))
                                            else:
                                                gpd_list.append(itag[2])
                                                gpd_tags.append(itag[1])

                            item = pattern_itemnbr.search(i[0]).group(1)
                            indenture = int(
                                pattern_indent.search(i[0]).group(1)) * '.'

                            chg = pattern_chg.search(i[0]).group(1)
                            deleted = chg == "D" or partKeyword == "DELETED"
                            if deleted:
                                if partKeyword != "DELETED":
                                    partKeyword = "DELETED"
                                shortName = ""
                                mfr = ""

                            if len(use_list) > 0:
                                if len(use_list) == 1:
                                    use = use_list[0]
                                else:
                                    use = ""
                                    for c, u in enumerate(use_list):
                                        use += u
                                        if c == len(use_list) - 1:
                                            break
                                        use += ','
                            else:
                                use = " "

                            i_refs = ""
                            for ref in refto:
                                i_refs += GenerateReferTos(ref)

                            if len(gpd_list) > 0 or dd != "":
                                i_gpd = "\n<genericPartDataGroup>"
                                if dd != "":
                                    i_gpd += "\n<genericPartData genericPartDataName=\"dd\"><genericPartDataValue>%s</genericPartDataValue></genericPartData>" % (
                                        dd)
                                for t, d in enumerate(gpd_list):
                                    i_gpd += "\n<genericPartData genericPartDataName=\"%s\"><genericPartDataValue>%s</genericPartDataValue></genericPartData>" % (
                                        gpd_tags[t], d)
                                i_gpd += "\n</genericPartDataGroup>"
                            else:
                                i_gpd = ""

                            # All the values for each item to be converted
                            i_delete = "changeType=\"delete\" " if deleted and partNumberValue != "" else ""
                            i_fignbr = fignbr if fignbr[-1].isnumeric(
                            ) else fignbr[0:-1]
                            i_fignbrzero = "0" * (2 - len(i_fignbr))
                            i_figvar = "figureNumberVariant=\"%s\" " % fignbr[-1] if fignbr[-1].isalpha(
                            ) else ""
                            i_key = pattern_key.search(i[0]).group(1)
                            i_itemzero = "0" * (3 - len(item)
                                                ) if item[-1].isnumeric() else "0" * (4 - len(item))
                            i_item = item if item[-1].isnumeric() else item[0:-1]
                            i_itemvar = item[-1] if item[-1].isalpha() else ""

                            if i_itemvar == "I":
                                skip_item = i_item
                                itemvar_skip = 1
                                if fignbr not in skip_figure_items:
                                    skip_figure_items[fignbr] = [skip_item]
                                else:
                                    skip_figure_items[fignbr].append(skip_item)

                            if (itemvar_skip == 1 and i_itemvar == "N") or (itemvar_skip == 0 and i_itemvar == "O"):
                                itemvar_skip += 1

                            if skip_item == i_item:
                                i_itemvar = chr(ord(i_itemvar) + itemvar_skip)
                            else:
                                itemvar_skip = 0
                            # if i_itemvar != "":
                                # if ord(i_itemvar) >= ord('I'):
                                # i_itemvar = chr(ord(i_itemvar) + 1)
                                # if ord(i_itemvar) >= ord('O'):
                                # i_itemvar = chr(ord(i_itemvar) + 1)
                            i_mfr = mfr if mfr != "" else cage
                            i_attach = "\n<attachStoreShipPart attachStoreShipPartCode=\"1\"/>" if attach == "1" else ""
                            i_illust = "\n<notIllustrated></notIllustrated>" if illustrated == " - " else ""
                            i_use = "\n<applicabilitySegment><usableOnCodeEquip>%s</usableOnCodeEquip></applicabilitySegment>" % use if use != " " else ""
                            i_desc = "%s%s%s" % (
                                partKeyword, '-'
                                if partKeyword[-1] != '-' else "", shortName)
                            i_kwd = "%s%s" % (
                                partKeyword, '-'
                                if shortName != "" and partKeyword[-1] != '-' else
                                "")
                            i_adt = shortName

                            # Generate and add a new catalogSeqNumber to the IPL file
                            ipl_item = """\n<catalogSeqNumber assyCode="%s" %sfigureNumber="%s%s"
    %sid="%s" indenture="%s" item="%s%s"
    subSubSystemCode="%s" subSystemCode="%s" systemCode="%s">
    <itemSeqNumber itemSeqNumberValue="00%s">
    <quantityPerNextHigherAssy>%s</quantityPerNextHigherAssy>
    <partRef manufacturerCodeValue="%s" partNumberValue="%s"></partRef>
    <partSegment>
    <itemIdentData><descrForPart>%s</descrForPart>
    <partKeyword>%s</partKeyword>
    <shortName>%s</shortName>
    </itemIdentData></partSegment>
    <partLocationSegment>%s%s%s</partLocationSegment>%s%s
    </itemSeqNumber></catalogSeqNumber>""" % (unit, i_delete, i_fignbrzero, i_fignbr, i_figvar, i_key, len(
                                indenture) + 1, i_itemzero, i_item, sect[-1], sect[0], chapt, i_itemvar, quantity, i_mfr, partNumberValue, i_desc, i_kwd, i_adt, i_attach, i_illust, i_refs, i_use, i_gpd)
                            ipl_item = re.sub(
                                r"<partLocationSegment></partLocationSegment>\n",
                                '', ipl_item)
                            figtext += '\n' + ipl_item

                        # For a deleted figure, we need to add a 'blank' CSN
                        if graphic == "":
                            i_fignbr = fignbr if fignbr[-1].isnumeric(
                            ) else fignbr[0:-1]
                            i_fignbrzero = "0" * (2 - len(i_fignbr))
                            i_figvar = "figureNumberVariant=\"%s\" " % fignbr[-1] if fignbr[-1].isalpha(
                            ) else ""
                            figtext += '\n' + """<catalogSeqNumber assyCode="%s" changeType="delete" figureNumber="%s%s" %s indenture="0"
    item="000" subSubSystemCode="%s" subSystemCode="%s" systemCode="%s">
    <itemSeqNumber itemSeqNumberValue="00">
    <quantityPerNextHigherAssy>0</quantityPerNextHigherAssy>
    <partRef manufacturerCodeValue="%s" partNumberValue=""></partRef>
    <partSegment>
    <itemIdentData><descrForPart>DELETED-</descrForPart>
    <partKeyword>DELETED</partKeyword>
    <shortName></shortName>
    </itemIdentData></partSegment>
    <partLocationSegment>
    <notIllustrated></notIllustrated></partLocationSegment>
    </itemSeqNumber></catalogSeqNumber>""" % (unit, i_fignbrzero, i_fignbr, i_figvar, sect[-1], sect[0], chapt, cage)
                        # Close off tags and write the file
                        figtext += "\n</illustratedPartsCatalog>\n</content>\n</dmodule>"
                        (path_name / module_name).write_text(figtext, encoding='utf-8')
                        isVariant = False if i_figvar == "" else True
                        pm_text += "%s%s" % ("\n</pmEntry>\n<pmEntry>"
                                             if not firstEntry and not isVariant else
                                             "", dmref)
                        if firstEntry:
                            firstEntry = False

                    pm_text += "</pmEntry>\n</pmEntry>"
                pm_text += "\n</pmEntry>"
            pm_text += "\n</pmEntry>"
        log_print("\nGenerated IPL Files")
    except Exception:
        log_print("Failed while generating data modules:")
        log_print(traceback.format_exc())

    try:  # Print to Bluesheet
        WriteToBluesheet(bluesheet_data, source_text)
    except Exception:
        log_print("Failed while printing to Bluesheet:")
        log_print(traceback.format_exc())

    try:  # Print PMC
        text = pmc.read_text(encoding="utf-8")
        if pm_text != "":
            report = re.subn(pattern_dpl2, r"\1%s" % pm_text.strip('\n'), text)
            if report[1] == 0:
                log_print(
                    "Failed to print to PMC; Detailed Parts List not found")
                log_print(
                    "Please ensure the Vendor Code List has the proper pmt code (pmt90)")
            else:
                text = report[0]
                pmc.write_text(text, encoding="utf-8")
                log_print("\nAdded IPL modules to PMC.")
        else:
            log_print("\nNo IPL Modules To Add")
    except Exception:
        log_print("Failed to print new 941 files to PMC:")
        log_print(traceback.format_exc())

    # Get a list of all DMCs in the directory
    files = sorted(path_name.glob('DMC*.xml'))

    if issueDate is not None and issueNumber is not None and pmTitle is not None:
        for file in files:  # Update issueDate, issueNumber, techName
            try:
                text = file.read_text(encoding="utf-8")

                text = re.sub(pattern_issdate, issueDate, text)
                text = re.sub(pattern_issnum, issueNumber, text)
                techName = pmTitle.replace('pmTitle', 'techName')
                text = re.sub(pattern_techname, techName, text)

                file.write_text(text, encoding="utf-8")

            except Exception:
                log_print("Failed to update Issue info in %s:" % file)
                log_print(traceback.format_exc())
        else:
            log_print("\nPopulated data module header info.")
    else:
        log_print(
            "\nFailed to update DMC header info; issueDate, issueNumber, and/or pmTitle not defined.")

    log_print("\nDone!")

    exit_handler()


if __name__ == "__main__":
    EIPCConvert(get_source(), Path('.'), None)
