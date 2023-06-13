import re
import sys
import traceback
from functools import partial
from pathlib import Path

from openpyxl import load_workbook
from collections import defaultdict
try:
    from ._IPL_Converter import DoTheThing
except ImportError:
    from _IPL_Converter import DoTheThing

import shutil

# try:
#     from launchpad.functions import (
#         ConversionClient, exit_handler_partial,
#         log_print_partial, get_input_partial)
# except ImportError:
#     pass

# Source processing
p_ftnote = re.compile(r'<ftnote.*?</ftnote>', re.S)
p_ftnotesub = re.compile(r'</tbody>.*?</ftnote>(?=\s?</table>)', re.S)
p_techname = re.compile(r'<techName>[^<]+</techName>')
p_note = re.compile(r'<note>.*?</note>', re.S)
p_entity = re.compile(r'&(.+?);', re.S)
p_rev = re.compile(r"<\?Pub\s?\n?/?_rev\n?\?>\n?")
p_entityref = re.compile(r'<!ENTITY ([^\s]+) "([^"]+)">')
p_grphcref = re.compile(r'<grphcref.*?refid="([^"]+).*?</grphcref>', re.S)
p_refint = re.compile(r'<refint.*?refid="([^"]+)".*?</refint>(\.?)', re.S)
p_rht = re.compile(r'<\?PubTbl row rht[^>]+>\n?')
p_para = re.compile(r"<para>(.*?)</para>", re.S)
p_table = re.compile(r'<table(.*?)</table>', re.S)
p_entry = re.compile(r'(<entry[^>]*>(.*?)</entry>)', re.S)
p_fixsteps_tabfig = re.compile(r'((?:</(?:proceduralStep|levelledPara)>\n?)*'
                               r'</(?:proceduralStep|levelledPara)>)\n*'
                               r'(<(table|figure).*?</\3>)', re.S)
p_fixsteps_list = re.compile(r'((?:</(?:proceduralStep|levelledPara)>\n?)*'
                             r'</(?:proceduralStep|levelledPara)>)\n*'
                             r'(<para><(?:random|sequential)List>.*?'
                             r'</(?:random|sequential)></para>)', re.S)
p_fixsteps = re.compile(r'((?:</(?:proceduralStep|levelledPara)>\n?)*)<(?:proceduralStep|levelledPara)>\n*(<(?:table|para><(?:random|sequential)List>|figure).*?)</(?:proceduralStep|levelledPara)>', re.S)
p_equ = re.compile(r'<equ>(.*?)</equ></para>', re.S)
p_colspec = re.compile(r'(<colspec[^/>]*>\s?)(?!</colspec>)')
p_subtitle = re.compile(r'<prcitem>\n?<title>(.*?)</title>', re.S)
p_graphic = re.compile(r'<graphic(.*?<title>(.*?)</title>.*?)</graphic>', re.S)
p_fig = re.compile(r'<figure(.*?<graphic.*?<title>(.*?)</title>.*?)</graphic>', re.S)
p_sheet = re.compile(r'<sheet.*?gnbr="([^"]+)', re.S)
p_sub_super = re.compile(r'<(sub|super)>([^<]+)</\1>')
p_sgm_attr = re.compile(r'(?:chapnbr|sectnbr|subjnbr|func|seq|confltr|varnbr|confnbr|pgblknbr|scalefit|sheetnbr|reprowid|reprodep)="[^"]+"\s?')
p_split_graphic = re.compile(r'(</table>\n?</gdesc>\n?</sheet>\n?)(<sheet[^>]*key="([^"]*)"[^>]*>(?!(?:</table>\n?)?</gdesc>).*?)', re.S)
p_title = re.compile(r'<title>(.*?)</title>', re.S)
p_id = re.compile(r'id="([^"]*)"', re.S)

# For Tools and Cons tables
p_tgroup = re.compile(r'<tgroup.*?</tgroup>', re.S)
p_tool2 = re.compile(r'<(std|ted)>(.*?)</\1>', re.S)
p_cons2 = re.compile(r'<con>(.*?)</con>', re.S)

# SBL, RTR, VCL TOI, Prelim
p_proptary = re.compile(r'<fullstmt>(.*?)</fullstmt>', re.S)
p_vendata = re.compile(r"(<vendata>.*?</vendata>)", re.S)
p_v = re.compile(r"row rowsep=\"0\">(\n)?<entry><para>([^V])")
p_mfrcont = re.compile(r"</?mfrcont>")
p_sblist = re.compile(r"<sblist.*?</sblist>", re.S)
p_sbdata = re.compile(r"(<sbdata>.*?</sbdata>)", re.S)
p_sbdatatag = re.compile(r"</?sbdata>")
p_trlist = re.compile(r"<trlist.*?</trlist>", re.S)
p_trdata = re.compile(r"<trdata>.*?</trdata>", re.S)
p_transltr = re.compile(r'<transltr[^>]*>.*?<title>[^<]*</title>\n?(.*?)</transltr>', re.S)

# Book metadata
p_cage = re.compile(r'spl="(\w{5})"')
p_tsn = re.compile(r'tsn="(\d+)"')
p_oidate = re.compile(r'oidate="(\d{8})"')
p_revdate = re.compile(r'revdate="(\d{8})"')
p_eccn = re.compile(r'ECCN:[^<]+')
p_copyright = re.compile(r'<geninfo>.*?((?:\d{4}, )\d{4})', re.S)
p_docnbr = re.compile(r'docnbr="([^"]+)')
p_model = re.compile(r'model="([^"]+)')
p_cmpnom = re.compile(r'<cmpnom>(.*?)</cmpnom>', re.S)

# PM Structure
p_pgblk = re.compile(r'<(pgblk|ipl)([^>]+)>\n?(?:<effect>(.*?)</effect>)?\n?(?:<title>(.*?)</title>)?(.*?)</\1>', re.S)
p_tasks = re.compile(r'<task(?:.*?func="(\w+)")?.*?key="([^"]+)"[^>]*>.*?<title>(.*?)</title>(.*?)</task>', re.S)
p_subtask = re.compile(r'<subtask.*?func="(\w+)".*?key="([^"]+)"[^>]*>(.*?)</subtask>', re.S)
p_nonpgblk = re.compile(r'<(/?)(?:trlist|sblist|transltr)', re.S)
p_vendlist_tag = re.compile(r'<(/?)vendlist', re.S)

# Repository stuff
p_tool = re.compile(r'<(std|ted)>.*?</\1>', re.S)
p_cons = re.compile(r'<con>.*?</con>', re.S)
p_cautwarn = re.compile(r'()<((?:proceduralStep|levelledPara)[^>]*)>\n*<(?:caution|warning)>.*?</(?:caution|warning)>(?!\n*<(?:warning|caution)>)\n*', re.S)
p_cautwarn2 = re.compile(r'(?<=</title>)\n*<(?:caution|warning)>.*?</(?:caution|warning)>\n?(<note>.*?</note>\n?)?<((?:proceduralStep|levelledPara)[^>]*)>', re.S)

# 00P
p_partinfo = re.compile(r'<partinfo[^>]+>(.*?)</partinfo>', re.S)
p_mfrpnr = re.compile(r'<mfrpnr>(.*?)</mfrpnr>', re.S)
p_mfr = re.compile(r'<mfr>(.*?)</mfr>')
p_pnr = re.compile(r'<pnr>(.*?)</pnr>')

months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

page_blocks = {
    "0": ("INTRODUCTION", "pmt58", "018", "D"),
    "1": ("DESCRIPTION AND OPERATION", "pmt59", "040", "D"),
    "1000": ("TESTING AND FAULT ISOLATION", "pmt60", "400", "P"),
    "2000": ("SCHEMATICS AND WIRING DIAGRAMS", "pmt61", "050", "D"),
    "3000": ("DISASSEMBLY", "pmt62", "500", "P"),
    "4000": ("CLEANING", "pmt63", "250", "P"),
    "5000": ("INSPECTION/CHECK", "pmt64", "300", "P"),
    "6000": ("REPAIR", "pmt65", "600", "P"),
    "7000": ("ASSEMBLY", "pmt66", "700", "P"),
    "8000": ("FITS AND CLEARANCES", "pmt67", "711", "P"),
    "9000": ("SPECIAL TOOLS, FIXTURES, EQUIPMENT, AND CONSUMABLES",
             "pmt68", "900", "D"),
    "10000": ("ILLUSTRATED PARTS LIST", "pmt75", "018", "D"),
    "11000": ("SPECIAL PROCEDURES", "pmt69", "900", "P"),
    "12000": ("REMOVAL", "pmt70", "520", "P"),
    "13000": ("INSTALLATION", "pmt71", "720", "P"),
    "14000": ("SERVICING", "pmt72", "200", "P"),
    "15000": ("STORAGE (INCLUDING TRANSPORTATION)", "pmt73", "800", "P"),
    "16000": ("REWORK (SERVICE BULLETIN ACCOMPLISHMENT PROCEDURES)", "pmt74",
              "664", "P"),
    "17000": ("APPENDIX", "pmt76", "100", "P")
}


def exit_handler_partial(logText, status=0, filename="log.txt", cc=None):
    if cc is not None:
        cc.exit_handler(status)
    elif len(logText) > 0:
        try:
            log = ""
            for l in logText:
                log += l + '\n'
            Path(filename).write_text(log.strip('\n'))
        except Exception:
            print(traceback.format_exc())
            if status == 0:
                status = 1
        finally:
            input("Press any key to continue")
            sys.exit(status)
    else:
        input("Press any key to continue")
        sys.exit(status)


def log_print_partial(logText, message, printToLogOnly=False, cc=None):
    if cc is not None:
        cc.log_print(message, printToLogOnly)
    else:
        if(not printToLogOnly):
            print(message)
        logText.append(str(message))


# Gets input using Launchpad or generic method
def get_input_partial(message, default=None, cc=None):
    if not cc:
        return input(message)
    else:
        return cc.get_user_input(message, default)


def get_source():  # Get the source file
    source = list(Path('.').glob('*.sgm'))
    if len(source) == 1:
        print("Found source file.")
        return Path(source[0])
    else:
        print(">>>Please input the Source file:")
        source = Path(input())
        while not source.is_file():
            print(">>>File does not exist. Please try again:")
            source = Path(input())
        return source


class Publication():
    how_to = False

    def lookup_task_number(self, func):
        if func in self.t:
            self.t[func] += 1
        else:
            self.t[func] = 801

        return add_leading_zero(self.t[func], 3)

    def lookup_subtask_number(self, func):
        if func in self.st:
            self.st[func] += 1
        else:
            self.st[func] = 1

        return add_leading_zero(self.st[func], 3)

    def lookup_code_var(self, infocode, generic):
        if not generic:
            code_var = self.info_codes[infocode][0]
            if self.info_codes[infocode][0] == 'Z':
                self.info_codes[infocode][0] = 'A'
                self.info_codes[infocode][1] += 1
            else:
                self.info_codes[infocode][0] = chr(ord(self.info_codes[infocode][0]) + 1)

            return code_var
        else:
            code_var = self.generic_info_codes[infocode][0]
            if self.generic_info_codes[infocode][0] == 'Z':
                self.generic_info_codes[infocode][0] = 'A'
                self.generic_info_codes[infocode][1] += 1
            else:
                self.generic_info_codes[infocode][0] = chr(ord(self.generic_info_codes[infocode][0]) + 1)
            return code_var

    def __init__(self):
        self.st = {}
        self.t = {}
        self.info_codes = defaultdict(lambda: ['A', 1])
        # self.info_codes = {
        #     "018": ['A', 1], "040": ['A', 1], "050": ['A', 1], "100":  ['A', 1], "400":  ['A', 1],
        #     "051":  ['A', 1], "500":  ['A', 1], "200":  ['A', 1], "250": ['A', 1],
        #     "300":  ['A', 1],  "600":  ['A', 1],  "700":  ['A', 1],
        #     "711":  ['A', 1], "710": ['A', 1], "900":  ['A', 1],
        #     "520":  ['A', 1],  "720":  ['A', 1], "250":  ['A', 1],
        #     "800":  ['A', 1], "664":  ['A', 1], "910": ['B', 1]}
        self.generic_info_codes = defaultdict(lambda: ['A', 0])
        # self.generic_info_codes = {
        #     "018": ['A', 0], "100": ['A', 0], "400": ['A', 0],
        #     "051": ['A', 0], "500": ['A', 0], "200": ['A', 0],
        #     "300": ['A', 0],  "600": ['A', 0],  "700": ['A', 0],
        #     "711": ['A', 0], "710": ['A', 0], "900": ['A', 0],
        #     "520": ['A', 0],  "720": ['A', 0], "250": ['A', 0],
        #     "800": ['A', 0], "664": ['A', 0], "910": ['B', 0]}


class ToolAndCons():
    tools = []
    cons = []

    def get_cir_id(self, list):
        for item in list:
            if item.pn == self.pn:
                if item.nom == self.nom:
                    self.id = item.id
                    break
        else:
            self.id = "{}-{}".format(self.type, add_leading_zero(len(list) + 1, 4))
            list.append(self)

    def create_sed(self, type):
        return """<{10}Descr id="{0}"><name>{1}{2}{8}</name>
    <shortName>{1}{2}</shortName>
    <{9}Ref {9}Number="{0}"{11}>
    <refs>
    <dmRef><dmRefIdent><dmCode assyCode="{3}" disassyCode="00"
    disassyCodeVariant="A" infoCode="00{12}" infoCodeVariant="A"
    itemLocationCode="D" modelIdentCode="HON{4}" subSubSystemCode="{5}"
    subSystemCode="{6}" systemCode="{7}" systemDiffCode="EAA"/></dmRefIdent>
    </dmRef>
    </refs>
    </{9}Ref>
    <reqQuantity>AR</reqQuantity>
    </{10}Descr>""".format(self.id, "{} ".format(self.pn) if self.pn != "" else "", self.nom, self.book_info['ata'][6:8], self.book_info['cage'], self.book_info['ata'][4], self.book_info['ata'][3], self.book_info['ata'][0:2], " ({})".format(self.desc) if self.desc != "" else "", type if type == "tool" else "supply", "supportEquip" if type == "tool" else "supply", ' supplyNumberType="sp01"' if type == "cons" else "", "L" if type == "cons" else "N")

    def create_spec(self, type):
        if type == "cons":
            return """<supplySpec><supplyIdent id="{0}" supplyNumber="{0}" supplyNumberType="sp01"/><name>{1}</name><shortName>{2}{1}</shortName></supplySpec>\n""".format(self.id, self.nom, "{} ".format(self.pn) if self.pn != "" else "")
        else:
            return """<toolSpec><toolIdent id="{0}" manufacturerCodeValue="{3}" toolNumber="{0}"/><itemIdentData><descrForPart>{1}</descrForPart><shortName>{2}{1}</shortName></itemIdentData><procurementData></procurementData><techData></techData><toolAlts><tool><itemDescr></itemDescr></tool></toolAlts></toolSpec>\n""".format(self.id, self.nom, "{} ".format(self.pn) if self.pn != "" else "", self.src)

    def __init__(self, pn, nom, src, desc, type, book_info):
        self.book_info = book_info
        self.pn = pn
        self.nom = nom.strip()
        self.src = src
        self.type = type
        self.desc = desc
        self.get_cir_id(ToolAndCons.tools if type == "tool" else ToolAndCons.cons)
        self.sed = self.create_sed(type)
        self.spec = self.create_spec(type)


class Task():

    def __init__(self, parent_key, key, func, pgblk, title, book_info, pub):
        self.parent_key = parent_key  # pgblk key?
        self.key = key
        self.authdoc = "{}-{}-{}-A01".format(book_info['ata'], func, pub.lookup_task_number(func)) if func is not None else ""
        self.title = title
        self.children = []
        self.pgblk = pgblk


class DataModule():  # args[content, key, func, parent, infocode, pgblk, not_applicable=False, generic=False, filename=None, infoname=None]
    """Class containing methods and data to create a data module"""
    @staticmethod
    def get_module_name(dmref):
        try:
            dmref = re.sub(r' authorityDocument="[^"]+"', '', dmref)
            vars = re.split(' |\n', dmref)
            vars = [var.split('\"')[1] for var in vars if(len(var.split('\"')) >= 2)]
            dmc = "DMC-%s-%s-%s-%s%s-%s-%s%s-%s%s-%s_sx-US.XML" % (vars[6], vars[10], vars[9], vars[8], vars[7], vars[0], vars[1], vars[2], vars[3], vars[4], vars[5])
            return dmc
        except Exception:
            print(traceback.format_exc())
            return ""

    def get_dmref(self, filename=None):
        if filename is None:
            info_var = self.pub.lookup_code_var(self.infocode, self.generic)
            if not self.generic:
                disassy = add_leading_zero(self.pub.info_codes[self.infocode][1], 2)
                return (
                    '<dmRef{5}><dmRefIdent><dmCode assyCode="{0}"'
                    ' disassyCode="{8}" disassyCodeVariant="A" infoCode="{6}"'
                    ' infoCodeVariant="{7}" itemLocationCode="C"'
                    ' modelIdentCode="HON{1}" subSubSystemCode="{2}"'
                    ' subSystemCode="{3}" systemCode="{4}" systemDiffCode="EAA"/>'
                    '</dmRefIdent></dmRef>'
                    .format(
                        self.book_info['ata'][6:8],
                        self.book_info['cage'],
                        self.book_info['ata'][4],
                        self.book_info['ata'][3],
                        self.book_info['ata'][0:2],
                        (' authorityDocument="{}"'.format(self.authdoc))
                        if self.authdoc != "" else "",
                        self.infocode, info_var, disassy))
            else:
                disassy = add_leading_zero(self.pub.generic_info_codes[self.infocode][1], 2)
                return """<dmRef{5}><dmRefIdent><dmCode assyCode="{0}" disassyCode="{8}"
    disassyCodeVariant="A" infoCode="{6}" infoCodeVariant="{7}"
    itemLocationCode="C" modelIdentCode="HON{1}" subSubSystemCode="{2}"
    subSystemCode="{3}" systemCode="{4}" systemDiffCode="EAA"/></dmRefIdent>
    </dmRef>""".format("00", "AERO", "0", "0", "00", (
                    ' authorityDocument="{}"'.format(self.authdoc))
                        if self.authdoc != "" else "",
                        self.infocode, info_var, disassy)
        else:
            if self.authdoc != "":
                return re.sub('<dmRef>', '<dmRef authorityDocument="{}">'.format(self.authdoc), get_dmref(filename))
            else:
                return get_dmref(filename)

    def __init__(self, content, key, func, parent, infocode, pgblk, book_info,
                 not_applicable=False, generic=False, filename=None,
                 infoname=None, authName=None, pub=None):
        self.book_info = book_info
        self.pub = pub
        self.parent = parent
        self.content = content if not not_applicable \
            else "<title>Not Applicable</title>"
        self.key = key
        self.authdoc = "{}-{}-{}-A01".format(
            self.book_info['ata'], func,
            pub.lookup_subtask_number(func)) if func is not None else ""

        self.infocode = infocode
        if infocode == "018" and not pub.how_to:
            self.filename = "DMC-HONAERO-EAA-{}-{}-{}-01A-018A-C_sx-US.XML".format(self.book_info['ata'][0:2], self.book_info['ata'][3:5], self.book_info['ata'][6:8])
            self.generic = False
            self.dmref = self.get_dmref(self.filename)
            pub.how_to = True
        else:
            self.generic = True if infocode == "018" else generic
            if filename is None:
                self.dmref = self.get_dmref()
                self.filename = DataModule.get_module_name(self.dmref)
            else:
                self.filename = filename
                self.dmref = self.get_dmref(filename)
        self.table_refs = re.findall(re.compile(r'<table.*?id="([^"]+)', re.S), self.content)

        figures = re.findall(re.compile(r'<figure.*?id="([^"]+)(.*?)</figure>', re.S), self.content)
        self.graph_refs = [i[0] for i in figures]
        self.sheet_refs = {}
        for f in figures:
            sheets = re.findall(re.compile(r'<graphic.*?id="([^"]+)', re.S), f[1])
            for i, s in enumerate(sheets):
                self.sheet_refs[s] = (f[0], i+1)

        self.tool_data = ""
        self.cons_data = ""
        self.warn_and_caut_data = ""
        self.infoname = page_blocks.get(pgblk, ("UNIDENTIFIED PAGEBLOCK", "pmt01", "000", "D"))[0] if infoname is None else infoname
        self.authname = 'Paragraph {}'.format(authName)


def get_dmref(filename):
    components = re.split('-', filename)

    comp_dict = {
            "model": components[1],
            "diff": components[2],
            "sys": components[3],
            "subsys": components[4][0],
            "subsub": components[4][1],
            "assy": components[5],
            "disassy": components[6][0:2],
            "disvar": components[6][2],
            "info": components[7][0:3],
            "infovar": components[7][3],
            "loc": components[8].split('_')[0][0]
        }

    return """\n<dmRef><dmRefIdent><dmCode assyCode=\"%s\" disassyCode=\"%s\"
    disassyCodeVariant=\"%s\" infoCode=\"%s\" infoCodeVariant=\"%s\"
    itemLocationCode=\"%s\" modelIdentCode=\"%s\" subSubSystemCode=\"%s\"
    subSystemCode=\"%s\" systemCode=\"%s\" systemDiffCode=\"%s\"/></dmRefIdent>
    </dmRef>""" % (
                comp_dict["assy"], comp_dict["disassy"], comp_dict["disvar"], comp_dict["info"],
                comp_dict["infovar"], comp_dict["loc"], comp_dict["model"], comp_dict["subsub"],
                comp_dict["subsys"], comp_dict["sys"], comp_dict["diff"])


def add_leading_zero(number, digits):
    return "0" * (digits - len(str(number))) + str(number)


def check_match(match, group_num=1):
    if match is not None:
        return(match.group(group_num))
    else:
        return("")
        

def CMMConvert(file, directory, cc=None):
    pub = Publication()
    log_file_name = "cmm_converter_log.txt"
    if cc is None:
        print("Warning: Launchpad Not Found")
    else:
        cc.log_name = log_file_name
    logText = []

    # Define partials. Makes function calls a little cleaner so cc doesn't need to be passed in each call.
    log_print = partial(log_print_partial, logText, cc=cc)
    # get_input = partial(get_input_partial, cc=cc)
    exit_handler = partial(exit_handler_partial, logText, filename=log_file_name, cc=cc)

    source_file = Path(file)
    # user = username
    dir = Path(directory)
    log_print(f"Converting CMM '{source_file.name}'")

    def process_bluesheet(source):
        if not Path('./launchpad/XML_Conv/BlueSheet Template.xlsx').is_file():
            return
        # cage = check_match(p_cage.search(source))
        pbs = p_pgblk.findall(source)
        shutil.copyfile(str(Path('./launchpad/XML_Conv/BlueSheet Template.xlsx').resolve()), str(dir / 'BlueSheet.xlsx'))
        new_workbook = load_workbook(str(dir / 'BlueSheet.xlsx'))
        ws = new_workbook.worksheets[1]
        bluesheet_data = []
        for p in pbs:
            gph = p_graphic.findall(p[4])
            if p[0] == "ipl":
                gph += p_fig.findall(p[4])
            for i, g in enumerate(gph):
                sht = p_sheet.findall(g[0])
                for j, s in enumerate(sht):
                    id = s
                    # icn = ("ICN-{}-0000{}-001-01".format(id, s)) if id != "" else ""

                    for b in bluesheet_data:
                        if id == b[1]:
                            break
                    else:
                        fignbr = check_match(re.search(r'fignbr="([^"]+)"', g[0]))
                        title = re.sub(re.compile(r'<acronym>.*?<acronymTerm>([^<]+).*?</acronym>', re.S), r'\1', g[1])
                        title = re.sub('\n', ' ', title)
                        d = ["I", id, "", "", (fignbr if fignbr != "" else str(i+1)) + ((",SH%s" % (j + 1)) if len(sht) > 1 else ""), title, "A"]
                        bluesheet_data.append(d)
                        ws.append(d)
        ws = new_workbook.worksheets[0]
        
        if cc is not None:
            user_info = cc.get_user_info()
            if user_info is not None:
                ws['B1'] = user_info.fname + " " + user_info.lname
                ws['B2'] = user_info.email

        ws['B4'] = check_match(re.search(r'(?s)<cmm[^>]+?docnbr="([^"]+)"', source))
        ws['B5'] = check_match(re.search(r'\b\d{2}-\d{2}-\d{2}\b', source), 0)
        ws['B6'] = re.sub(r'(?s)<acronym>.*?<acronymTerm>([^<]+).*?</acronym>', r'\1', check_match(re.search(r'<cmpnom>(.*?)</cmpnom>', source)))
        ws['B7'] = check_match(re.search(r'(?s)<cmm[^>]+?spl="([^"]+)"', source))
        ws['B8'] = check_match(re.search(r'<!ENTITY ECCN\w+ "([^"]+)">', source)) or check_match(re.search(r'ECCN:?\s?(\w+)', source))
        ws['B9'] = check_match(re.search(r'(?s)<cmm[^>]+?type="([^"]+)"', source)).upper()

        new_workbook.save(str(dir / 'BlueSheet.xlsx'))
        return source

    def process_source(source_file):  # Do series of find/replaces to remove/modify ATA2200 tags -> S1000D tags
        nonlocal is_irm

        def sub_acro(match):
            content = match.group(0)
            content = re.sub('(acro|abbr)', 'acronym', content)
            content = re.sub('acronymterm', 'acronymTerm', content)
            content = re.sub('acronymname', 'acronymDefinition', content)
            content = re.sub(r'(?s)<acronymTerm>\s*(.*?)\s*</acronymTerm>', r'<acronymTerm>\1</acronymTerm>', content)
            # Add to our list of Acronyms and Abbreviations
            term = check_match(re.search('<acronymTerm>([^<]+)</acronymTerm>', content)).strip()
            defn = check_match(re.search('<acronymDefinition>([^<]+)</acronymDefinition>', content)).replace('\n', ' ').strip()
            for a in acros_and_abbrs:
                if a[0] == term and a[1] == defn:
                    break
            else:
                acros_and_abbrs.append([term, defn])
            return content

        def note_para(match):
            note = re.sub('para>', 'notePara>', match.group(0))
            note = re.sub(r'<unlist[^>]*>', r'<notePara>\n<attentionRandomList>', note)
            note = re.sub(r'</unlist>', r'</attentionRandomList></notePara>', note)
            note = re.sub(r'<numlist[^>]*>', r'<notePara><attentionSequentialList>', note)
            note = re.sub(r'</numlist>', r'</attentionSequentialList></notePara>', note)
            note = re.sub(r'<(/?)unlitem[^>]*>', r'<\1attentionRandomListItem>', note)
            note = re.sub(r'<(/?)numlitem[^>]*>', r'<\1attentionSequentialListItem>', note)
            note = re.sub(r'\n?<(attentionRandomListItem|attentionSequentialListItem)>\n?<notePara>', r'\n<\1><attentionListItemPara>', note)
            note = re.sub(r'</notePara>\n?</(attentionRandomListItem|attentionSequentialListItem)>\n?', r'</attentionListItemPara></\1>\n', note)
            return note

        def replace_entities(text):
            header_entities = p_entityref.findall(text)  # Get entities from the header of the source file

            # If the MasterTextEntities.xml file is present, concatenate the entities from there to the other header entities to make one long list
            master_entities_file = Path('MasterTextEntities_CAPU.xml')
            if master_entities_file.is_file():
                master_entities = p_entityref.findall(master_entities_file.read_text(encoding='utf-8'))
                master_entities_file.unlink()
            else:
                master_entities = []

            header_entities += master_entities

            # Sub out the entities in source
            for e in header_entities:
                text = re.sub('&%s;' % e[0], e[1], text)
            return text

        def replace_graphics(match):
            def replace_gdesc(match, title):
                gdesc = match.group(2)
                dim = ""
                legend = ""
                notes = ""
                for table in p_table.finditer(gdesc):
                    table = table.group(0)
                    if "<title>Key</title>" not in table:  # If there's no table, probably okay to delete the info
                        id = check_match(p_id.search(match.group(1)))
                        table = re.sub(r'<title>(Dimensional|Serviceable)\sLimits</title>', r'<title>\1 Limits for <internalRef internalRefId="{}" internalRefTargetType="irtt09"></internalRef></title>'.format(id), table)
                        dim = table
                    else:
                        table = re.sub('</?para>\n?', '', table)
                        entries = p_entry.findall(table)
                        spanned_entries = [e[1] for e in entries if "namest" in e[0]]
                        single_entries = [e[1] for e in entries if "namest" not in e[0]]

                        pairs = []
                        num_of_cols = len(re.findall(r'<colspec', table)) // 2

                        for i in range(0, num_of_cols):
                            for j, entry in enumerate(single_entries[i * 2:-1:2 * num_of_cols]):
                                pairs.append((entry.strip(), single_entries[j*2*num_of_cols+1+i*2].strip()))

                        if len(pairs) > 0:
                            legend = "<legend>\n<definitionList>\n"
                            for pair in pairs:
                                if pair[0] != "" or pair[1] != "":
                                    if pair[0] == "":
                                        key = " "
                                    else:
                                        key = re.sub(r'\.', '', pair[0])
                                    legend += """<definitionListItem>
    <listItemTerm>{}</listItemTerm>
    <listItemDefinition><para>{}</para>
    </listItemDefinition>
    </definitionListItem>\n""".format(key, pair[1])
                            legend += "</definitionList>\n</legend>\n"
                            if len(spanned_entries) > 0:
                                notes = '<table{}><tgroup cols="1"><colspec colname="col1"/><tbody><row>\n<entry>'.format(' frame="bottom"' if "<note" in spanned_entries[0] else ' frame="none"')
                                for se in spanned_entries:
                                    if "<para>" not in se:
                                        notes += "<para>{}</para>".format(se.strip())
                                    else:
                                        notes += se.strip()
                                notes += "\n</entry></row>\n</tbody></tgroup></table>\n"
                        else:
                            legend = ""
                gdesc = re.sub(p_table, '', gdesc).strip()
                return "{}{}</graphic>{}{}{}".format(match.group(1), "<gdesc>{}</gdesc>".format(gdesc) if gdesc != '' else "", legend, notes, dim)

            def generate_icn(match):
                graphic_nbr = check_match(re.search(r'\d{6}', match.group(1)), 0)
                if graphic_nbr:
                    icn = "ICN-{}-0000{}-001-01".format(book_info['cage'], graphic_nbr)
                else:
                    icn = match.group(1)
                return 'infoEntityIdent="{}"'.format(icn)

            text = match.group(0)
            title = check_match(p_title.search(text))
            text = re.sub(p_sgm_attr, '', text)
            text = re.sub(p_split_graphic, r'\1</graphic>\n<graphic key="\3_f">\n<title>{}</title>\n\2'.format(title), text)
            text = re.sub(r'<(/?)graphic', r'<\1figure', text)
            text = re.sub(r'<(/?)sheet', r'<\1graphic', text)
            text = re.sub('key', 'id', text)
            text = re.sub('gnbr="([^"]+)"', generate_icn, text)
            text = re.sub(r'(\s?)imgarea="hl"', r'\1reproductionWidth="355.6 mm" reproductionHeight="209.55 mm"', text)
            text = re.sub(r'\s?imgarea="[^"]+\n?"', '', text)
            gdesc_sub_report = re.subn(re.compile(r'(<graphic[^>]*>)(?!</graphic>).*?<gdesc>(.*?)</gdesc>\n?</graphic>', re.S), partial(replace_gdesc, title=title), text)

            text = gdesc_sub_report[0]
            # if gdesc_sub_report[1] == 1:
            text = re.sub(re.compile(r'(<table.*?</table>)\n*</figure>', re.S), r'</figure>\n\1', text)
            return text

        def process_tables(match):
            def FootnoteRefRepl(match):
                for i, refid in enumerate(note_ids):
                    if match.group(1) == refid:
                        return "<superScript>%s</superScript>" % str(i+1)
                else:
                    return match.group(0)
            table = match.group(0)
            table = re.sub(p_rht, "", table)
            table = re.sub(p_sgm_attr, '', table)
            # Deal with footnotes in tables. Will add each footnote to a attention sequential list item in the last row of the table, and replace the refids with superscript numbers.
            ftnotes = p_ftnote.findall(table)
            if len(ftnotes) > 0:
                cols = re.findall(r'colname="([^"]+)"', table)
                end_col = cols[-1]
                st_col = cols[0]
                note_ids = []
                note_content = "\n<row>\n<entry colsep=\"0\" nameend=\"%s\" namest=\"%s\">\n<?PubTbl cell border-left-style=\"none\" border-right-style=\"none\"?>\n<note>\n<attentionSequentialList>\n" % (end_col, st_col)
                for fn in ftnotes:
                    note_ids.append(check_match(re.search('ftnoteid="([^"]+)"', fn)))
                    note_content += "<attentionSequentialListItem>"
                    note_paras = p_para.findall(fn)
                    for p in note_paras:
                        note_content += "<attentionListItemPara>%s</attentionListItemPara>\n" % p
                    note_content += "</attentionSequentialListItem>"
                note_content += "\n</attentionSequentialList></note>\n</entry>\n</row></tbody></tgroup>"
                table = re.sub(p_ftnotesub, note_content, table)
                table = re.sub(p_refint, FootnoteRefRepl, table)

            return table

        def sub_ipl(match):

            def sub_ipl_refint(match):
                if match.group(1) in ipl_figs:
                    return "IPL FIGURE {}".format(ipl_figs[match.group(1)])
                else:
                    return match.group(0)
            ipl = match.group(0)
            ipl_figs = re.findall(r'<figure[^>]*?fignbr="([^"]+)"[^>]*?key="([^"]+)"', ipl)
            ipl_figs = dict([(f[1], f[0]) for f in ipl_figs])

            ipl = p_refint.sub(sub_ipl_refint, ipl)
            return ipl
        source = source_file.read_text(encoding='utf-8')
        source = replace_entities(source)
        source_file.write_text(source, encoding='utf-8')

        is_irm = check_match(p_title.search(source)) == "Inspection/Repair Manual"

        # Sub out unsupported elements
        source = re.sub(r'[\r\n]+>', '>\n', source)
        # source = re.sub(r'\\', '/', source)
        source = re.sub(r'<(\?[^>?]+)>', r'<\1?>', source)
        source = re.sub(r'<\?Pub Dtl\?>\n?', '', source)
        source = re.sub(r'<(?:revst|revend)>[\n\s]?', '', source)
        source = re.sub(r'\s?chg="\w"', '', source)
        source = re.sub(r'<\?ITG-STRIP-REV[^>]+>[\n\s]?', '', source)
        source = re.sub(re.compile(r'<chgdesc.*?</chgdesc>\n?', re.S), '', source)
        source = re.sub(r'<!--.*?-->\n?', '', source)
        source = re.sub(p_rev, "", source)
        source = re.sub(p_subtitle, r'<para>\1</para>', source)
        source = re.sub(p_equ, r'</para><para><emphasis emphasisType="em01">\1</emphasis></para>', source)
        source = re.sub(r'</equ><equ>', r'</emphasis></para><para><emphasis emphasisType="em01">', source)
        source = re.sub(p_note, note_para, source)
        source = re.sub(r'<unlist[^>]*>', r'<para><randomList>', source)
        source = re.sub(r'</unlist>', r'</randomList></para>', source)
        source = re.sub(r'<numlist[^>]*>', r'<para><sequentialList>', source)
        source = re.sub(r'</numlist>', r'</sequentialList></para>', source)
        source = re.sub(r'(?:num|un)litem[^>]*>', r'listItem>', source)
        source = re.sub(re.compile(r'<(acro|abbr)>.*?</\1>', re.S), sub_acro, source)

        source = re.sub(r'</?topic>\n?', '', source)
        source = re.sub(r'</?(?:prc)?list\d>\n?', '', source)
        source = re.sub(r'</?(?:prc)?item>\n?', '', source)
        source = re.sub(p_sub_super, r'<\1Script>\2</\1Script>', source)
        source = re.sub(p_colspec, r'\1</colspec>', source)

        source = re.sub(re.compile(r'<table.*?</table>', re.S), process_tables, source)
        book_info['cage'] = check_match(p_cage.search(source))
        book_info['revdate'] = check_match(p_revdate.search(source))
        source = re.sub(r'\s?revdate="[^"]+"', '', source)
        try:
            process_bluesheet(source)  # Process bluesheet stuff before messing with the images
        except Exception:
            print("Bluesheet failed to process:\n")
            print(traceback.format_exc())
        source = re.sub(re.compile(r'<ipl.*?</ipl>', re.S), sub_ipl, source)

        source = re.sub(p_graphic, replace_graphics, source)
        source = re.sub(r'<prcitem(\d)>\n?</prcitem\1>', '', source)
        source = re.sub(r'<l(\d)item>\n?</l\1item>', '', source)  # Found in OHM

        return source

    def create_cir(type, data):
        def sub_refs(match):
            id = match.group(1)
            lst = ToolAndCons.tools if "tool" in match.group(1) else ToolAndCons.cons
            for t in lst:
                if t.id == id:
                    return ("{} {}".format(t.pn, t.nom)).upper()
            else:
                for tsk in tasks:
                    # Check for task references
                    if tsk.key == id:
                        authname = tsk.children[0].authname.split('.')[0] if tsk.children[0].authname is not None else ''
                        if authname != '':
                            authname = "{}. {}".format(authname, tsk.title)
                        return re.sub(' authorityDocument="[^"]+"', ' authorityName="{}"'.format(authname), tsk.children[0].dmref)

                    # Check for subtask or graphic references
                    for dm2 in tsk.children:
                        if dm2.key == id:
                            return re.sub(' authorityDocument="[^"]+"', (' authorityName="{}"'.format(dm2.authname)) if dm2.authname is not None else '', dm2.dmref)
                        elif id in dm2.graph_refs or id in dm2.table_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(id), dm2.dmref)
                        elif id in dm2.sheet_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[id][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[id][1]))
                        elif id in dm2.graph_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(id), dm2.dmref)
                        elif id in dm2.sheet_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[id][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[id][1]))  # Will return a reference to the figure, not the sheet

        content = ""
        infocode_var = "A"
        infocode = "012"
        if type == "tool" or type == "cons":
            content = "<{}Repository>\n".format("tool" if type == "tool" else "supply")
            for d in data:
                content += d.spec
            content += "</{}Repository>\n".format("tool" if type == "tool" else "supply")
            infocode = "00L" if type == "cons" else "00N"
            title = "Supplies CIR" if type == "cons" else "Support Equipment CIR"
        else:
            content = "<{}Repository>\n".format("caution" if type == "caut" else "warning")
            for d in data:
                content += """<{2}Spec><{2}Ident id="{0}" {2}IdentNumber="{0}"/>{1}</{2}Spec>\n""".format(d, warn[d] if type == "warn" else caut[d], "caution" if type == "caut" else "warning")
            content += "</{}Repository>\n".format("caution" if type == "caut" else "warning")
            infocode_var = "B" if type == "caut" else "A"
            title = "Cautions CIR" if type == "caut" else "Warnings CIR"

            content = re.sub(r'<internalRef internalRefId="([^"]+)"[^>]*>\n?</internalRef>', sub_refs, content)

        dmcode = """<dmCode assyCode="{0}" disassyCode="00" disassyCodeVariant="A" infoCode="{5}" infoCodeVariant="{6}" itemLocationCode="D" modelIdentCode="HON{1}" subSubSystemCode="{2}" subSystemCode="{3}" systemCode="{4}" systemDiffCode="EAA"/>""".format(book_info['ata'][6:8], book_info['cage'], book_info['ata'][4], book_info['ata'][3], book_info['ata'][0:2], infocode, infocode_var)
        return """<?xml version="1.0" encoding="UTF-8"?>
<!--Arbortext, Inc., 1988-2014, v.4002-->
<!DOCTYPE dmodule [
<!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
%ISOEntities;
]>
<?Pub Inc?>
<dmodule xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/comrep.xsd">
<identAndStatusSection>
<dmAddress>
<dmIdent>
{0}
<language countryIsoCode="US" languageIsoCode="sx"/>
<issueInfo inWork="00" issueNumber="001"/></dmIdent>
<dmAddressItems>
<issueDate day="{7}" month="{6}" year="{5}"/>
<dmTitle><techName>{4}</techName><infoName>{2}</infoName></dmTitle>
</dmAddressItems></dmAddress>
<dmStatus issueType="new">
<security securityClassification="01"/>
<responsiblePartnerCompany enterpriseCode="{8}"></responsiblePartnerCompany>
<originator enterpriseCode="{8}"></originator>
<applic>
<displayText>
<simplePara>ALL</simplePara>
</displayText>
</applic>
<brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
<qualityAssurance>
<unverified/></qualityAssurance></dmStatus>
</identAndStatusSection>
<content>
<commonRepository>
{1}
</commonRepository>
</content>
</dmodule>""".format(dmcode, content, title, add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8], book_info['cage'])

    def create_appliccrossreftable(filename):
        dmcode = check_match(re.search('<dmCode [^>]+>', get_dmref(filename)), 0)
        return """<?xml version="1.0" encoding="UTF-8"?>
<!--Arbortext, Inc., 1988-2014, v.4002-->
<!DOCTYPE dmodule [
<!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
%ISOEntities;
]>
<?Pub Inc?>
<dmodule
xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/appliccrossreftable.xsd"
xmlns:dc="http://www.purl.org/dc/elements/1.1/"
xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
xmlns:xlink="http://www.w3.org/1999/xlink"
xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<?MENTORPATH ?>
<identAndStatusSection>
<dmAddress>
<dmIdent>
{0}
<language countryIsoCode="US" languageIsoCode="sx"/>
<issueInfo inWork="00" issueNumber="001"/></dmIdent>
<dmAddressItems>
<issueDate day="{10}" month="{9}" year="{8}"/>
<dmTitle><techName>{7}</techName><infoName>Applicability
Cross-reference Table (ACT)</infoName></dmTitle>
</dmAddressItems></dmAddress>
<dmStatus issueType="new">
<security securityClassification="01"/>
<responsiblePartnerCompany enterpriseCode="59364">
</responsiblePartnerCompany>
<originator enterpriseCode="59364"></originator>
<applic id="app-0001">
<displayText>
<simplePara>ALL</simplePara>
</displayText>
</applic>
<brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
<qualityAssurance>
<unverified/></qualityAssurance></dmStatus>
</identAndStatusSection>
<content>
<applicCrossRefTable>
<productAttributeList>
<productAttribute id="PN" productIdentifier="primary">
<name>Part Number</name><?Pub Caret 11?>
<displayName>Part Number</displayName>
<descr>Part Number</descr>
</productAttribute>
<productAttribute id="cage" productIdentifier="primary">
<name>CAGE</name>
<displayName>CAGE</displayName>
<descr>CAGE</descr>
</productAttribute>
</productAttributeList>
<productCrossRefTableRef><dmRef><dmRefIdent><dmCode assyCode="{1}"
disassyCode="00" disassyCodeVariant="A" infoCode="00P"
infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HON{2}"
subSubSystemCode="{3}" subSystemCode="{4}" systemCode="{5}"
systemDiffCode="EAA"/></dmRefIdent></dmRef></productCrossRefTableRef>
</applicCrossRefTable>
</content>
</dmodule>
<?INMEDLNG sx-US?>""".format(dmcode, book_info['ata'][6:8], book_info['cage'], book_info['ata'][4], book_info['ata'][3], book_info['ata'][0:2], add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8])

    def create_prdcrossreftable(filename):  # Create 00P file
        content = ""
        partinfo = p_partinfo.findall(source)
        for pi in partinfo:
            mfrpnr = p_mfrpnr.findall(pi)
            for mp in mfrpnr:
                mfr = check_match(p_mfr.search(mp))
                pnr = check_match(p_pnr.search(mp))
                content += """<product>
<assign applicPropertyIdent="PN" applicPropertyType="prodattr" applicPropertyValue="{}"/>
<assign applicPropertyIdent="cage" applicPropertyType="prodattr" applicPropertyValue="{}"/></product>""".format(pnr, mfr)
                altpnr = re.search(r'<altpnr>\n?<pnr>([^<]+)</pnr>\n?<mfr>([^<]+)</mfr>', mp)
                if altpnr is not None:
                    content += """<product>
<assign applicPropertyIdent="PN" applicPropertyType="prodattr" applicPropertyValue="{}"/>
<assign applicPropertyIdent="cage" applicPropertyType="prodattr" applicPropertyValue="{}"/></product>""".format(altpnr.group(1), altpnr.group(2))

        dmcode = check_match(re.search('<dmCode [^>]+>', get_dmref(filename)), 0)

        return """<?xml version="1.0" encoding="UTF-8"?>
<!--Arbortext, Inc., 1988-2014, v.4002-->
<!DOCTYPE dmodule [
<!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
%ISOEntities;
]>
<?Pub Inc?>
<dmodule xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/prdcrossreftable.xsd" xmlns:dc="http://www.purl.org/dc/elements/1.1/" xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
xmlns:xlink="http://www.w3.org/1999/xlink" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<identAndStatusSection>
<dmAddress>
<dmIdent>
{0}
<language countryIsoCode="US" languageIsoCode="sx"/>
<issueInfo inWork="00" issueNumber="001"/></dmIdent>
<dmAddressItems>
<issueDate day="{7}" month="{6}" year="{5}"/>
<dmTitle><techName>{4}</techName><infoName>Product Cross-reference Table (PCT)</infoName></dmTitle>
</dmAddressItems></dmAddress>
<dmStatus issueType="new">
<security securityClassification="01"/>
<responsiblePartnerCompany enterpriseCode="{1}"></responsiblePartnerCompany>
<originator enterpriseCode="{1}"></originator>
<applic id="app-0001">
<displayText>
<simplePara>ALL</simplePara>
</displayText>
</applic>
<brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
<qualityAssurance>
<firstVerification verificationType="tabtop"/></qualityAssurance></dmStatus>
</identAndStatusSection>
<content>
<productCrossRefTable>
{2}
</productCrossRefTable>
</content>
</dmodule>
<?INMEDLNG sx-US?>
    """.format(dmcode, book_info['cage'], content, add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8])

    def create_descript(dm, add_levelled_para=True):  # Create a Descriptive DM
        dmcode = check_match(re.search('<dmCode [^>]+>', dm.dmref), 0)
        return """<?xml version="1.0" encoding="UTF-8"?>
<!--Arbortext, Inc., 1988-2014, v.4002-->
<!DOCTYPE dmodule [
<!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
%ISOEntities;
]>
<?Pub Inc?>
<dmodule xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/descript.xsd" xmlns:dc="http://www.purl.org/dc/elements/1.1/" xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#" xmlns:xlink="http://www.w3.org/1999/xlink"
xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<identAndStatusSection>
<dmAddress>
<dmIdent>
{0}
<language countryIsoCode="US" languageIsoCode="sx"/>
<issueInfo inWork="00" issueNumber="001"/></dmIdent>
<dmAddressItems>
<issueDate day="{8}" month="{7}" year="{6}"/>
<dmTitle><techName>{5}</techName><infoName>{3}</infoName></dmTitle>
</dmAddressItems></dmAddress>
<dmStatus issueType="new">
<security securityClassification="01"/>
<responsiblePartnerCompany enterpriseCode="{1}"></responsiblePartnerCompany>
<originator enterpriseCode="{1}"></originator>
<applic id="app-0001a">
<displayText>
<simplePara>ALL</simplePara>
</displayText>
</applic>
<brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
<qualityAssurance>
<firstVerification verificationType="tabtop"/></qualityAssurance></dmStatus>
</identAndStatusSection>
<content><description>
{2}
</description></content></dmodule>
<?INMEDCLNPRJ A350_CMMs_-_Honeywell?>
<?INMEDBSPATH A350_CMMs_-_Honeywell?>
<?INMEDLNG sx-US?>
    """.format(dmcode, book_info['cage'], ('<levelledPara>{}</levelledPara>'.format(dm.content)) if add_levelled_para else dm.content, dm.infoname, add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8])

    def create_proced(dm):  # Create a Procedural DM
        dmcode = check_match(re.search('<dmCode [^>]+>', dm.dmref), 0)
        tool_data = dm.tool_data
        cons_data = dm.cons_data
        return """<?xml version="1.0" encoding="UTF-8"?>
<!--Arbortext, Inc., 1988-2014, v.4002-->
<!DOCTYPE dmodule [
<!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
%ISOEntities;
]>
<?Pub Inc?>
<dmodule xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/proced.xsd" xmlns:dc="http://www.purl.org/dc/elements/1.1/" xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#" xmlns:xlink="http://www.w3.org/1999/xlink"
xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<identAndStatusSection>
<dmAddress>
<dmIdent>
{0}
<language countryIsoCode="US" languageIsoCode="sx"/>
<issueInfo inWork="00" issueNumber="001"/></dmIdent>
<dmAddressItems>
<issueDate day="{11}" month="{10}" year="{9}"/>
<dmTitle><techName>{8}</techName><infoName>{6}</infoName></dmTitle>
</dmAddressItems></dmAddress>
<dmStatus issueType="new">
<security securityClassification="01"/>
<responsiblePartnerCompany enterpriseCode="{1}"></responsiblePartnerCompany>
<originator enterpriseCode="{1}"></originator>
<applic id="app-0001a">
<displayText>
<simplePara>ALL</simplePara>
</displayText>
</applic>
<brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
<qualityAssurance>
<firstVerification verificationType="tabtop"/></qualityAssurance></dmStatus>
</identAndStatusSection>
<content>
{5}
<procedure><preliminaryRqmts>
<reqCondGroup><noConds/></reqCondGroup>
<reqSupportEquips>
{3}
</reqSupportEquips>
<reqSupplies>
{4}
</reqSupplies>
<reqSpares><noSpares/></reqSpares>
<reqSafety><noSafety/></reqSafety>
</preliminaryRqmts>
<mainProcedure>
<proceduralStep>
{2}
</proceduralStep>
</mainProcedure>
<closeRqmts>
<reqCondGroup>
<noConds/></reqCondGroup>
</closeRqmts>
</procedure></content></dmodule>
<?INMEDCLNPRJ A350_CMMs_-_Honeywell?>
<?INMEDBSPATH A350_CMMs_-_Honeywell?>
<?INMEDLNG sx-US?>
    """.format(dmcode, book_info['cage'], dm.content, "<supportEquipDescrGroup>{}</supportEquipDescrGroup>".format(tool_data) if tool_data.strip() != "" else "<noSupportEquips/>", "<supplyDescrGroup>{}</supplyDescrGroup>".format(cons_data) if cons_data != "" else "<noSupplies/>", ("<warningsAndCautionsRef>{}</warningsAndCautionsRef>".format(dm.warn_and_caut_data)) if dm.warn_and_caut_data != "" else "", dm.infoname, add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8])

    def create_pmc(content):  # Create the PMC
        return """<?xml version="1.0" encoding="UTF-8"?>
<!--Arbortext, Inc., 1988-2014, v.4002-->
<!DOCTYPE pm [
<!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
%ISOEntities;
]>
<?Pub Inc?>
<pm
xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/pm.xsd"
xmlns:dc="http://www.purl.org/dc/elements/1.1/"
xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
xmlns:xlink="http://www.w3.org/1999/xlink"
xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<identAndStatusSection><pmAddress>
<pmIdent>
<pmCode modelIdentCode="HON{0}" pmIssuer="{0}" pmNumber="00001"
pmVolume="01"/><language countryIsoCode="US" languageIsoCode="sx"/>
<issueInfo inWork="00" issueNumber="{6}"/></pmIdent>
<pmAddressItems>
<issueDate day="{9}" month="{8}" year="{7}"/>
<pmTitle>{13}</pmTitle>
<shortPmTitle>{14}</shortPmTitle>
<externalPubCode pubCodingScheme="initialDate">{12} {11} {10}</externalPubCode>
<externalPubCode pubCodingScheme="CMP">{3}</externalPubCode>
<externalPubCode pubCodingScheme="INT">{5}</externalPubCode>
</pmAddressItems>
</pmAddress><pmStatus issueType="new">
<security securityClassification="01"/>
<dataRestrictions>
<restrictionInstructions>
<dataDistribution></dataDistribution>
<exportControl>
<exportRegistrationStmt>
<simplePara>These items are controlled by the U.S. government and
authorized for export only to the country of ultimate destination
for use by the ultimate consignee or end-user(s) herein identified.
They may not be resold, transferred, or otherwise disposed of, to
any other country or to any person other than the authorized ultimate
consignee or end-user(s), either in their original form or after being
incorporated into other items, without first obtaining approval from
the U.S. government or as otherwise authorized by U.S. law and regulations.</simplePara>
<simplePara>{2}</simplePara>
</exportRegistrationStmt>
</exportControl>
<dataHandling></dataHandling>
<dataDestruction></dataDestruction>
<dataDisclosure></dataDisclosure>
</restrictionInstructions>
<restrictionInfo>
<copyright>
<copyrightPara>&copy; Honeywell International Inc. Do not copy without
express permission of Honeywell.</copyrightPara>
</copyright>
<policyStatement></policyStatement>
<dataConds></dataConds>
</restrictionInfo>
</dataRestrictions>
<responsiblePartnerCompany enterpriseCode="{0}">
</responsiblePartnerCompany>
<originator enterpriseCode="{0}"></originator>
<applicCrossRefTableRef>
{4}
</applicCrossRefTableRef>
<applic>
<displayText>
<simplePara>ALL</simplePara>
</displayText>
</applic>
<brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
<qualityAssurance>
<unverified/></qualityAssurance>
</pmStatus></identAndStatusSection>
<content>
{1}
</content>
</pm>
<?INMEDLNG sx-US?>""".format(book_info['cage'],
                             content,
                             book_info['eccn'],
                             book_info['ata'],
                             get_dmref("DMC-HON{}-EAA-{}-00A-00WA-D_sx-US.XML".format(book_info['cage'],
                                                                                      book_info['ata'])),
                                                                                      book_info['docnbr'],
                                                                                      add_leading_zero(book_info['issue_no'], 3),
                                                                                      book_info['revdate'][0:4],
                                                                                      book_info['revdate'][4:6],
                                                                                      book_info['revdate'][6:8],
                                                                                      book_info['oidate'][0:4],
                                                                                      months[int(book_info['oidate'][4:6])],
                                                                                      book_info['oidate'][6:8],
                                                                                      book_info['cmpnom'],
                                                                                      book_info['model'])

    def GetServiceBulletin():
        nonlocal pub

        def ReplaceRow(match):
            if '/' in match.group(0):
                return "</row>"
            else:
                return "<row rowsep=\"0\">"

        def ReplaceEntry(match):
            if '/' in match.group(0):
                return "</para></entry>"
            else:
                return "<entry><para>"

        filename_008A = "DMC-HONAERO-EAB-00-00-00-00A-008A-D_sx-US.XML"

        sblist = re.search(p_sblist, source)

        table_body = ""
        sbdata = p_sbdata.findall(sblist.group(0))

        for s in sbdata:
            s = re.sub('</sbtitle><issdate>', '</sbtitle><entry></entry><issdate>', s)
            s = re.sub(r"<(?!\?)/?(?!sbdata)\w+>", ReplaceEntry, s)
            table_body += s + '\n'

        table_body = re.sub(p_sbdatatag, ReplaceRow, table_body)

        content_008A = """<table id="SB-1" frame="topbot"><tgroup cols="4">
<?PubTbl tgroup dispwid="700.00px"?>
<colspec colname="col1" colsep="0" colwidth="25.18*" /><colspec
colname="col2" colsep="0" colwidth="43.81*" /><colspec colname="col3"
colsep="0" colwidth="14.56*" /><colspec colname="COLSPEC0"
colwidth="16.45*" /><thead><row rowsep="1">
<entry>
<para>Service Bulletin /</para>
<para>Revision Number</para>
</entry>
<entry>
<para>Title</para>
</entry>
<entry>
<para>Modification</para>
</entry>
<entry>
<para>Date Put In</para>
<para>Manual</para>
</entry>
</row></thead><tbody>
%s
</tbody></tgroup></table>""" % (table_body)
        return DataModule(content_008A, "sbl", None, None, None, None, book_info, filename=filename_008A, infoname="SERVICE BULLETIN LIST", pub=pub)

    def GetTemporaryRevision():
        nonlocal pub
        empty_entry = "\n<entry></entry>"

        filename_003C = "DMC-HONAERO-EAB-00-00-00-00A-003C-D_sx-US.XML"

        trlist = re.search(p_trlist, source)  # Get data

        table_body = ""
        trdata = p_trdata.findall(trlist.group(0))

        for t in trdata:
            table_body += t + '\n'

        for r in (("<trdata>", "<row rowsep=\"0\">"), ("</trdata>", "%s</row>" % (empty_entry * 5)), ("<trnbr>", "<entry><para>"), ("<trstatus>", "<entry><para>"),  ("</trnbr>", "</para></entry>"), ("</trstatus>", "</para></entry>"), ("<trloc>", "<entry><para>"), ("</trloc>", "</para></entry>")):
            table_body = table_body.replace(*r)

        content_003C = """<para>Instructions on each page of a temporary revision tell you where to put the pages in your manual. Remove the temporary revision pages only when discard instructions are given. For each temporary revision, put the
applicable data in the record columns on this page.</para>
<para>Definition of Status column: A <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> may be active, incorporated, or deleted. &ldquo;Active&rdquo; is entered by the holder of the manual. &ldquo;Incorporated&rdquo;
means a <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> has been incorporated into the manual and includes the revision number of the manual when the <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary
revision</acronymDefinition></acronym> was incorporated. &ldquo;Deleted&rdquo; means a <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> has been replaced by another <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary
revision</acronymDefinition></acronym>, a <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> number will not be issued, or a <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> has
been deleted.</para>
<table id="TR-1" frame="topbot"><tgroup cols="8">
<?PubTbl tgroup dispwid="700.00px"?>
<colspec align="center" colname="col1" colsep="0" colwidth="1.21*"/><colspec colname="col2" colsep="0" colwidth="1.48*"/>
<colspec align="center" colname="col3" colsep="0" colwidth="0.98*"/><colspec align="center" colname="col4" colsep="0" colwidth="1.30*"/>
<colspec align="center" colname="col5" colsep="0" colwidth="1.30*"/><colspec colname="col6" colsep="0" colwidth="0.87*"/>
<colspec colname="col7" colsep="0" colwidth="1.15*"/><colspec colname="col8" colsep="0" colwidth="0.70*"/><thead><row rowsep="1">
<entry colsep="0" valign="bottom">
<?PubTbl cell border-left-style="none"?>
<para>Temporary</para>
<para>Revision</para>
<para>Number</para>
</entry>
<entry align="center" colsep="0" valign="bottom">
<para>Status</para>
</entry>
<entry colsep="0" valign="bottom">
<para>Page Number</para>
</entry>
<entry colsep="0" valign="bottom">
<para>Issue Date</para>
</entry>
<entry colsep="0" valign="bottom">
<para>Date Put In</para>
<para>Manual</para>
</entry>
<entry align="center" colsep="0" valign="bottom">
<para>By</para>
</entry>
<entry align="center" colsep="0" valign="bottom">
<para>Date</para>
<para>Removed</para>
<para>From</para>
<para>Manual</para>
</entry>
<entry align="center" valign="bottom">
<?PubTbl cell border-right-style="none"?>
<para>By</para>
</entry>
</row></thead><tbody>
%s
</tbody></tgroup></table>""" % (table_body)
        # TODO: Validate number of entries?
        return DataModule(content_003C, "rotr", None, None, None, None, book_info, filename=filename_003C, infoname="RECORD OF TEMPORARY REVISIONS", pub=pub)

    def GetVendorCodes():
        """Creates 018Z File"""
        nonlocal pub

        def AddV(match):
            """Adds the 'V' to the beginning of the vendor code"""
            return "row rowsep=\"0\">%s<entry><para>V%s" % (match.group(1) if match.group(1) is not None else '', match.group(2))

        filename_018Z = "DMC-HON{}-EAA-00-00-00-00A-018Z-C_sx-US.XML".format(book_info['cage'])

        vendata = re.findall(p_vendata, source)
        if len(vendata) == 0:
            print("Could Not Locate Vendor Code List Data In The Source.")
            return

        table_body = ""

        for v in vendata:
            table_body += v + '\n'

        table_body = re.sub(p_mfrcont, "", table_body)

        for r in (
                ("<vendata>", "<row rowsep=\"0\">"),
                ("</vendata>", "</row>"),
                ("<mad>", "<entry><para>"), ("<mfr>", "<entry><para>"),
                ("</mad>", "</para></entry>"),
                ("</mfr>", "</para></entry>")):
            table_body = table_body.replace(*r)

        table_body = re.subn(p_v, AddV, table_body)[0]

        content_018Z = """<note><notePara>The vendor codes and part numbers that are shown in the DPL must not be construed as an authorization of the vendor, pursuant to the FAA regulations, to ship directly to the user. Neither must it be construed
as a certification by Honeywell that parts shipped by vendors directly to users will conform to the type design or that such parts are airworthy or safe for installation.</notePara></note>
<table frame="topbot"><tgroup cols="2">
<?PubTbl tgroup dispwid="700.00px"?>
<colspec colname="col1" colsep="0" colwidth="0.28*"/><colspec
colname="col2" colwidth="1.72*"/><thead><row rowsep="0">
<entry rowsep="1">
<para>Code</para>
</entry>
<entry rowsep="1">
<para>Vendor</para>
</entry>
</row></thead><tbody>
%s
</tbody></tgroup></table>""" % (table_body)

        return DataModule(content_018Z, "vendlist", None, None, None, None, book_info, filename=filename_018Z, infoname="Vendor Code List", pub=pub)

    def GetTransmittalInfo():
        nonlocal pub
        filename_003A = "DMC-HONAERO-EAB-00-00-00-00A-003A-D_sx-US.XML"
        transltr = check_match(p_transltr.search(source))

        return DataModule(transltr, "transltr", None, None, None, None, book_info, filename=filename_003A, infoname="TRANSMITTAL INFORMATION", pub=pub)

    def GetRecordOfRevisions():
        nonlocal pub
        filename_003B = "DMC-HONAERO-EAB-00-00-00-00A-003B-D_sx-US.XML"
        content = """<para>For each revision, write the revision number, revision date, date put in the manual, and your initials in the applicable column.</para><note>
<notePara>Refer to the Revision History in the <dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="003" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0"
systemCode="00" systemDiffCode="EAB"/></dmRefIdent></dmRef> section for revision data.</notePara>
</note><table><tgroup cols="9">
<?PubTbl tgroup dispwid="1230.00px"?>
<colspec colname="col1" colwidth="0.92in"/><colspec colname="col2" colwidth="0.86in"/><colspec colname="col3" colwidth="0.78in"/><colspec colname="col4" colwidth="0.51in"/><colspec colname="col5" colwidth="0.31in"/><colspec colname="col6" colwidth="0.76in"/>
<colspec colname="col7" colwidth="0.75in"/><colspec colname="col8" colwidth="0.74in"/><colspec colname="col9" colwidth="0.47in"/><thead><row>
<entry align="center">
<para>Revision Number</para>
</entry>
<entry align="center">
<para>Revision Date</para>
</entry>
<entry align="center">
<para>Date Put in Manual</para>
</entry>
<entry align="center">
<para>By</para>
</entry>
<entry></entry>
<entry align="center">
<para>Revision Number</para>
</entry>
<entry align="center">
<para>Revision Date</para>
</entry>
<entry align="center">
<para>Date Put in Manual</para>
</entry>
<entry align="center">
<para>By</para>
</entry>
</row></thead><tbody>{}</tbody></tgroup></table>""".format("""<row>\n<entry></entry>\n<entry></entry>\n<entry></entry>\n<entry></entry>\n<entry></entry>\n<entry></entry>\n<entry></entry>\n<entry></entry>\n<entry></entry>\n</row>""" * 27)
        return DataModule(content, "ror", None, None, None, None, book_info, filename=filename_003B, infoname="RECORD OF REVISIONS", pub=pub)

    def GetBookInfo(source):  # Get metadata from book
        # Get ATA number
        book_info['ata'] = "{}-{}-{}".format(check_match(re.search(r'chapnbr="([^"]+)"', source)), check_match(re.search(r'sectnbr="([^"]+)"', source)), check_match(re.search(r'subjnbr="([^"]+)"', source)))

        book_info['issue_no'] = check_match(p_tsn.search(source))
        book_info['oidate'] = check_match(p_oidate.search(source))
        book_info['eccn'] = check_match(p_eccn.search(source), 0)
        book_info['cpyrgt'] = check_match(p_copyright.search(source))
        book_info['docnbr'] = check_match(p_docnbr.search(source))
        book_info['model'] = check_match(p_model.search(source))
        book_info['cmpnom'] = check_match(p_cmpnom.search(source))

        # PM Title

        return book_info

    def GetPrelim():
        nonlocal pub
        names = ["DMC-HONAERO-EAA-00-00-00-00A-023A-D_sx-US.XML", "DMC-HONAERO-EAA-00-00-00-00A-010A-D_sx-US.XML", "DMC-HONAERO-EAA-00-00-00-00A-012A-D_sx-US.XML", "DMC-HONAERO-EAA-00-00-00-00A-012B-D_sx-US.XML"]
        prelims = []
        proptary = p_proptary.findall(source)
        for p in zip(proptary, names):
            prelims.append(DataModule(p[0], "proptary", None, None, None, None, book_info, filename=p[1], infoname="Preliminary Information", pub=pub))
        cpy = """<title>Copyright - Notice</title>
    <para>Copyright {} Honeywell International Inc. All rights
    reserved.</para>
    <para>Honeywell is a registered trademark of Honeywell International
    Inc.</para>
    <para>All other marks are owned by their respective companies.</para>""".format(book_info['cpyrgt'])
        prelims.append(DataModule(cpy, "cpy", None, None, None, None, book_info, filename="DMC-HONAERO-EAA-00-00-00-00A-021A-D_sx-US.XML", infoname="Copyright Information", pub=pub))
        return prelims

    def get_irm_pmt(title, pgblknbr):
        title = title.upper()
        if "GENERAL CHECK" in title:
            return "pmt78"
        elif "CONTINUE-TIME" in title:
            return "pmt80"
        elif "ZERO-TIME" in title:
            return "pmt81"
        elif "INSPECTION/CHECK" in title:
            return "pmt79"
        elif "GENERAL REPAIR" in title:
            return "pmt82"
        elif "REPAIR" in title:
            return "pmt83"
        else:
            return page_blocks.get(pgblknbr, ("UNIDENTIFIED PAGEBLOCK", "pmt01", "000", "D"))[1]

    def GetPMCData():
        nonlocal vcl
        nonlocal pub
        try:  # Prelim Info
            pm_text = '<pmEntry pmEntryType="pmt77">\n<pmEntryTitle>Proprietary Information</pmEntryTitle>'
            prelims = GetPrelim()
            for p in prelims:
                front_matter.append(p)
                pm_text += p.dmref

            pm_text += '</pmEntry>'

            # Add Transmittal Info
            pm_text += '<pmEntry pmEntryType="pmt52">\n<pmEntryTitle>TRANSMITTAL INFORMATION</pmEntryTitle>\n'
            transltr_dm = GetTransmittalInfo()
            front_matter.append(transltr_dm)
            pm_text += transltr_dm.dmref
            pm_text += '</pmEntry>'

            # Add Record of Revisions
            pm_text += '<pmEntry pmEntryType="pmt53">\n<pmEntryTitle>RECORD OF REVISIONS</pmEntryTitle>\n'
            ror_dm = GetRecordOfRevisions()
            front_matter.append(ror_dm)
            pm_text += ror_dm.dmref
            pm_text += '</pmEntry>'

            # Add Record of Temporary Revisions
            pm_text += '<pmEntry pmEntryType="pmt54">\n<pmEntryTitle>RECORD OF TEMPORARY REVISIONS</pmEntryTitle>\n'
            rotr_dm = GetTemporaryRevision()
            front_matter.append(rotr_dm)
            pm_text += rotr_dm.dmref
            pm_text += '</pmEntry>'

            # Add SBL
            pm_text += '<pmEntry pmEntryType="pmt55">\n<pmEntryTitle>SERVICE BULLETIN LIST</pmEntryTitle>\n'
            sbl_dm = GetServiceBulletin()
            front_matter.append(sbl_dm)
            pm_text += sbl_dm.dmref
            pm_text += '</pmEntry>'

            pm_text += """<pmEntry pmEntryType="pmt56">
    <pmEntryTitle>LIST OF EFFECTIVE PAGES</pmEntryTitle>
    <externalPubRef authorityDocument="LEP"><externalPubRefIdent></externalPubRefIdent></externalPubRef>
    </pmEntry>"""
        except Exception:
            print("Failed while adding front matter")
            print(traceback.format_exc())

        pgblks = p_pgblk.findall(source)
        for pb in pgblks:
            if pb[0] == "ipl":
                pub.generic_info_codes["018"] = ["T", 0]
            pb_key = check_match(re.search(r'key="([^"]+)"', pb[1]))
            pgblknbr = check_match(re.search(r'pgblknbr="([^"]+)"', pb[1]))
            pgblknbr = pgblknbr[:-1] + "0" if len(pgblknbr) > 1 else pgblknbr
            pblk = page_blocks.get(pgblknbr, ("UNIDENTIFIED PAGEBLOCK", "pmt01", "000", "D"))
            title = pb[3].upper() if pb[3] != "" else pblk[0]
            confnbr = check_match(re.search(r'confnbr="([^"]+)"', pb[1]))
            effect = pb[2]
            authName = ""
            if effect != "":
                title += ", PN {}".format(effect)
                authName = effect
            elif confnbr != "" and confnbr != "1":
                title += " {}".format(confnbr)
                authName = confnbr
            content = pb[4]
            
            pmt_code = pblk[1] if not is_irm else get_irm_pmt(title, pgblknbr)
            pm_text += '<pmEntry {}pmEntryType="{}">\n<pmEntryTitle>{}</pmEntryTitle>\n'.format("" if authName == "" else 'authorityName="{}" '.format(authName), pmt_code, title)
            source_tasks = p_tasks.findall(content)
            if content != "<isempty/>" and content != "" and len(source_tasks) > 0:
                if len(source_tasks) > 0:
                    for i, task in enumerate(source_tasks):
                        task_title = task[2]
                        task_title = re.sub(re.compile(r'<acronym>\n?<acronymTerm>([^<]+).*?</acronym>', re.S), r'\1', task_title)
                        task_obj = Task(pb_key, task[1], task[0], pgblknbr, task_title, book_info, pub)
                        task_content = task[3]
                        pm_text += "<pmEntry{}>\n<pmEntryTitle>{}</pmEntryTitle>\n".format(' authorityDocument="{}"'.format(task_obj.authdoc) if task_obj.authdoc != "" and not is_irm else "", task_title)
                        subtasks = p_subtask.findall(task_content)
                        for j, st in enumerate(subtasks):
                            st_key = st[1]
                            st_func = st[0]
                            st_content = st[2]
                            if "<title>Job Setup Data</title>" in st_content:
                                st_content = fill_tools_and_cons(content, st_content)
                            dm_obj = DataModule(st_content, st_key, st_func, task_obj, pblk[2], pgblknbr, book_info, authName="{}.{}{}.".format(i + 1, chr((ord('A') - 1) + j // 26) if j >= 26 else "", chr((ord('A') + j % 26))), infoname=title, pub=pub)
                            dm_text = dm_obj.dmref if not is_irm else re.sub(r'\sauthorityDocument="[^"]*"', '', dm_obj.dmref)
                            pm_text += "{}\n".format(dm_text)
                            task_obj.children.append(dm_obj)
                        tasks.append(task_obj)
                        pm_text += "</pmEntry>\n"
            else:
                dm_obj = DataModule(None, None, None, None, pblk[2], pgblknbr, book_info, True, pub=pub)
                not_applicable.append(dm_obj)
                pm_text += dm_obj.dmref
                pm_text += '</pmEntry>\n'
                continue
            if pb[0] == "ipl":
                vendlist_dm = GetVendorCodes()
                if vendlist_dm is not None:
                    vcl = vendlist_dm
                    # Add VCL
                    pm_text += '<pmEntry>\n<pmEntryTitle>Vendor Code List</pmEntryTitle>'

                    pm_text += vendlist_dm.dmref
                    # Add DPL
                    pm_text += '</pmEntry>\n'
                pm_text += '<pmEntry>\n<pmEntryTitle>Detailed Parts List</pmEntryTitle>\n'
                pm_text += '<pmEntry></pmEntry>\n'  # Replace this eventually
                pm_text += '</pmEntry>'
            pm_text += "\n</pmEntry>"
        return pm_text

    def find_duplicate_id(obj, list):
        for item in list:
            if item.id == obj.id:
                return True
        else:
            return False

    def process_data_module(dm, schema):
        local_tools = []
        local_cons = []
        local_warn_caut = []

        def get_warncaut_ref(id):
            return """<warningRef id="{0}" warningIdentNumber="{0}"><dmRef><dmRefIdent><dmCode assyCode="{1}" disassyCode="00"
disassyCodeVariant="A" infoCode="012" infoCodeVariant="{6}" itemLocationCode="D" modelIdentCode="HON{2}"
subSubSystemCode="{3}" subSystemCode="{4}" systemCode="{5}" systemDiffCode="EAA"/></dmRefIdent></dmRef></warningRef>\n""".format(id, book_info['ata'][6:8], book_info['cage'], book_info['ata'][4], book_info['ata'][3], book_info['ata'][0:2], "A" if "warn" in id else "B")

        def replace_tool(match):
            tool = re.sub(r'\n', ' ', match.group(0))
            pn = check_match(re.search(re.compile(r'<(?:tool|std)nbr>(.*?)</(?:tool|std)nbr>', re.S), tool))
            nom = check_match(re.search(re.compile(r'<(?:tool|std)name>(.*?)</(?:tool|std)name>', re.S), tool))
            src = check_match(re.search(re.compile(r'<(?:tool|std)src>(.*?)</(?:tool|std)src>', re.S), tool))
            desc = check_match(re.search(re.compile(r'<(?:tool|std)desc>(.*?)</(?:tool|std)desc>', re.S), tool))
            new_tool = ToolAndCons(pn, re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.S), r'\1', nom), src, re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.S), r'\1', desc), "tool", book_info)  # Create new tool object. ID is generated on creation

            # If that object doesn't exist as a tool in this data module yet, add it
            if not find_duplicate_id(new_tool, local_tools):
                local_tools.append(new_tool)

            return '<internalRef internalRefId="{}" internalRefTargetType="irtt05"></internalRef>'.format(new_tool.id)

        def replace_cons(match):
            cons = re.sub(r'\n', ' ', match.group(0))
            pn = check_match(re.search(re.compile(r'<connbr>(.*?)</connbr>', re.S), cons))
            nom = check_match(re.search(re.compile(r'<conname>(.*?)</conname>', re.S), cons))
            src = check_match(re.search(re.compile(r'<consrc>(.*?)</consrc>', re.S), cons))
            desc = check_match(re.search(re.compile(r'<condesc>(.*?)</condesc>', re.S), cons))
            new_cons = ToolAndCons(pn, re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.S), r'\1', nom), src, re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.S), r'\1', desc), "cons", book_info)  # Create new cons object. ID is generated on creation

            # If that object doesn't exist as a tool in this data module yet, add it
            if not find_duplicate_id(new_cons, local_cons):
                local_cons.append(new_cons)

            return '<internalRef internalRefId="{}" internalRefTargetType="irtt04"></internalRef>'.format(new_cons.id)

        def replace_table_warn_caut(match):
            return "<{0}>{1}</{0}>".format(match.group(1), re.sub(r'<(/?)para>', r'<\1warningAndCautionPara>', match.group(2)))

        def replace_warn_and_caut(match):
            caut_ref = ""
            warn_ref = ""
            all = re.sub('<(/?)para>', r'<\1warningAndCautionPara>', match.group(0))
            all = re.sub(r'\n', ' ', all)
            warn_caut = re.findall(re.compile(r'<(warning|caution)>(.*?)</\1>', re.S), all)
            for wc in warn_caut:
                if wc[0] == "warning":
                    for k, w in warn.items():
                        if w == wc[1]:
                            id = k
                            break
                    else:
                        id = '{}-{}'.format("warn", add_leading_zero(len(warn) + 1, 4))
                        warn[id] = re.sub(re.compile(r'<acronym>.*?<acronymDefinition>(.*?)</.*?</acronym>', re.S), r'\1', wc[1])
                    warn_ref += "{} ".format(id)
                else:
                    for k, c in caut.items():
                        if c == wc[1]:
                            id = k
                            break
                    else:
                        id = '{}-{}'.format("caut", add_leading_zero(len(caut) + 1, 4))
                        caut[id] = re.sub(re.compile(r'<acronym>.*?<acronymDefinition>(.*?)</.*?</acronym>', re.S), r'\1', wc[1])
                    caut_ref += "{} ".format(id)

                if id not in local_warn_caut:
                    local_warn_caut.append(id)

            caut_ref = caut_ref.strip()
            warn_ref = warn_ref.strip()

            step = match.group(2)
            if caut_ref != "":
                if "cautionRefs=" in step:
                    step = re.sub(r'cautionRefs="([^"]+)"', r'cautionRefs="\1 {}"'.format(caut_ref), step)
                else:
                    step += ' cautionRefs="{}"'.format(caut_ref)

            if warn_ref != "":
                if "warningRefs=" in step:
                    step = re.sub(r'warningRefs="([^"]+)"', r'warningRefs="\1 {}"'.format(warn_ref), step)
                else:
                    step += ' warningRefs="{}"'.format(warn_ref)

            return '{}\n<{}>'.format(match.group(1) if match.group(1) is not None else "", step)

        def link_graphic_references(match):  # Return a cross reference to a figure
            gid = match.group(1)
            if gid in dm.graph_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt01"></internalRef>'.format(gid)
            elif gid in dm.sheet_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt09"></internalRef>'.format(gid)
            else:
                for t2 in tasks:
                    for dm2 in t2.children:
                        if dm2 != dm:
                            if gid in dm2.graph_refs:
                                return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(gid), dm2.dmref)
                            elif gid in dm2.sheet_refs:
                                return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[gid][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[gid][1]))  # Will return a reference to the figure, not the sheet
                else:
                    return match.group(0)

        def ref_pgblk(match, do_return=False):  # Return the first child of the first task with a parent pgblk key matching the refid
            for tsk in tasks:
                if tsk.parent_key == match.group(1) and len(tsk.children) > 0:
                    return re.sub(' authorityDocument="[^"]+"', '', tsk.children[0].dmref)
            else:
                if not do_return:
                    return ref_int(match, True)
                else:
                    return match.group(0)

        def ref_int(match, do_return=False):
            refid = match.group(1)
            if refid in dm.graph_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt01"></internalRef>'.format(refid)
            elif refid in dm.table_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt02"></internalRef>'.format(refid)
            elif refid in dm.sheet_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt09"></internalRef>'.format(refid)
            elif refid == dm.key:
                return re.sub(' authorityDocument="[^"]+"', (' authorityName="{}"'.format(dm.authname if match.group(2) == "" else dm.authname.strip('.'))) if dm.authname is not None else '', dm.dmref)

            for tsk in tasks:
                if tsk.key == refid:
                    return re.sub(' authorityDocument="[^"]+"', (' authorityName="{}"'.format(tsk.children[0].authname.split('.')[0])) if tsk.children[0].authname is not None else '', tsk.children[0].dmref)
                for dm2 in tsk.children:
                    if dm == dm2:
                        continue
                    if dm2.key == refid:
                        return re.sub(' authorityDocument="[^"]+"', (' authorityName="{}"'.format(dm2.authname if match.group(2) == "" else dm2.authname.strip('.'))) if dm2.authname is not None else '', dm2.dmref)
                    elif refid in dm2.graph_refs or refid in dm2.table_refs:
                        return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(refid), dm2.dmref)
                    elif refid in dm2.sheet_refs:
                        return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[refid][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[refid][1]))  # Will return a reference to the figure, not the sheet
            else:
                if not do_return:
                    return ref_pgblk(match, True)
                else:
                    return match.group(0)

        def delegate_refint(match):
            if 'reftype="pgblk"' in match.group(0) or 'reftype="ipl"' in match.group(0):
                return ref_pgblk(match) + match.group(2)
            elif 'reftype="vendlist"' in match.group(0) and vcl is not None:
                return vcl.dmref + match.group(2)
            else:
                return ref_int(match) + match.group(2)

        dm.content = re.sub(p_grphcref, link_graphic_references, dm.content)
        dm.content = re.sub(p_refint, delegate_refint, dm.content)

        # Replace tools, create header data
        dm.content = re.sub(p_tool, replace_tool, dm.content)
        for tool in local_tools:
            dm.tool_data += tool.sed

        # Replace tools, create header data
        dm.content = re.sub(p_cons, replace_cons, dm.content)
        for cons in local_cons:
            dm.cons_data += cons.sed

        dm.content = re.sub(re.compile(r'<gdesc>.*?</gdesc>', re.S), '', dm.content)

        dm.content = re.sub(r'<(/?)prcitem\d>', r'<\1proceduralStep>' if schema != "descript" else r'<\1levelledPara>', dm.content)
        dm.content = re.sub(r'<(/?)l\ditem>', r'<\1proceduralStep>' if schema != "descript" else r'<\1levelledPara>', dm.content)
        while True:
            fix1 = re.subn(p_fixsteps, r'\2\n\1', dm.content)
            dm.content = fix1[0]
            fix2 = re.subn(p_fixsteps_list, r'\2\n\1', dm.content)
            dm.content = fix2[0]
            fix3 = re.subn(p_fixsteps_tabfig, r'\2\n\1', dm.content)
            dm.content = fix3[0]
            if fix1[1] + fix2[1] + fix3[1] == 0:
                break

        dm.content = re.sub(p_cautwarn, replace_warn_and_caut, dm.content)
        dm.content = re.sub(p_cautwarn2, replace_warn_and_caut, dm.content)  # This one deals with warns/cauts at the top of the module

        dm.content = re.sub(re.compile(r'<(warning|caution)>(.*?)</\1>', re.S), replace_table_warn_caut, dm.content)
        for id in local_warn_caut:
            dm.warn_and_caut_data += get_warncaut_ref(id)

        if dm.parent is not None:
            if dm.parent.title == "Acronyms and Abbreviations":
                dm.content = create_acronym_table(dm.content)
        if "special tools" in dm.infoname.lower() and "Job Setup Data" in dm.content:
            dm.content = create_master_tools_and_cons_tables(dm.content)

    def fix_the_other_things(ipl_module_list):
        def fix_ids(match):
            key = match.group(1)
            if key in ipl_module_list:
                return re.sub(r'\n?<dmRef>', r'<dmRef referredFragment="{}">'.format(key), ipl_module_list[key])
            else:
                return match.group(0)

        try:  # Fix IDs
            if ipl_module_list is None:
                return
            # Get a list of all DMCs in the directory
            files = sorted(dir.glob('DMC*.xml'))

            # Process each file in turn
            for file in files:
                # Read content into a string
                text = file.read_text(encoding="utf-8")

                text = re.sub(p_refint, fix_ids, text)
                file.write_text(text, encoding="utf-8")
        except Exception:
            log_print(traceback.format_exc())

    def sort_list(lst):
        num_list = []
        for li in lst[::-1]:
            if li[0] != "":
                if li[0][0].isnumeric():
                    num_list.append(li)
                    lst.remove(li)
        lst.sort(key=lambda s: (s[0].lower(), s[1].lower()))
        if len(num_list) > 0:
            num_list.sort(key=lambda s: (s[0].lower(), s[1].lower()))
            lst = num_list + lst
        return lst

    def replace_tables(match, tool_table, cons_table):
        nonlocal tab
        t = tool_table if tab else cons_table
        orig_table_rows = match.group(0).count('<row>') - 1
        new_table_rows = t.count('<row>')
        if orig_table_rows > new_table_rows:
            return match.group(0)
        tgroup = """<tgroup cols="3"><colspec colname="col1"/><colspec colname="col2"/>
<colspec colname="col3"/><thead><row>
<entry>
<para>Number</para>
</entry>
<entry>
<para>Description</para>
</entry>
<entry>
<para>Source</para>
</entry>
</row></thead><tbody>{}</tbody></tgroup>""".format(t if t.strip() != "" else """<row><entry>
<para>Not applicable</para>
</entry>
<entry>
<para>Not applicable</para>
</entry>
<entry>
<para>Not applicable</para>
</entry></row>""")
        tab = False
        return tgroup

    def fill_tools_and_cons(content, st_content):
        nonlocal tab

        def create_tools_table(tools):
            nonlocal master_tools_list
            tool_list = []
            rows = ""
            for t in tools:
                cnt = re.sub(r'\n>', '>', t[1])
                cnt = re.sub(r'\n', ' ', cnt)
                pn = check_match(re.search(re.compile(r'<(?:tool|std)nbr>(.*?)</(?:tool|std)nbr>', re.S), cnt)).strip()
                nom = check_match(re.search(re.compile(r'<(?:tool|std)name>(.*?)</(?:tool|std)name>', re.S), cnt)).strip()
                src = check_match(re.search(re.compile(r'<(?:tool|std)src>(.*?)</(?:tool|std)src>', re.S), cnt)).strip()
                desc = check_match(re.search(re.compile(r'<(?:tool|std)desc>(.*?)</(?:tool|std)desc>', re.S), cnt)).strip()
                if (pn, nom, desc, src) in tool_list or nom == "":  # Check for duplicates
                    continue
                tool_list.append((pn, nom, desc, src))

            master_tools_list += tool_list
            tool_list = sort_list(tool_list)
            for t in tool_list:
                rows += "<row>\n<entry><para>{}</para></entry>\n<entry><para>{}{}</para></entry>\n<entry><para>{}</para></entry>\n</row>".format(t[0], t[1], (" ({})".format(t[2])) if t[2] != "" else "", t[3])
            return rows

        def create_cons_table(cons):
            nonlocal master_cons_list

            cons_list = []
            rows = ""
            for c in cons:
                cnt = re.sub(r'\n>', '>', c)
                cnt = re.sub(r'\n', ' ', cnt)
                pn = check_match(re.search(re.compile(r'<connbr>(.*?)</connbr>', re.S), cnt)).strip()
                nom = check_match(re.search(re.compile(r'<conname>(.*?)</conname>', re.S), cnt)).strip()
                src = check_match(re.search(re.compile(r'<consrc>(.*?)</consrc>', re.S), cnt)).strip()
                desc = check_match(re.search(re.compile(r'<condesc>(.*?)</condesc>', re.S), cnt)).strip()
                if (pn, nom, desc, src) in cons_list or nom == "":  # Check for duplicates
                    continue

                cons_list.append((pn, nom, desc, src))

            master_cons_list += cons_list
            cons_list = sort_list(cons_list)
            for c in cons_list:
                rows += "<row>\n<entry><para>{}</para></entry>\n<entry><para>{}{}</para></entry>\n<entry><para>{}</para></entry>\n</row>".format(c[0], c[1], (" ({})".format(c[2])) if c[2] != "" else "", c[3])
            return rows

        if '<title>Index of Repairs</title>' in st_content:
            return st_content

        tools = p_tool2.findall(content)  # Find tools
        tool_table = create_tools_table(tools)  # Create tool table
        cons = p_cons2.findall(content)  # Find consumables
        cons_table = create_cons_table(cons)  # Create consumables table
        tab = True
        return re.sub(p_tgroup, partial(replace_tables, tool_table=tool_table, cons_table=cons_table), st_content, 2)  # Replace 2 tables in the data module with the tools and consumables tables

    def create_master_tools_and_cons_tables(dm_content):
        nonlocal master_cons_list
        nonlocal master_tools_list
        nonlocal tab

        tool_table = ""
        master_tools_list = list(dict.fromkeys(master_tools_list))
        master_tools_list = sort_list(master_tools_list)
        for t in master_tools_list:
            tool_table += "<row>\n<entry><para>{}</para></entry>\n<entry><para>{}{}</para></entry>\n<entry><para>{}</para></entry>\n</row>".format(t[0], t[1], (" ({})".format(t[2])) if t[2] != "" else "", t[3])

        cons_table = ""
        master_cons_list = list(dict.fromkeys(master_cons_list))
        master_cons_list = sort_list(master_cons_list)
        for c in master_cons_list:
            cons_table += "<row>\n<entry><para>{}</para></entry>\n<entry><para>{}{}</para></entry>\n<entry><para>{}</para></entry>\n</row>".format(c[0], c[1], (" ({})".format(c[2])) if c[2] != "" else "", c[3])

        tab = True
        return re.sub(p_tgroup, partial(replace_tables, tool_table=tool_table, cons_table=cons_table), dm_content, 2)

    def create_acronym_table(dm_content):
        nonlocal acros_and_abbrs
        sorted_list = sorted(acros_and_abbrs, key=lambda s: s[0].lower())
        table = """<table frame="topbot" id="t001">
    <title>List of Acronyms and Abbreviations</title><?Pub Caret -1?>
    <tgroup cols="2">
    <?PubTbl tgroup dispwid="605.00px"?>
    <colspec colname="col1" colsep="0" colwidth="0.39*"/><colspec colname="col2" colwidth="1.61*"/><tbody><row>
    <entry>
    <para>Term</para>
    </entry>
    <entry>
    <para>Full Term</para>
    </entry>
    </row>"""
        for a in sorted_list:
            table += """<row>
    <entry rowsep="0">
    <para>{}</para>
    </entry>
    <entry rowsep="0">
    <para>{}</para>
    </entry>
    </row>""".format(a[0], a[1])
        table += "</tbody></tgroup></table></levelledPara>"
        return re.sub(r'</levelledPara><levelledPara><para>AA WILL BE GENERATED</para>\n</levelledPara>', table, dm_content)
        print("</body>\n</html")

    def create_dm(dm):
        try:
            # Decide whether to create a Descriptive or Procedural module
            if page_blocks.get(dm.parent.pgblk, ("UNIDENTIFIED PAGEBLOCK", "pmt01", "000", "D"))[3] == "D" and dm.parent.title != "Planning Data" and (("<caution>" not in dm.content and "<warning>" not in dm.content) or dm.parent.pgblk == "0"):
                process_data_module(dm, "descript")
                module = create_descript(dm)
            else:
                process_data_module(dm, "proced")
                module = create_proced(dm)

            (dir / dm.filename).write_text(module, encoding='utf-8')
        except Exception:
            # log_print("Failed to create {} data module.".format(dm.filename))
            log_print(traceback.format_exc())

    acros_and_abbrs = []
    book_info = {}  # Contains metadata about the book
    tasks = []
    not_applicable = []
    front_matter = []
    vcl = None
    warn = {}
    caut = {}
    is_irm = False
    tab = True
    master_cons_list = []
    master_tools_list = []

    try:

        try:  # Process the source data before reading it
            print("\nProcessing source data...", end='\r')
            source = process_source(source_file)
        except Exception:
            log_print("Processing source data... Error!")
            print(traceback.format_exc())
        else:
            log_print("Processing source data... Done!")

        try:
            print("\nGetting book info...", end='\r')
            book_info = GetBookInfo(source)
        except Exception:
            log_print("Getting book info... Error!")
            print(traceback.format_exc())
        else:
            log_print("Getting book info... Done!")

        try:
            print("\nProcessing data modules...", end='\r')
            pmc = create_pmc(GetPMCData())
        except Exception:
            log_print("Processing data modules... Error!")
            print(traceback.format_exc())
        else:
            log_print("Processing data modules... Done!")

        try:
            print("\nCreating data modules...", end='\r')
            # Create the Data Modules
            for t in tasks:
                for dm in t.children:
                    create_dm(dm)

            # Create Front Matter Data Modules
            for f in front_matter:
                try:
                    process_data_module(f, "descript")
                    (dir / f.filename).write_text(create_descript(f), encoding='utf-8')
                except Exception:
                    log_print("Failed to create {} data module".format(f.filename))

            # Create N/A Data Modules
            for n in not_applicable:
                try:
                    (dir / n.filename).write_text(create_proced(n), encoding='utf-8')
                except Exception:
                    log_print("Failed to create {} data module".format(n.filename))
            try:
                if vcl is not None:
                    (dir / vcl.filename).write_text(create_descript(vcl, add_levelled_para=False), encoding='utf-8')
            except Exception:
                log_print("Failed to create 018Z data module")

            # Create CIRs
            try:
                (dir / "DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "00NA")).write_text(create_cir("tool", ToolAndCons.tools), encoding='utf-8')
            except Exception:
                log_print("Failed to create 00NA data module")

            try:
                (dir / "DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "00LA")).write_text(create_cir("cons", ToolAndCons.cons), encoding='utf-8')
            except Exception:
                log_print("Failed to create 00LA data module")

            try:
                (dir / "DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "012A")).write_text(create_cir("warn", warn), encoding='utf-8')
            except Exception:
                log_print("Failed to create 012A data module")

            try:
                (dir / "DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "012B")).write_text(create_cir("caut", caut), encoding='utf-8')
            except Exception:
                log_print("Failed to create 012B data module")

            try:
                # Create 00P
                filename_00P = "DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "00PA")
                (dir / filename_00P).write_text(create_prdcrossreftable(filename_00P), encoding='utf-8')
            except Exception:
                log_print("Failed to create 00PA data module")

            try:
                # Create 00W
                filename_00W = "DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "00WA")
                (dir / filename_00W).write_text(create_appliccrossreftable(filename_00W), encoding='utf-8')
            except Exception:
                log_print("Failed to create 00WA data module")
        except Exception:
            log_print("Creating data modules... Error!")
            print(traceback.format_exc())
        else:
            log_print("Creating data modules... Done!")

        try:
            print("\nCreating PMC...", end='\r')

            def to_lower(match):
                """Replace string with lowercase string"""
                return match.group(0).lower()
            pmc = re.sub(r'&([A-Z]+);', to_lower, pmc)
            (dir / 'PMC-HON{0}-{0}-00001-01_sx-US.XML'.format(book_info['cage'])).write_text(pmc, encoding='utf-8')
        except Exception:
            log_print("Creating PMC... Error!")
            log_print(traceback.format_exc())
        else:
            log_print("Creating PMC... Done!")
    except Exception:
        log_print(traceback.format_exc())

    try:
        log_print("\nConverting IPL: ")
        ipl_module_list, ipl_conv_log = DoTheThing(source, dir, book_info, cc)  # Convert the IPL. This is in another module
        log_print(ipl_conv_log, True)
        fix_the_other_things(ipl_module_list)  # Convert the links referencing IPL Figures
    except Exception:
        print(traceback.format_exc())

    log_print("Done!")
    # If we reach this point, run the exit function
    return exit_handler(), cc


if __name__ == "__main__":
    CMMConvert(get_source(), Path('.'), None)
