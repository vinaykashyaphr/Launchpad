import ntpath
import re
import sys
import atexit
import importlib
from msvcrt import getch
import traceback
from pathlib import Path
from openpyxl import Workbook, load_workbook
from functools import partial
from collections import defaultdict
import shutil

try:
    from launchpad.functions import ConversionClient
except ImportError:
    pass

#Source processing
pattern_ftnote = re.compile(r'<ftnote.*?</ftnote>', re.DOTALL)
pattern_ftnotesub = re.compile(r'</tbody>.*?</ftnote>(?=\s?</table>)', re.DOTALL)
pattern_techname = re.compile(r'<techName>[^<]+</techName>')
pattern_note = re.compile(r'<note>.*?</note>', re.DOTALL)
pattern_entity = re.compile(r'&(.+?);', re.DOTALL)
pattern_rev = re.compile(r"<\?Pub\s?\n?/?_rev\n?\?>\n?")
pattern_entityref = re.compile(r'<!ENTITY ([^\s]+) "([^"]+)">')
pattern_grphcref = re.compile(r'<grphcref.*?refid="([^"]+).*?</grphcref>', re.DOTALL)
pattern_refint = re.compile(r'<refint.*?refid="([^"]+)".*?</refint>', re.DOTALL)
pattern_rht = re.compile(r'<\?PubTbl row rht[^>]+>\n?')
pattern_para = re.compile(r"<para>(.*?)</para>", re.DOTALL)
pattern_table = re.compile(r'<table(.*?)</table>', re.DOTALL)
pattern_entry = re.compile(r'(<entry[^>]*>(.*?)</entry>)', re.DOTALL)
pattern_fixsteps_tabfig = re.compile(r'((?:</(?:proceduralStep|levelledPara)>\n?)*</(?:proceduralStep|levelledPara)>)\n*(<(table|figure).*?</\3>)', re.DOTALL) 
pattern_fixsteps_list = re.compile(r'((?:</(?:proceduralStep|levelledPara)>\n?)*</(?:proceduralStep|levelledPara)>)\n*(<para><(?:random|sequential)List>.*?</(?:random|sequential)></para>)', re.DOTALL)
pattern_fixsteps = re.compile(r'((?:</(?:proceduralStep|levelledPara)>\n?)*)<(?:proceduralStep|levelledPara)>\n*(<(?:table|para><(?:random|sequential)List>|figure).*?)</(?:proceduralStep|levelledPara)>', re.DOTALL)
#pattern_equ = re.compile(r'<equ>(.*?)</equ>(.*?)(?=<(?:/para|equ)>)', re.DOTALL)
pattern_equ = re.compile(r'(<para>\s*)?<equ>(.*?)</equ>([^<]*?)(?=</para>)', re.DOTALL)
pattern_colspec = re.compile(r'(<colspec[^/>]*>\s?)(?!</colspec>)')
pattern_subtitle = re.compile(r'<prcitem>\n?<title>(.*?)</title>', re.DOTALL)
pattern_graphic = re.compile(r'<graphic(.*?<title>(.*?)</title>.*?)</graphic>', re.DOTALL)
pattern_fig = re.compile(r'<figure(.*?<graphic.*?<title>(.*?)</title>.*?)</graphic>', re.DOTALL)
pattern_sheet = re.compile(r'<sheet.*?gnbr="([^"]+)', re.DOTALL)
pattern_sub_super = re.compile(r'<(sub|super)>([^<]+)</\1>')
pattern_sgm_attr = re.compile(r'(?:chapnbr|sectnbr|subjnbr|func|seq|confltr|varnbr|confnbr|pgblknbr|scalefit|sheetnbr|reprowid|reprodep|formal)="[^"]*"\s?')
pattern_refext = re.compile(r'</?refext[^>]*>', re.DOTALL)

#SBL, RTR, VCL TOI, Prelim
pattern_proptary = re.compile(r'<fullstmt>(.*?)</fullstmt>', re.DOTALL)
pattern_vendata = re.compile(r"(<vendata>.*?</vendata>)", re.DOTALL)
pattern_v = re.compile(r"row rowsep=\"0\">(\n)?<entry><para>([^V])")
pattern_mfrcont = re.compile(r"</?mfrcont>")
pattern_sblist = re.compile(r"<sblist.*?</sblist>", re.DOTALL)
pattern_sbdata = re.compile(r"(<sbdata>.*?</sbdata>)", re.DOTALL)
pattern_sbdatatag = re.compile(r"</?sbdata>")
pattern_trlist = re.compile(r"<trlist.*?</trlist>", re.DOTALL)
pattern_trdata = re.compile(r"<trdata>.*?</trdata>", re.DOTALL)
pattern_transltr = re.compile(r'<transltr[^>]*>.*?<title>[^<]*</title>\n?(.*?)</transltr>', re.DOTALL)

#Book metadata
pattern_cage = re.compile(r'spl="(\w{5})"')
pattern_tsn = re.compile(r'tsn="(\d+)"')
pattern_oidate = re.compile(r'oidate="(\d{8})"')
pattern_revdate = re.compile(r'revdate="(\d{8})"')
pattern_eccn = re.compile(r'ECCN:[^<]+')
pattern_copyright = re.compile(r'<geninfo>.*?((?:\d{4}, )\d{4})', re.DOTALL)
pattern_docnbr = re.compile(r'docnbr="([^"]+)')
pattern_model = re.compile(r'model="([^"]+)')
pattern_cmpnom = re.compile(r'<partinfo[^>]+>\n?<title>(.*?)</title>', re.DOTALL)

#PM Structure
pattern_pgblk = re.compile(r'<(pgblk|ipl)([^>]+)>\n?(?:<effect>(.*?)</effect>)?\n?(?:<title>(.*?)</title>)?(.*?)</\1>', re.DOTALL)
#pattern_tasks = re.compile(r'<task(?:.*?func="(\w+)")?.*?key="([^"]+)"[^>]*>.*?<title>(.*?)</title>(.*?)</task>', re.DOTALL)
pattern_tasks = re.compile(r'<task([^>]+)>.*?<title>(.*?)</title>(.*?)</task>', re.DOTALL)
pattern_subtask = re.compile(r'<subtask([^>]+)>(.*?)</subtask>', re.DOTALL)
pattern_nonpgblk = re.compile(r'<(/?)(?:trlist|sblist|transltr)', re.DOTALL)
pattern_vendlist_tag = re.compile(r'<(/?)vendlist', re.DOTALL)

#Repository stuff
pattern_tool = re.compile(r'<(std|ted)>.*?</\1>', re.DOTALL)
pattern_cons = re.compile(r'<con>.*?</con>', re.DOTALL)
pattern_cautwarn = re.compile(r'()<((?:proceduralStep|levelledPara)[^>]*)>\n*<(?:caution|warning)>.*?</(?:caution|warning)>(?!\n*<(?:warning|caution)>)\n*', re.DOTALL)
pattern_cautwarn2 = re.compile(r'(?<=</title>)\n*<(?:caution|warning)>.*?</(?:caution|warning)>\n?(<note>.*?</note>\n?)?<((?:proceduralStep|levelledPara)[^>]*)>', re.DOTALL)

#00P
pattern_partinfo = re.compile(r'<partinfo([^>]+)>(.*?)</partinfo>', re.DOTALL)
pattern_mfrpnr = re.compile(r'<mfrpnr>(.*?)</mfrpnr>', re.DOTALL)
pattern_mfr = re.compile(r'<mfr>(.*?)</mfr>')
pattern_pnr = re.compile(r'<pnr>(.*?)</pnr>')

#EM stuff
pattern_chapnbr = re.compile(r'chapnbr="([^"]+)"')
pattern_sectnbr = re.compile(r'sectnbr="([^"]+)"')
pattern_subjnbr = re.compile(r'subjnbr="([^"]+)"')
pattern_chapter = re.compile(r'<chapter([^>]+)>\n?<title>(.*?)</title>(.*?)</chapter>', re.DOTALL)
pattern_section = re.compile(r'<section([^>]+)>\n?<title>(.*?)</title>(.*?)</section>', re.DOTALL)
pattern_subject = re.compile(r'<subject([^>]+)>\n?<title>(.*?)</title>(.*?)</subject>', re.DOTALL)
pattern_title = re.compile(r'<title>(.*?)</title>', re.DOTALL)
#pattern_sh_eff = re.compile(
pattern_effect = re.compile(r'<effect efftext="([^"]+)"[^<]+</effect>\n*<title>(.*?)</title>', re.DOTALL)
pattern_effect2 = re.compile(r'(?:(<graphic[^>]*>)\n?(?:<title>.*?</title>\n?)?)?(<sheet[^>]*key="([^"]*)"[^<]+)<effect efftext="([^"]+)"[^<]+</effect>(.*?</sheet>)', re.DOTALL)

#Tools and Cons table
pattern_tbody = re.compile(r'<tbody>.*?</tbody>', re.DOTALL)
pattern_tool2 = re.compile(r'<(std|ted)>(.*?)</\1>', re.DOTALL)
pattern_cons2 = re.compile(r'<con>(.*?)</con>', re.DOTALL)

pattern_split_graphic = re.compile(r'(</table>\n?</gdesc>\n?</sheet>\n?)(<sheet[^>]*key="([^"]*)"[^>]*>(?!(?:</table>\n?)?</gdesc>).*?)', re.DOTALL)
pattern_title = re.compile(r'<title>(.*?)</title>', re.DOTALL)
pattern_id = re.compile(r'id="([^"]*)"', re.DOTALL)

months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
page_blocks = {
"0" : ("INTRODUCTION", "pmt58", "018", "D"), #Not a page block
"1" : ("SYSTEM DESCRIPTION", "pmt59", "040", "D"),
"1000" : ("FAULT ISOLATION", "pmt60", "420", "P"),
"2000" : ("MAINTENANCE PRACTICES", "pmt83", "913", "P"),
"3000" : ("REMOVAL", "pmt84", "520", "P"),
"4000" : ("INSTALLATION", "pmt85", "720", "P"), 
"5000" : ("DISASSEMBLY", "pmt86", "500", "P"), 
"6000" : ("CLEANING", "pmt87", "250", "P"), 
"7000" : ("", "pmt88", "910", "P"),
"8000" : ("INSPECTION/CHECK", "pmt89", "300", "P"),
"9000" : ("REPAIR", "pmt68", "600", "P"),
"10000" : ("ASSEMBLY", "pmt76", "700", "P"),
"11000" : ("SERVICING", "pmt69", "200", "P"),
"12000" : ("STORAGE", "pmt70", "800", "P"),
"13000" : ("TESTING", "pmt71", "400", "P"),
"14000" : ("", "pmt72", "910", "P"),
"15000" : ("", "pmt73", "910", "P"),
"16000" : ("", "pmt74", "910", "P")
}

class Publication():
    how_to = False

    def lookup_code_var(self, infocode, generic):
        if not generic:
            code_var = self.info_codes[infocode][0]
            if self.info_codes[infocode][0] == 'Z':
                self.info_codes[infocode][0] = 'A'
                self.info_codes[infocode][1] += 1
            else:
                self.info_codes[infocode][0] = chr(ord(self.info_codes[infocode][0]) + 1)
            
            return code_var
        else:
            code_var = self.generic_info_codes[infocode][0]
            if self.generic_info_codes[infocode][0] == 'Z':
                self.generic_info_codes[infocode][0] = 'A'
                self.generic_info_codes[infocode][1] += 1
            else:
                self.generic_info_codes[infocode][0] = chr(ord(self.generic_info_codes[infocode][0]) + 1)
            return code_var

    def __init__(self):
        self.info_codes = defaultdict(lambda:['A', 1])
        self.generic_info_codes = defaultdict(lambda:['A', 0])
    # 	self.info_codes = {"018": ['A', 1], "040" : ['A', 1], "100":  ['A', 1], "400":  ['A', 1], "051":  ['A', 1], "500":  ['A', 1], "200":  ['A', 1],  "300":  ['A', 1],  "600":  ['A', 1],  "700":  ['A', 1], "711":  ['A', 1], "710": ['A', 1], \
    # "900":  ['A', 1], "520":  ['A', 1],  "720":  ['A', 1], "250":  ['A', 1], "800":  ['A', 1], "664":  ['A', 1], "910" : ['B', 1], "913" : ['A', 1], "420" : ['A', 1]} 
    
    # 	self.generic_info_codes = {"018": ['A', 0], "040" : ['A', 0], "100": ['A', 0], "400": ['A', 0], "051": ['A', 0], "500": ['A', 0], "200": ['A', 0],  "300": ['A', 0],  "600": ['A', 0],  "700": ['A', 0], "711": ['A', 0], "710": ['A', 0], \
    # "900": ['A', 0], "520": ['A', 0],  "720": ['A', 0], "250": ['A', 0], "800": ['A', 0], "664": ['A', 0], "910": ['B', 0], "913" : ['A', 0], "420" : ['A', 0]}

class ToolAndCons():
    tools = []
    cons = []
    
    def get_cir_id(self, list):
        for item in list:
            if item.pn == self.pn:
                if item.nom == self.nom:
                    self.id = item.id
                    break
        else:
            self.id = "{}-{}".format(self.type, add_leading_zero(len(list) + 1, 4))
            list.append(self)
    
    def create_sed(self, type):
            return """<{10}Descr id="{0}"><name>{1}{2}{8}</name>
<shortName>{1}{2}</shortName>
<{9}Ref {9}Number="{0}"{11}>
<refs>
<dmRef><dmRefIdent><dmCode assyCode="{3}" disassyCode="00"
disassyCodeVariant="A" infoCode="00{12}" infoCodeVariant="A"
itemLocationCode="D" modelIdentCode="HON{4}" subSubSystemCode="{5}"
subSystemCode="{6}" systemCode="{7}" systemDiffCode="EAA"/></dmRefIdent>
</dmRef>
</refs>
</{9}Ref>
<reqQuantity>AR</reqQuantity>
</{10}Descr>""".format(self.id, "{}, ".format(self.pn) if self.pn != "" else "", self.nom, self.book_info['ata'][6:8], self.book_info['cage'], self.book_info['ata'][4], self.book_info['ata'][3], self.book_info['ata'][0:2], " ({})".format(self.desc) if self.desc != "" else "", type if type == "tool" else "supply", "supportEquip" if type == "tool" else "supply", ' supplyNumberType="sp01"' if type == "cons" else "", "L" if type == "cons" else "N")
    
    def create_spec(self, type):
        if type == "cons":
            return """<supplySpec><supplyIdent id="{0}" supplyNumber="{0}" supplyNumberType="sp01"/><name>{1}</name><shortName>{2}{1}</shortName></supplySpec>\n""".format(self.id, self.nom, "{} ".format(self.pn) if self.pn != "" else "")
        else:
            return """<toolSpec><toolIdent id="{0}" manufacturerCodeValue="{3}" toolNumber="{0}"/><itemIdentData><descrForPart>{1}</descrForPart><shortName>{2}{1}</shortName></itemIdentData><procurementData></procurementData><techData></techData><toolAlts><tool><itemDescr></itemDescr></tool></toolAlts></toolSpec>\n""".format(self.id, self.nom, "{} ".format(self.pn) if self.pn != "" else "", self.src)
            
    def __init__(self, pn, nom, src, desc, type, book_info):
        self.book_info = book_info
        self.pn = pn
        self.nom = nom.strip()
        self.src = src
        self.type = type
        self.desc = desc
        self.get_cir_id(ToolAndCons.tools if type == "tool" else ToolAndCons.cons)
        self.sed = self.create_sed(type)
        self.spec = self.create_spec(type)

class Task():
    def __init__(self, parent_key, key, authdoc, pgblk, title, subj_key):
        self.parent_key = parent_key #pgblk key?
        self.key = key
        if authdoc is not None:
            self.authdoc = authdoc if "--" not in authdoc else ""
        else:
            self.authdoc = ""
        self.subj_key = subj_key
        self.title = title
        self.children = []
        self.pgblk = pgblk
        self.subj_num = check_match(re.search('\d\d-\d\d-\d\d', authdoc), 0)

class DataModule(): #args[content, key, func, parent, infocode, pgblk, not_applicable=False, generic=False, filename=None, infoname=None]
    @staticmethod
    def get_module_name(dmref):
        try:
            dmref = re.sub(r' authorityDocument="[^"]+"', '', dmref)
            vars = re.split(' |\n', dmref)
            vars = [var.split('\"')[1] for var in vars if(len(var.split('\"')) >= 2)]
            dmc = "DMC-%s-%s-%s-%s%s-%s-%s%s-%s%s-%s_sx-US.XML" % (vars[6].upper(), vars[10], vars[9], vars[8], vars[7], vars[0], vars[1], vars[2], vars[3], vars[4], vars[5])
            return dmc
        except Exception as e:
            log_print("Error while parsing dmRef:")
            log_print(traceback.format_exc())
            return None
    
    def get_dmref(self, filename=None):
        if filename is None:
            info_var = self.pub.lookup_code_var(self.infocode, self.generic)
            if self.generic == False:
                disassy = add_leading_zero(self.pub.info_codes[self.infocode][1], 2)
                return """<dmRef{5}><dmRefIdent><dmCode assyCode="{0}" disassyCode="{8}"
    disassyCodeVariant="A" infoCode="{6}" infoCodeVariant="{7}"
    itemLocationCode="C" modelIdentCode="HON{1}" subSubSystemCode="{2}"
    subSystemCode="{3}" systemCode="{4}" systemDiffCode="EAA"/></dmRefIdent>
    </dmRef>""".format(\
                self.parent.subj_num[6:8], \
                self.book_info['cage'], \
                self.parent.subj_num[4], \
                self.parent.subj_num[3], \
                self.parent.subj_num[0:2], \
                (' authorityDocument="{}"'.format(self.authdoc)) if self.authdoc != "" else "", \
                self.infocode, \
                info_var, \
                disassy)
            else:
                disassy = add_leading_zero(self.pub.generic_info_codes[self.infocode][1], 2)
                return """<dmRef{5}><dmRefIdent><dmCode assyCode="{0}" disassyCode="{8}"
    disassyCodeVariant="A" infoCode="{6}" infoCodeVariant="{7}"
    itemLocationCode="C" modelIdentCode="HON{1}" subSubSystemCode="{2}"
    subSystemCode="{3}" systemCode="{4}" systemDiffCode="EAA"/></dmRefIdent>
    </dmRef>""".format(\
                "00", \
                "AERO", \
                "0", \
                "0", \
                "00", \
                (' authorityDocument="{}"'.format(self.authdoc)) if self.authdoc != "" else "", \
                self.infocode, \
                info_var, \
                disassy)
        else:
            if self.authdoc != "":
                return re.sub('<dmRef>', '<dmRef authorityDocument="{}">'.format(self.authdoc), get_dmref(filename))
            else:
                return get_dmref(filename)
                
    def __init__(self, content, key, authdoc, parent, infocode, pgblk, book_info, not_applicable=False, generic=False, filename=None, infoname=None, authName=None, pub=None): #TODO: Pass parent info? Or create parent class containing child DataModule objects
        self.book_info = book_info
        self.parent = parent
        self.pub = pub
        self.content = content if not_applicable == False else "<title>Not Applicable</title>"
        self.key = key
        if authdoc is not None:
            self.authdoc = authdoc if "--" not in authdoc else ""
        else:
            self.authdoc = ""
        self.infocode = infocode
        if infocode == "018" and pub.how_to == False:
            self.filename = "DMC-HONAERO-EAA-{}-{}-{}-01A-018A-C_sx-US.XML".format(book_info['ata'][0:2], book_info['ata'][3:5], book_info['ata'][6:8])
            self.generic = False
            self.dmref = self.get_dmref(self.filename)
            pub.how_to = True
        else:
            self.generic = True if infocode == "018" else generic	
            if filename is None:
                self.dmref = self.get_dmref()
                self.filename = DataModule.get_module_name(self.dmref)
            else:
                self.filename = filename
                self.dmref = self.get_dmref(filename)
        self.table_refs = re.findall(re.compile(r'<table.*?id="([^"]+)', re.DOTALL), self.content)
        
        figures = re.findall(re.compile(r'<figure.*?id="([^"]+)(.*?)</figure>', re.DOTALL), self.content)
        self.graph_refs = [i[0] for i in figures]
        self.sheet_refs = {}
        for f in figures:
            sheets = re.findall(re.compile(r'<graphic.*?id="([^"]+)', re.DOTALL), f[1])
            for i,s in enumerate(sheets):
                self.sheet_refs[s] = (f[0], i+1) #Sheet associated with a figure id and a sheet number

        self.tool_data = ""
        self.cons_data = ""
        self.warn_and_caut_data = ""
        self.infoname = page_blocks[pgblk][0] if infoname is None else re.sub(r'&([A-Z]+);', to_lower, infoname)
        title = check_match(pattern_title.search(self.content))
        title = re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.DOTALL), r'\1', title)
        self.authname = 'Paragraph {}{}{}'.format(authName, ". " if title != "" else "", re.sub(r'^\((Pre|Post) SB [^)]+\) ', '', title))

def exit_handler_partial(logText, status=0, filename="log.txt", cc=None):
    if cc is not None:
        cc.exit_handler(status)
    elif len(logText) > 0:
        try:
            log = ""
            for l in logText:
                log += l + '\n'
            Path(filename).write_text(log.strip('\n'))
        except:
            print(traceback.format_exc())
            if status == 0:
                status = 1
        finally:
            input("Press any key to continue")
            sys.exit(status)
    else:
        input("Press any key to continue")
        sys.exit(status)

def log_print_partial(logText, message, printToLogOnly=False, cc=None):
    if cc is not None:
        cc.log_print(message, printToLogOnly)
    else:
        if(not printToLogOnly):
            print(message)	
        logText.append(str(message))

#Gets input using Launchpad or generic method
def get_input_partial(message, default=None, cc=None):
    if not cc:
        return input(message)
    else:
        return cc.get_user_input(message, default)

def get_dmref(filename):
    components = re.split('-', filename)
    try:
        comp_dict = {
                "model": components[1],
                "diff": components[2],
                "sys": components[3],
                "subsys": components[4][0],
                "subsub": components[4][1],
                "assy": components[5],
                "disassy": components[6][0:2],
                "disvar" : components[6][2],
                "info" : components[7][0:3],
                "infovar" : components[7][3],
                "loc" : components[8].split('_')[0][0]
            }
    except IndexError:
        print(f"IndexError parsing {filename} to dmRef")
        
    return """\n<dmRef><dmRefIdent><dmCode assyCode=\"%s\" disassyCode=\"%s\"
disassyCodeVariant=\"%s\" infoCode=\"%s\" infoCodeVariant=\"%s\"
itemLocationCode=\"%s\" modelIdentCode=\"%s\" subSubSystemCode=\"%s\"
subSystemCode=\"%s\" systemCode=\"%s\" systemDiffCode=\"%s\"/></dmRefIdent>
</dmRef>""" % (comp_dict["assy"], comp_dict["disassy"], comp_dict["disvar"], comp_dict["info"], \
    comp_dict["infovar"], comp_dict["loc"], comp_dict["model"], comp_dict["subsub"], \
    comp_dict["subsys"], comp_dict["sys"], comp_dict["diff"])
    
def to_lower(match):
    return match.group(0).lower()

def check_match(match, group_num = 1):
        if match is not None:
            return(match.group(group_num))
        else:
            return("")
    
def add_leading_zero(number, digits):
    return "0" * (digits - len(str(number))) + str(number)	

def get_source(): #Get the source file
        source = list(Path('.').glob('*.sgm'))
        if len(source) == 1:
            print("Found source file.")
            return source[0]
        else:
            print(">>>Please input the Source file:")
            source = Path(input())
            while not source.is_file():
                print(">>>File does not exist. Please try again:")
                source = Path(input())
            return source
    
def EMConvert(file, directory, cc):
    pub = Publication()
    log_file_name = "em_converter_log.txt"
    if cc is None:
        print("Warning: Launchpad Not Found")
    else:
        cc.log_name = log_file_name
    
    logText = []
    errors = False
    
    #Define partials. Makes function calls a little cleaner so cc doesn't need to be passed in each call.
    log_print = partial(log_print_partial, logText, cc=cc)
    get_input = partial(get_input_partial, cc=cc)
    exit_handler = partial(exit_handler_partial, logText, filename=log_file_name, cc=cc)

    acros_and_abbrs = []
    book_info = {} #Contains metadata about the book
    tasks = []
    not_applicable = []
    front_matter = []
    introduction = []
    intro_key = None
    vcl = None
    warn = {}
    caut = {}
    tab = True
    # user = username

    def process_bluesheet(source):
        # nonlocal dir
        # cage = check_match(pattern_cage.search(source))
        pbs = pattern_pgblk.findall(source)
        shutil.copyfile(str(Path('./launchpad/XML_Conv/BlueSheet Template.xlsx').resolve()), str(dir / 'BlueSheet.xlsx'))
        new_workbook = load_workbook(str(dir / 'BlueSheet.xlsx'))
        ws = new_workbook.worksheets[1]
        bluesheet_data = []
        for p in pbs:
            gph = pattern_graphic.findall(p[4])
            if p[0] == "ipl":
                gph += pattern_fig.findall(p[4])
            for i,g in enumerate(gph):
                sht = pattern_sheet.findall(g[0])
                for j,s in enumerate(sht):
                    id = s
                    #icn = ("ICN-{}-0000{}-001-01".format(id, s)) if id != "" else ""
                    
                    for b in bluesheet_data:
                        if id == b[1]:
                            break
                    else:
                        fignbr = check_match(re.search(r'fignbr="([^"]+)"', g[0]))
                        title = re.sub(re.compile(r'<acronym>.*?<acronymTerm>([^<]+).*?</acronym>', re.DOTALL), r'\1', g[1])
                        title = re.sub('\n', ' ', title)
                        d = ["I", id, "", "", (fignbr if fignbr != "" else str(i+1)) + ((",SH%s" % (j + 1)) if len(sht) > 1 else ""), title, "A"]
                        bluesheet_data.append(d)
                        ws.append(d)
        ws = new_workbook.worksheets[0]
        if cc is not None:
            user_info = cc.get_user_info()
            if user_info is not None:
                ws['B1'] = user_info.fname + " " + user_info.lname
                ws['B2'] = user_info.email

        ws['B4'] = check_match(re.search(r'(?s)<em[^>]+?docnbr="([^"]+)"', source))
        ws['B5'] = check_match(re.search(r'\b\d{2}-\d{2}-\d{2}\b', source), 0)
        ws['B6'] = re.sub(r'(?s)<acronym>.*?<acronymTerm>([^<]+).*?</acronym>', r'\1', check_match(re.search(r'<partinfo[^>]*>\s+<title>(.*?)</title>', source)))
        ws['B7'] = check_match(re.search(r'(?s)<em[^>]+?spl="([^"]+)"', source))
        ws['B8'] = check_match(re.search(r'<!ENTITY ECCN\w+ "([^"]+)">', source)) or check_match(re.search(r'ECCN:?\s?(\w+)', source))
        ws['B9'] = check_match(re.search(r'(?s)<em[^>]+?type="([^"]+)"', source)).upper()

        new_workbook.save(str(dir / 'BlueSheet.xlsx'))
        return source

    def process_source(source_file): #Do series of find/replaces to remove/modify ATA2200 tags -> S1000D tags
        
        def SubAcro(match):
            content = match.group(0)
            content = re.sub('(acro|abbr)', 'acronym', content)
            content = re.sub('acronymterm', 'acronymTerm', content)
            content = re.sub('acronymname', 'acronymDefinition', content)
            content = re.sub(r'(?s)<acronymTerm>\s*(.*?)\s*</acronymTerm>', r'<acronymTerm>\1</acronymTerm>', content)
            #Add to our list of Acronyms and Abbreviations
            term = check_match(re.search('<acronymTerm>([^<]+)</acronymTerm>', content)).strip()
            defn = check_match(re.search('<acronymDefinition>([^<]+)</acronymDefinition>', content)).replace('\n', ' ').strip()
            for a in acros_and_abbrs:
                if a[0] == term and a[1] == defn:
                    break
            else:
                acros_and_abbrs.append([term, defn])	
            return content

        def NotePara(match):
            note = re.sub('para>', 'notePara>', match.group(0))
            note = re.sub(r'<unlist[^>]*>', r'<notePara>\n<attentionRandomList>', note)
            note = re.sub(r'</unlist>', r'</attentionRandomList></notePara>', note)
            note = re.sub(r'<numlist[^>]*>', r'<notePara><attentionSequentialList>', note)
            note = re.sub(r'</numlist>', r'</attentionSequentialList></notePara>', note)
            note = re.sub(r'<(/?)unlitem[^>]*>', r'<\1attentionRandomListItem>', note)
            note = re.sub(r'<(/?)numlitem[^>]*>', r'<\1attentionSequentialListItem>', note)
            note = re.sub(r'\n?<(attentionRandomListItem|attentionSequentialListItem)>\n?<notePara>', r'\n<\1><attentionListItemPara>', note)
            note = re.sub(r'</notePara>\n?</(attentionRandomListItem|attentionSequentialListItem)>\n?', r'</attentionListItemPara></\1>\n', note)
            return note

        def replace_entities(text):
            header_entities = pattern_entityref.findall(text) #Get entities from the header of the source file
            
            #If the MasterTextEntities.xml file is present, concatenate the entities from there to the other header entities to make one long list
            master_entities_file = Path('launchpad/XML_Conv/MasterTextEntities.xml')
            if master_entities_file.is_file():
                master_entities =  pattern_entityref.findall(master_entities_file.read_text(encoding='utf-8'))
                #master_entities_file.unlink()
            else:
                master_entities = []
                
            header_entities += master_entities
            
            #Sub out the entities in source
            for e in header_entities:
                text = re.sub('&%s;' % e[0], e[1], text)
            return text
            
        def replace_graphics(match):
            def replace_gdesc(match):
                gdesc = match.group(2)
                dim = ""
                legend = ""
                notes = ""
                for table in pattern_table.finditer(gdesc):
                    table = table.group(0)
                    if "<title>" in table and "<title>Key" not in table:
                        id = check_match(pattern_id.search(match.group(1)))
                        table = re.sub(r'<title>Dimensional\sLimits</title>', r'<title>Dimensional Limits for <internalRef internalRefId="{}" internalRefTargetType="irtt09"></internalRef></title>'.format(id), table)
                        dim = table
                    else:
                        table = re.sub('</?para>\n?', '', table)
                        entries = pattern_entry.findall(table)
                        spanned_entries = [e[1] for e in entries if "namest" in e[0]]
                        single_entries = [e[1] for e in entries if "namest" not in e[0]]	
                        
                        pairs = []
                        num_of_cols = len(re.findall(r'<colspec', table)) // 2
                        
                        for i in range (0, num_of_cols):
                            for j, entry in enumerate(single_entries[i * 2:-1:2 * num_of_cols]):
                                pairs.append((entry.strip(), single_entries[j*2*num_of_cols+1+i*2].strip()))
        
                        if len(pairs) > 0:
                            legend = "<legend>\n<definitionList>\n"
                            for pair in pairs:
                                if pair[0] != "" or pair[1] != "":
                                    if pair[0] == "":
                                        key = " "
                                    else:
                                        key = re.sub(r'\.', '', pair[0])
                                    legend += """<definitionListItem>
    <listItemTerm>{}</listItemTerm>
    <listItemDefinition><para>{}</para>
    </listItemDefinition>
    </definitionListItem>\n""".format(key, pair[1])
                            legend += "</definitionList>\n</legend>\n"
                            if len(spanned_entries) > 0:
                                notes = '<table{}><tgroup cols="1"><colspec colname="col1"/><tbody><row>\n<entry>'.format(' frame="bottom"' if "<note" in spanned_entries[0] else "")
                                for se in spanned_entries:
                                    notes += se.strip()
                                notes += "\n</entry></row>\n</tbody></tgroup></table>\n"
                        else:
                            legend = ""
                return "{}</graphic>{}{}{}".format(match.group(1), legend, notes, dim)
                
            def generate_icn(match):
                return 'infoEntityIdent="{}"'.format("ICN-{}-0000{}-001-01".format(book_info['cage'], check_match(re.search(r'\d{6}', match.group(1)), 0)))	
            
            def replace_sheet_effect(match, title):
                return '{1}\n<title>({2}) {0}</title>\n{3}{4}'.format(\
                title, \
                '</graphic>\n<graphic key="{}_f">'.format(match.group(3)) if match.group(1) is None else match.group(1),\
                match.group(4),\
                match.group(2),\
                match.group(5)
                )
            
            text = match.group(0)
            title = check_match(pattern_title.search(text))
            text = re.sub(pattern_sgm_attr, '', text)
            text = re.sub(pattern_effect2, partial(replace_sheet_effect, title=title), text)
            text = re.sub(pattern_split_graphic, r'\1</graphic>\n<graphic key="\3_f">\n<title>{}</title>\n\2'.format(title), text)
            text = re.sub(r'<(/?)graphic', r'<\1figure', text)
            text = re.sub(r'<(/?)sheet', r'<\1graphic', text)
            text = re.sub('key', 'id', text)
            text = re.sub('gnbr="([^"]+)"', generate_icn, text)
            text = re.sub(r'(\s?)imgarea="hl"', r'\1reproductionWidth="355.6 mm" reproductionHeight="209.55 mm"', text)
            text = re.sub(r'\s?imgarea="[^"]+\n?"', '', text)
            text = re.sub(re.compile(r'(<graphic[^>]*>)(?!\n?</graphic>).*?<gdesc>(.*?)</gdesc>\n?</graphic>', re.DOTALL), replace_gdesc, text)
            text = re.sub(re.compile(r'(<table.*?</table>)\n*</figure>', re.DOTALL), r'</figure>\n\1', text)
            return text

        def process_tables(match):
            def FootnoteRefRepl(match):
                for i,refid in enumerate(note_ids):		
                    if match.group(1) == refid:
                        return "<superScript>%s</superScript>" % str(i+1)
                else:
                    return match.group(0)
                    
            table = match.group(0)
            table = re.sub(pattern_rht, "", table)
            table = re.sub(pattern_sgm_attr, '', table)
            #Deal with footnotes in tables. Will add each footnote to a attention sequential list item in the last row of the table, and replace the refids with superscript numbers.
            ftnotes = pattern_ftnote.findall(table)
            if len(ftnotes) > 0:
                cols = re.findall(r'colname="([^"]+)"', table)
                end_col = cols[-1]
                st_col = cols[0]
                note_ids = []
                note_content = "\n<row>\n<entry colsep=\"0\" nameend=\"%s\" namest=\"%s\">\n<?PubTbl cell border-left-style=\"none\" border-right-style=\"none\"?>\n<note>\n<attentionSequentialList>\n" % (end_col, st_col)
                for fn in ftnotes:
                    note_ids.append(check_match(re.search('ftnoteid="([^"]+)"', fn)))
                    note_content += "<attentionSequentialListItem>"
                    note_paras = pattern_para.findall(fn)
                    for p in note_paras:
                        note_content += "<attentionListItemPara>%s</attentionListItemPara>\n" % p
                    note_content += "</attentionSequentialListItem>"
                note_content += "\n</attentionSequentialList></note>\n</entry>\n</row></tbody></tgroup>"
                note_content = re.sub(pattern_note, NotePara, note_content)
                table = re.sub(pattern_ftnotesub, note_content, table)
                table = re.sub(pattern_refint, FootnoteRefRepl, table)
                
            return table
        
        def sub_equ(match):
            if match.group(3) == "":
                return r'{}<emphasis emphasisType="em01">{}</emphasis>'.format("</para><para>" if match.group(1) is None else "<para>", match.group(2))
            else:
                return r'{}<emphasis emphasisType="em01">{}</emphasis></para><para>{}'.format("</para><para>" if match.group(1) is None else "<para>", match.group(2), match.group(3))
        
        
        source = source_file.read_text(encoding='utf-8')
        source = replace_entities(source)
        source_file.write_text(source, encoding='utf-8')
        #Sub out unsupported elements
        source = re.sub(r'[\r\n]+>', '>\n', source)
        source = re.sub(r'<(\?[^>?]+)>', r'<\1?>', source)
        source = re.sub(r'<\?Pub Dtl\?>\n?', '', source)
        source = re.sub('<(?:revst|revend)>[\n\s]?', '', source)
        source = re.sub(r'\s?chg="\w"', '', source)
        source = re.sub(r'<\?ITG-STRIP-REV[^>]+>[\n\s]?', '', source)
        source = re.sub(re.compile(r'<chgdesc.*?</chgdesc>\n?', re.DOTALL), '', source)
        source = re.sub(r'<!--.*?-->\n?', '', source)
        source = re.sub(pattern_rev, "", source)
        source = re.sub(pattern_subtitle, r'<para>\1</para>', source)
        source = re.sub(pattern_equ, sub_equ, source)
        source = re.sub(r'</equ><equ>', r'</emphasis></para><para><emphasis emphasisType="em01">', source)
        source = re.sub(pattern_note, NotePara, source)

        source = re.sub(re.compile(r'<(acro|abbr)>.*?</\1>', re.DOTALL), SubAcro, source)
        source = re.sub(pattern_refext, '', source)
        source = re.sub(r'</?(?:prc)?list\d>\n?', '', source)
        source = re.sub(r'</?(?:prc)?item>\n?', '', source)
        source = re.sub(pattern_sub_super, r'<\1Script>\2</\1Script>', source)
        source = re.sub(pattern_colspec, r'\1</colspec>', source)
        
        source = re.sub(re.compile(r'<table.*?</table>', re.DOTALL), process_tables, source)
        source = re.sub(r'<unlist[^>]*>', r'<para><randomList>', source)
        source = re.sub(r'</unlist>', r'</randomList></para>', source)
        source = re.sub(r'<numlist[^>]*>', r'<para><sequentialList>', source)
        source = re.sub(r'</numlist>', r'</sequentialList></para>', source)
        source = re.sub(r'(?:num|un)litem[^>]*>', r'listItem>', source)
        
        book_info['cage'] = check_match(pattern_cage.search(source)).upper()
        book_info['revdate'] = check_match(pattern_revdate.search(source))
        source = re.sub(r'\s?revdate="[^"]+"', '', source)
        source = re.sub(re.compile(r'</subtask>\n?(<graphic.*?(?=<(sub|/)task))', re.DOTALL), r'\1\n</subtask>\n', source)
        #source = re.sub(pattern_sh_eff, '', 
        source = re.sub(pattern_effect, r'<title>(\1) \2</title>', source)
        
        try:
            process_bluesheet(source) #Process bluesheet stuff before messing with the images
        except:
            print("Bluesheet failed to process:\n")
            print(traceback.format_exc())
        source = re.sub(pattern_graphic, replace_graphics, source)
        source = re.sub(r'<prcitem(\d)>\n?</prcitem\1>', '', source)
        source = re.sub(r'<l(\d)item>\n?</l\1item>', '', source) #Found in OHM
        #(dir /'processed_source.sgml').write_text(source, encoding='utf-8')
        #Replace entity references in source with the actual text
        
        return source

    def create_cir(type, data):
        def sub_refs(match):
            id = match.group(1)
            lst =  ToolAndCons.tools if "tool" in match.group(1) else ToolAndCons.cons
            for t in lst:
                if t.id == id:
                    return ("{} {}".format(t.pn, t.nom)).upper()
            else:
                for tsk in tasks:
                    #Check for task references
                    if tsk.key == id:
                        authname = tsk.children[0].authname.split('.')[0] if tsk.children[0].authname is not None else ''
                        if authname != '':
                            authname = "{}. {}".format(authname, tsk.title)
                        return re.sub(' authorityDocument="[^"]+"', ' authorityName="{}"'.format(authname), tsk.children[0].dmref)
                    
                    #Check for subtask or graphic references
                    for dm2 in tsk.children:
                        if dm2.key == id:
                            return re.sub(' authorityDocument="[^"]+"', (' authorityName="{}"'.format(dm2.authname)) if dm2.authname is not None else '', dm2.dmref) 	
                        elif id in dm2.graph_refs or id in dm2.table_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(id), dm2.dmref)
                        elif id in dm2.sheet_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[id][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[id][1]))
                        elif id in dm2.graph_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(id), dm2.dmref)
                        elif id in dm2.sheet_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[id][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[id][1])) #Will return a reference to the figure, not the sheet			
                else:
                    #Check for graphic references to the Intro
                    for dm2 in introduction:
                        if id in dm2.graph_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(id), dm2.dmref)
                        elif id in dm2.sheet_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[id][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[id][1])) #Will return a reference to the figure, not the sheet
                    else:
                        return match.group(0)
        
        content = ""
        infocode_var = "A"
        infocode = "012"
        if type == "tool" or type == "cons":
            content = "<{}Repository>\n".format("tool" if type == "tool" else "supply")
            for d in data:
                content += d.spec
            content += "</{}Repository>\n".format("tool" if type == "tool" else "supply")
            infocode = "00L" if type == "cons" else "00N"
            title = "Supplies CIR" if type == "cons" else "Support Equipment CIR"
        else:
            content = "<{}Repository>\n".format("caution" if type == "caut" else "warning")
            for d in data:
                content += """<{2}Spec><{2}Ident id="{0}" {2}IdentNumber="{0}"/>{1}</{2}Spec>\n""".format(d, warn[d] if type == "warn" else caut[d], "caution" if type == "caut" else "warning")
            content += "</{}Repository>\n".format("caution" if type == "caut" else "warning")
            infocode_var = "B" if type == "caut" else "A"
            title = "Cautions CIR" if type == "caut" else "Warnings CIR"
        
            content = re.sub(r'<internalRef internalRefId="([^"]+)"[^>]*>\n?</internalRef>', sub_refs, content)
        
        dmcode = """<dmCode assyCode="{0}" disassyCode="00" disassyCodeVariant="A" infoCode="{5}" infoCodeVariant="{6}" itemLocationCode="D" modelIdentCode="HON{1}" subSubSystemCode="{2}" subSystemCode="{3}" systemCode="{4}" systemDiffCode="EAA"/>""".format(book_info['ata'][6:8], book_info['cage'], book_info['ata'][4], book_info['ata'][3], book_info['ata'][0:2], infocode, infocode_var)
        return """<?xml version="1.0" encoding="UTF-8"?>
    <!--Arbortext, Inc., 1988-2014, v.4002-->
    <!DOCTYPE dmodule [
    <!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
    %ISOEntities;
    ]>
    <?Pub Inc?>
    <dmodule xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/comrep.xsd">
    <identAndStatusSection>
    <dmAddress>
    <dmIdent>
    {0}
    <language countryIsoCode="US" languageIsoCode="sx"/>
    <issueInfo inWork="00" issueNumber="001"/></dmIdent>
    <dmAddressItems>
    <issueDate day="{7}" month="{6}" year="{5}"/>
    <dmTitle><techName>{4}</techName><infoName>{2}</infoName></dmTitle>
    </dmAddressItems></dmAddress>
    <dmStatus issueType="new">
    <security securityClassification="01"/>
    <responsiblePartnerCompany enterpriseCode="{8}"></responsiblePartnerCompany>
    <originator enterpriseCode="{8}"></originator>
    <applic>
    <displayText>
    <simplePara>ALL</simplePara>
    </displayText>
    </applic>
    <brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
    <qualityAssurance>
    <unverified/></qualityAssurance></dmStatus>
    </identAndStatusSection>
    <content>
    <commonRepository>
    {1}
    </commonRepository>
    </content>
    </dmodule>""".format(dmcode, content, title, add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8], book_info['cage'])
        
    def create_appliccrossreftable(filename):
        dmcode = check_match(re.search('<dmCode [^>]+>', get_dmref(filename)), 0)
        return """<?xml version="1.0" encoding="UTF-8"?>
    <!--Arbortext, Inc., 1988-2014, v.4002-->
    <!DOCTYPE dmodule [
    <!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
    %ISOEntities;
    ]>
    <?Pub Inc?>
    <dmodule
    xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/appliccrossreftable.xsd"
    xmlns:dc="http://www.purl.org/dc/elements/1.1/"
    xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
    xmlns:xlink="http://www.w3.org/1999/xlink"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <?MENTORPATH ?>
    <identAndStatusSection>
    <dmAddress>
    <dmIdent>
    {0}
    <language countryIsoCode="US" languageIsoCode="sx"/>
    <issueInfo inWork="00" issueNumber="001"/></dmIdent>
    <dmAddressItems>
    <issueDate day="{10}" month="{9}" year="{8}"/>
    <dmTitle><techName>{7}</techName><infoName>Applicability
    Cross-reference Table (ACT)</infoName></dmTitle>
    </dmAddressItems></dmAddress>
    <dmStatus issueType="new">
    <security securityClassification="01"/>
    <responsiblePartnerCompany enterpriseCode="59364">
    </responsiblePartnerCompany>
    <originator enterpriseCode="59364"></originator>
    <applic id="app-0001">
    <displayText>
    <simplePara>ALL</simplePara>
    </displayText>
    </applic>
    <brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
    <qualityAssurance>
    <unverified/></qualityAssurance></dmStatus>
    </identAndStatusSection>
    <content>
    <applicCrossRefTable>
    <productAttributeList>
    <productAttribute id="PN" productIdentifier="primary">
    <name>Part Number</name><?Pub Caret 11?>
    <displayName>Part Number</displayName>
    <descr>Part Number</descr>
    </productAttribute>
    <productAttribute id="Model" productIdentifier="primary">
    <name>Model</name>
    <displayName>Model</displayName>
    <descr>Model</descr>
    </productAttribute>
    <productAttribute id="cage" productIdentifier="primary">
    <name>CAGE</name>
    <displayName>CAGE</displayName>
    <descr>CAGE</descr>
    </productAttribute>
    </productAttributeList>
    <productCrossRefTableRef><dmRef><dmRefIdent><dmCode assyCode="{1}"
    disassyCode="00" disassyCodeVariant="A" infoCode="00P"
    infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HON{2}"
    subSubSystemCode="{3}" subSystemCode="{4}" systemCode="{5}"
    systemDiffCode="EAA"/></dmRefIdent></dmRef></productCrossRefTableRef>
    </applicCrossRefTable>
    </content>
    </dmodule>
    <?INMEDLNG sx-US?>""".format(dmcode, book_info['ata'][6:8], book_info['cage'], book_info['ata'][4], book_info['ata'][3], book_info['ata'][0:2], add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8])
        
    def create_prdcrossreftable(filename): #Create 00P file
        content = ""
        partinfo = pattern_partinfo.findall(source)
        for pi in partinfo:
            model = check_match(pattern_model.search(pi[0]))
            mfrpnr = pattern_mfrpnr.findall(pi[1])
            for mp in mfrpnr:
                mfr = check_match(pattern_mfr.search(mp))
                pnr = check_match(pattern_pnr.search(mp))
                content += """\n<product>\n
    <assign applicPropertyIdent="PN" applicPropertyType="prodattr" applicPropertyValue="{}"/>\n
    <assign applicPropertyIdent="Model" applicPropertyType="prodattr" applicPropertyValue="{}"/>\n
    <assign applicPropertyIdent="cage" applicPropertyType="prodattr" applicPropertyValue="{}"/>\n</product>""".format(pnr, model, mfr)
                altpnr = re.search(r'<altpnr>\n?<pnr>([^<]+)</pnr>\n?<mfr>([^<]+)</mfr>', mp)
                if altpnr is not None:
                    content += """\n<product>\n
    <assign applicPropertyIdent="PN" applicPropertyType="prodattr" applicPropertyValue="{}"/>\n
    <assign applicPropertyIdent="Model" applicPropertyType="prodattr" applicPropertyValue="{}"/>\n
    <assign applicPropertyIdent="cage" applicPropertyType="prodattr" applicPropertyValue="{}"/>\n</product>""".format(altpnr.group(1), model, altpnr.group(2))
        
        dmcode = check_match(re.search('<dmCode [^>]+>', get_dmref(filename)), 0)
        
        return """<?xml version="1.0" encoding="UTF-8"?>
    <!--Arbortext, Inc., 1988-2014, v.4002-->
    <!DOCTYPE dmodule [
    <!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
    %ISOEntities;
    ]>
    <?Pub Inc?>
    <dmodule xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/prdcrossreftable.xsd" xmlns:dc="http://www.purl.org/dc/elements/1.1/" xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
    xmlns:xlink="http://www.w3.org/1999/xlink" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <identAndStatusSection>
    <dmAddress>
    <dmIdent>
    {0}
    <language countryIsoCode="US" languageIsoCode="sx"/>
    <issueInfo inWork="00" issueNumber="001"/></dmIdent>
    <dmAddressItems>
    <issueDate day="{7}" month="{6}" year="{5}"/>
    <dmTitle><techName>{4}</techName><infoName>Product Cross-reference Table (PCT)</infoName></dmTitle>
    </dmAddressItems></dmAddress>
    <dmStatus issueType="new">
    <security securityClassification="01"/>
    <responsiblePartnerCompany enterpriseCode="{1}"></responsiblePartnerCompany>
    <originator enterpriseCode="{1}"></originator>
    <applic id="app-0001">
    <displayText>
    <simplePara>ALL</simplePara>
    </displayText>
    </applic>
    <brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
    <qualityAssurance>
    <firstVerification verificationType="tabtop"/></qualityAssurance></dmStatus>
    </identAndStatusSection>
    <content>
    <productCrossRefTable>
    {2}
    </productCrossRefTable>
    </content>
    </dmodule>
    <?INMEDLNG sx-US?>
    """.format(dmcode, book_info['cage'], content, add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8])
            
    def create_descript(dm, add_levelled_para=True): #Create a Descriptive DM
        dmcode = check_match(re.search('<dmCode [^>]+>', dm.dmref), 0)
        return """<?xml version="1.0" encoding="UTF-8"?>
    <!--Arbortext, Inc., 1988-2014, v.4002-->
    <!DOCTYPE dmodule [
    <!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
    %ISOEntities;
    ]>
    <?Pub Inc?>
    <dmodule xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/descript.xsd" xmlns:dc="http://www.purl.org/dc/elements/1.1/" xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#" xmlns:xlink="http://www.w3.org/1999/xlink"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <identAndStatusSection>
    <dmAddress>
    <dmIdent>
    {0}
    <language countryIsoCode="US" languageIsoCode="sx"/>
    <issueInfo inWork="00" issueNumber="001"/></dmIdent>
    <dmAddressItems>
    <issueDate day="{8}" month="{7}" year="{6}"/>
    <dmTitle><techName>{5}</techName><infoName>{3}</infoName></dmTitle>
    </dmAddressItems></dmAddress>
    <dmStatus issueType="new">
    <security securityClassification="01"/>
    <responsiblePartnerCompany enterpriseCode="{1}"></responsiblePartnerCompany>
    <originator enterpriseCode="{1}"></originator>
    <applic id="app-0001a">
    <displayText>
    <simplePara>ALL</simplePara>
    </displayText>
    </applic>
    <brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
    <qualityAssurance>
    <firstVerification verificationType="tabtop"/></qualityAssurance></dmStatus>
    </identAndStatusSection>
    <content>
    {9}
    <description>
    {2}
    </description></content></dmodule>
    <?INMEDCLNPRJ A350_CMMs_-_Honeywell?>
    <?INMEDBSPATH A350_CMMs_-_Honeywell?>
    <?INMEDLNG sx-US?>
    """.format(dmcode, book_info['cage'], ('<levelledPara>{}</levelledPara>'.format(dm.content)) if add_levelled_para else dm.content, dm.infoname, add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8], ("<warningsAndCautionsRef>{}</warningsAndCautionsRef>".format(dm.warn_and_caut_data)) if dm.warn_and_caut_data != "" else "")

    def create_proced(dm): #Create a Procedural DM
        dmcode = check_match(re.search('<dmCode [^>]+>', dm.dmref), 0)
        tool_data = dm.tool_data
        cons_data = dm.cons_data
        return """<?xml version="1.0" encoding="UTF-8"?>
    <!--Arbortext, Inc., 1988-2014, v.4002-->
    <!DOCTYPE dmodule [
    <!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
    %ISOEntities;
    ]>
    <?Pub Inc?>
    <dmodule xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/proced.xsd" xmlns:dc="http://www.purl.org/dc/elements/1.1/" xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#" xmlns:xlink="http://www.w3.org/1999/xlink"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <identAndStatusSection>
    <dmAddress>
    <dmIdent>
    {0}
    <language countryIsoCode="US" languageIsoCode="sx"/>
    <issueInfo inWork="00" issueNumber="001"/></dmIdent>
    <dmAddressItems>
    <issueDate day="{11}" month="{10}" year="{9}"/>
    <dmTitle><techName>{8}</techName><infoName>{6}</infoName></dmTitle>
    </dmAddressItems></dmAddress>
    <dmStatus issueType="new">
    <security securityClassification="01"/>
    <responsiblePartnerCompany enterpriseCode="{1}"></responsiblePartnerCompany>
    <originator enterpriseCode="{1}"></originator>
    <applic id="app-0001a">
    <displayText>
    <simplePara>ALL</simplePara>
    </displayText>
    </applic>
    <brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
    <qualityAssurance>
    <firstVerification verificationType="tabtop"/></qualityAssurance></dmStatus>
    </identAndStatusSection>
    <content>
    {5}
    <procedure><preliminaryRqmts>
    <reqCondGroup><noConds/></reqCondGroup>
    <reqSupportEquips>
    {3}
    </reqSupportEquips>
    <reqSupplies>
    {4}
    </reqSupplies>
    <reqSpares><noSpares/></reqSpares>
    <reqSafety><noSafety/></reqSafety>
    </preliminaryRqmts>
    <mainProcedure>
    <proceduralStep>
    {2}
    </proceduralStep>
    </mainProcedure>
    <closeRqmts>
    <reqCondGroup><noConds/></reqCondGroup>
    </closeRqmts>
    </procedure></content></dmodule>
    <?INMEDCLNPRJ A350_CMMs_-_Honeywell?>
    <?INMEDBSPATH A350_CMMs_-_Honeywell?>
    <?INMEDLNG sx-US?>
    """.format(dmcode, book_info['cage'], dm.content, "<supportEquipDescrGroup>{}</supportEquipDescrGroup>".format(tool_data) if tool_data != "" else "<noSupportEquips/>", "<supplyDescrGroup>{}</supplyDescrGroup>".format(cons_data) if cons_data != "" else "<noSupplies/>", ("<warningsAndCautionsRef>{}</warningsAndCautionsRef>".format(dm.warn_and_caut_data)) if dm.warn_and_caut_data != "" else "", dm.infoname, add_leading_zero(book_info['issue_no'], 3), book_info['cmpnom'], book_info['revdate'][0:4], book_info['revdate'][4:6], book_info['revdate'][6:8])

    def create_pmc(content): #Create the PMC
        return """<?xml version="1.0" encoding="UTF-8"?>
    <!--Arbortext, Inc., 1988-2014, v.4002-->
    <!DOCTYPE pm [
    <!ENTITY % ISOEntities PUBLIC "ISO 8879-1986//ENTITIES ISO Character Entities 20030531//EN//XML" "http://www.s1000d.org/S1000D_4-1/ent/ISOEntities">
    %ISOEntities;
    ]>
    <?Pub Inc?>
    <pm
    xsi:noNamespaceSchemaLocation="http://www.s1000d.org/S1000D_4-1/xml_schema_flat/pm.xsd"
    xmlns:dc="http://www.purl.org/dc/elements/1.1/"
    xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
    xmlns:xlink="http://www.w3.org/1999/xlink"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <identAndStatusSection><pmAddress>
    <pmIdent>
    <pmCode modelIdentCode="HON{0}" pmIssuer="{0}" pmNumber="00001"
    pmVolume="01"/><language countryIsoCode="US" languageIsoCode="sx"/>
    <issueInfo inWork="00" issueNumber="{6}"/></pmIdent>
    <pmAddressItems>
    <issueDate day="{9}" month="{8}" year="{7}"/>
    <pmTitle>{13}</pmTitle>
    <shortPmTitle>{14}</shortPmTitle>
    <externalPubCode pubCodingScheme="initialDate">{12} {11} {10}</externalPubCode>
    <externalPubCode pubCodingScheme="CMP">{3}</externalPubCode>
    <externalPubCode pubCodingScheme="INT">{5}</externalPubCode>
    </pmAddressItems>
    </pmAddress><pmStatus issueType="new">
    <security securityClassification="01"/>
    <dataRestrictions>
    <restrictionInstructions>
    <dataDistribution></dataDistribution>
    <exportControl>
    <exportRegistrationStmt>
    <simplePara>These items are controlled by the U.S. government and
    authorized for export only to the country of ultimate destination
    for use by the ultimate consignee or end-user(s) herein identified.
    They may not be resold, transferred, or otherwise disposed of, to
    any other country or to any person other than the authorized ultimate
    consignee or end-user(s), either in their original form or after being
    incorporated into other items, without first obtaining approval from
    the U.S. government or as otherwise authorized by U.S. law and regulations.</simplePara>
    <simplePara>{2}</simplePara>
    </exportRegistrationStmt>
    </exportControl>
    <dataHandling></dataHandling>
    <dataDestruction></dataDestruction>
    <dataDisclosure></dataDisclosure>
    </restrictionInstructions>
    <restrictionInfo>
    <copyright>
    <copyrightPara>&copy; Honeywell International Inc. Do not copy without
    express permission of Honeywell.</copyrightPara>
    </copyright>
    <policyStatement></policyStatement>
    <dataConds></dataConds>
    </restrictionInfo>
    </dataRestrictions>
    <responsiblePartnerCompany enterpriseCode="{0}">
    </responsiblePartnerCompany>
    <originator enterpriseCode="{0}"></originator>
    <applicCrossRefTableRef>
    {4}
    </applicCrossRefTableRef>
    <applic>
    <displayText>
    <simplePara>ALL</simplePara>
    </displayText>
    </applic>
    <brexDmRef><dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="022" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0" systemCode="00" systemDiffCode="A"/></dmRefIdent></dmRef></brexDmRef>
    <qualityAssurance>
    <unverified/></qualityAssurance>
    </pmStatus></identAndStatusSection>
    <content>
    {1}
    </content>
    </pm>
    <?INMEDLNG sx-US?>""".format(book_info['cage'],
                                 content,
                                 book_info['eccn'],
                                 book_info['ata'],
                                 get_dmref("DMC-HON{}-EAA-{}-00A-00WA-D_sx-US.XML".format(book_info['cage'],
                                                                                           book_info['ata'])),
                                                                                          book_info['docnbr'],
                                                                                          add_leading_zero(book_info['issue_no'], 3),
                                                                                          book_info['revdate'][0:4],
                                                                                          book_info['revdate'][4:6],
                                                                                          book_info['revdate'][6:8],
                                                                                          book_info['oidate'][0:4],
                                                                                          months[int(book_info['oidate'][4:6])],
                                                                                          book_info['oidate'][6:8],
                                                                                          book_info['cmpnom'],
                                                                                          book_info['model'])

    def GetServiceBulletin():
        nonlocal pub
        def ReplaceRow(match):
            if '/' in match.group(0):
                return "</row>"
            else:
                return "<row rowsep=\"0\">"
            
        def ReplaceEntry(match):
            if '/' in match.group(0):
                return "</para></entry>"
            else:
                return "<entry><para>"
        
        filename_008A = "DMC-HONAERO-EAB-00-00-00-00A-008A-D_sx-US.XML"

        sblist = re.search(pattern_sblist, source)
        # if sblist is None:
            # log_print("Could Not Locate Service Bulletin List In Source")
            # return
            
        table_body = ""
        sbdata = pattern_sbdata.findall(sblist.group(0))
        
        for s in sbdata:
            s = re.sub('</sbtitle><issdate>', '</sbtitle><entry></entry><issdate>', s)
            s = re.sub(r"<(?!\?)/?(?!sbdata)\w+>", ReplaceEntry, s)
            table_body += s + '\n'

        table_body = re.sub(pattern_sbdatatag, ReplaceRow, table_body)

        content_008A = """<table id="SB-1" frame="topbot"><tgroup cols="4">
    <?PubTbl tgroup dispwid="700.00px"?>
    <colspec colname="col1" colsep="0" colwidth="25.18*" /><colspec
    colname="col2" colsep="0" colwidth="43.81*" /><colspec colname="col3"
    colsep="0" colwidth="14.56*" /><colspec colname="COLSPEC0"
    colwidth="16.45*" /><thead><row rowsep="1">
    <entry>
    <para>Service Bulletin /</para>
    <para>Revision Number</para>
    </entry>
    <entry>
    <para>Title</para>
    </entry>
    <entry>
    <para>Date Put In</para>
    <para>Manual</para>
    </entry>
    <entry>
    <para>Status</para>
    </entry>
    </row></thead><tbody>
    %s
    </tbody></tgroup></table>""" % (table_body)
        return DataModule(content_008A, "sbl", None, None, None, None, book_info, filename=filename_008A, infoname="SERVICE BULLETIN LIST", pub=pub)	

    def GetTemporaryRevision():
        nonlocal pub
        empty_entry = "\n<entry></entry>"

        filename_003C = "DMC-HONAERO-EAB-00-00-00-00A-003C-D_sx-US.XML"

        trlist = re.search(pattern_trlist, source) #Get data
        
        # if trlist is None: #Ensure we got the list
            # log_print("Could Not Locate Temporary Revision List In Source")
            # return
            
        table_body = ""
        trdata = pattern_trdata.findall(trlist.group(0))
        
        for t in trdata:
            table_body += t + '\n'
        
        for r in (("<trdata>", "<row rowsep=\"0\">"), ("</trdata>", "%s</row>" % (empty_entry * 5)), ("<trnbr>","<entry><para>"), ("<trstatus>","<entry><para>"),  ("</trnbr>","</para></entry>"), ("</trstatus>","</para></entry>"), ("<trloc>","<entry><para>"), ("</trloc>","</para></entry>")):
            table_body = table_body.replace(*r)
            
        content_003C = """<para>Instructions on each page of a temporary revision tell you where to put the pages in your manual. Remove the temporary revision pages only when discard instructions are given. For each temporary revision, put the
    applicable data in the record columns on this page.</para>
    <para>Definition of Status column: A <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> may be active, incorporated, or deleted. &ldquo;Active&rdquo; is entered by the holder of the manual. &ldquo;Incorporated&rdquo;
    means a <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> has been incorporated into the manual and includes the revision number of the manual when the <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary
    revision</acronymDefinition></acronym> was incorporated. &ldquo;Deleted&rdquo; means a <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> has been replaced by another <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary
    revision</acronymDefinition></acronym>, a <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> number will not be issued, or a <acronym><acronymTerm>TR</acronymTerm><acronymDefinition>temporary revision</acronymDefinition></acronym> has
    been deleted.</para>
    <table id="TR-1" frame="topbot"><tgroup cols="8">
    <?PubTbl tgroup dispwid="700.00px"?>
    <colspec align="center" colname="col1" colsep="0" colwidth="1.21*"/><colspec colname="col2" colsep="0" colwidth="1.48*"/>
    <colspec align="center" colname="col3" colsep="0" colwidth="0.98*"/><colspec align="center" colname="col4" colsep="0" colwidth="1.30*"/>
    <colspec align="center" colname="col5" colsep="0" colwidth="1.30*"/><colspec colname="col6" colsep="0" colwidth="0.87*"/>
    <colspec colname="col7" colsep="0" colwidth="1.15*"/><colspec colname="col8" colsep="0" colwidth="0.70*"/><thead><row rowsep="1">
    <entry colsep="0" valign="bottom">
    <?PubTbl cell border-left-style="none"?>
    <para>Temporary</para>
    <para>Revision</para>
    <para>Number</para>
    </entry>
    <entry align="center" colsep="0" valign="bottom">
    <para>Status</para>
    </entry>
    <entry colsep="0" valign="bottom">
    <para>Page Number</para>
    </entry>
    <entry colsep="0" valign="bottom">
    <para>Issue Date</para>
    </entry>
    <entry colsep="0" valign="bottom">
    <para>Date Put In</para>
    <para>Manual</para>
    </entry>
    <entry align="center" colsep="0" valign="bottom">
    <para>By</para>
    </entry>
    <entry align="center" colsep="0" valign="bottom">
    <para>Date</para>
    <para>Removed</para>
    <para>From</para>
    <para>Manual</para>
    </entry>
    <entry align="center" valign="bottom">
    <?PubTbl cell border-right-style="none"?>
    <para>By</para>
    </entry>
    </row></thead><tbody>
    %s
    </tbody></tgroup></table>""" % (table_body)
        #TODO: Validate number of entries?
        return DataModule(content_003C, "rotr", None, None, None, None, book_info, filename=filename_003C, infoname="RECORD OF TEMPORARY REVISIONS", pub=pub)	

    def GetVendorCodes():
        nonlocal pub
        def AddV(match):
            return "row rowsep=\"0\">%s<entry><para>V%s" % (match.group(1) if match.group(1) is not None else '', match.group(2))

        filename_018Z = "DMC-HON{}-EAA-00-00-00-00A-018Z-C_sx-US.XML".format(book_info['cage'])

        vendata = re.findall(pattern_vendata, source)
        if len(vendata) == 0:
            print("Could Not Locate Vendor Code List Data In The Source.")
            return
            
        table_body = ""
        
        for v in vendata:
            table_body += v + '\n'
            
        table_body = re.sub(pattern_mfrcont, "", table_body)
        
        for r in (("<vendata>", "<row rowsep=\"0\">"), ("</vendata>", "</row>"), ("<mad>","<entry><para>"), ("<mfr>","<entry><para>"),  ("</mad>","</para></entry>"), ("</mfr>","</para></entry>")):
            table_body = table_body.replace(*r)
            
        table_body = re.subn(pattern_v, AddV, table_body)[0]

        content_018Z = """<note><notePara>The vendor codes and part numbers that are shown in the DPL must not be construed as an authorization of the vendor, pursuant to the FAA regulations, to ship directly to the user. Neither must it be construed
    as a certification by Honeywell that parts shipped by vendors directly to users will conform to the type design or that such parts are airworthy or safe for installation.</notePara></note>
    <table frame="topbot"><tgroup cols="2">
    <?PubTbl tgroup dispwid="700.00px"?>
    <colspec colname="col1" colsep="0" colwidth="0.28*"/><colspec
    colname="col2" colwidth="1.72*"/><thead><row rowsep="0">
    <entry rowsep="1">
    <para>Code</para>
    </entry>
    <entry rowsep="1">
    <para>Vendor</para>
    </entry>
    </row></thead><tbody>
    %s
    </tbody></tgroup></table>""" % (table_body)

        return DataModule(content_018Z, "vendlist", None, None, None, None, book_info, filename=filename_018Z, infoname="Vendor Code List", pub=pub)

    def GetTransmittalInfo():
        nonlocal pub
        filename_003A = "DMC-HONAERO-EAB-00-00-00-00A-003A-D_sx-US.XML"
        transltr = check_match(pattern_transltr.search(source))
        
        return DataModule(transltr, "transltr", None, None, None, None, book_info, filename=filename_003A, infoname="TRANSMITTAL INFORMATION", pub=pub)

    def GetRecordOfRevisions():
        nonlocal pub
        filename_003B = "DMC-HONAERO-EAB-00-00-00-00A-003B-D_sx-US.XML"
        content = """<para>For each revision, write the revision number, revision date, date put in the manual, and your initials in the applicable column.</para><note>
    <notePara>Refer to the Revision History in the <dmRef><dmRefIdent><dmCode assyCode="00" disassyCode="00" disassyCodeVariant="A" infoCode="003" infoCodeVariant="A" itemLocationCode="D" modelIdentCode="HONAERO" subSubSystemCode="0" subSystemCode="0"
    systemCode="00" systemDiffCode="EAB"/></dmRefIdent></dmRef> section for revision data.</notePara>
    </note><table><tgroup cols="9">
    <?PubTbl tgroup dispwid="1230.00px"?>
    <colspec colname="col1" colwidth="0.92in"/><colspec colname="col2" colwidth="0.86in"/><colspec colname="col3" colwidth="0.78in"/><colspec colname="col4" colwidth="0.51in"/><colspec colname="col5" colwidth="0.31in"/><colspec colname="col6" colwidth="0.76in"/>
    <colspec colname="col7" colwidth="0.75in"/><colspec colname="col8" colwidth="0.74in"/><colspec colname="col9" colwidth="0.47in"/><thead><row>
    <entry align="center">
    <para>Revision Number</para>
    </entry>
    <entry align="center">
    <para>Revision Date</para>
    </entry>
    <entry align="center">
    <para>Date Put in Manual</para>
    </entry>
    <entry align="center">
    <para>By</para>
    </entry>
    <entry></entry>
    <entry align="center">
    <para>Revision Number</para>
    </entry>
    <entry align="center">
    <para>Revision Date</para>
    </entry>
    <entry align="center">
    <para>Date Put in Manual</para>
    </entry>
    <entry align="center">
    <para>By</para>
    </entry>
    </row></thead><tbody><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row><row>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    <entry></entry>
    </row></tbody></tgroup></table>"""
        return DataModule(content, "ror", None, None, None, None, book_info, filename=filename_003B, infoname="RECORD OF REVISIONS", pub=pub)
        
    def GetBookInfo(source): #Get metadata from book
        #Get ATA number
        book_info['issue_no'] = check_match(pattern_tsn.search(source))
        book_info['oidate'] = check_match(pattern_oidate.search(source))
        book_info['eccn'] = check_match(pattern_eccn.search(source), 0)
        book_info['cpyrgt'] = check_match(pattern_copyright.search(source))
        book_info['docnbr'] = check_match(pattern_docnbr.search(source))
        book_info['ata'] = book_info['docnbr'] if re.match(r'\d{2}-\d{2}-\d{2}', book_info['docnbr']) else get_input("ATA Number not found. Please enter one manually: ")
        book_info['model'] = check_match(pattern_model.search(source))
        book_info['cmpnom'] = check_match(pattern_cmpnom.search(source))

        #PM Title
        
        return book_info

    def GetPrelim():
        nonlocal pub
        names = ["DMC-HONAERO-EAA-00-00-00-00A-023A-D_sx-US.XML", "DMC-HONAERO-EAA-00-00-00-00A-010A-D_sx-US.XML", "DMC-HONAERO-EAA-00-00-00-00A-012A-D_sx-US.XML", "DMC-HONAERO-EAA-00-00-00-00A-012B-D_sx-US.XML"]
        prelims = []
        proptary = pattern_proptary.findall(source)
        for p in zip(proptary,names):
            prelims.append(DataModule(p[0], "proptary", None, None, None, None, book_info, filename=p[1], infoname="Preliminary Information"))
        cpy = """<title>Copyright - Notice</title>
    <para>Copyright {} Honeywell International Inc. All rights
    reserved.</para>
    <para>Honeywell is a registered trademark of Honeywell International
    Inc.</para>
    <para>All other marks are owned by their respective companies.</para>""".format(book_info['cpyrgt'])
        prelims.append(DataModule(cpy, "cpy", None, None, None, None, book_info, filename="DMC-HONAERO-EAA-00-00-00-00A-021A-D_sx-US.XML", infoname="Copyright Information", pub=pub))
        return prelims

    def fill_tools_and_cons(content, st_content):
        nonlocal tab
        def sort_list(lst):
            num_list = []
            for li in lst[::-1]:
                if li[0] != "":
                    if li[0][0].isnumeric():
                        num_list.append(li)
                        lst.remove(li)
            lst.sort(key=lambda s: (s[0].lower(), s[1].lower()))
            if len(num_list) > 0:
                num_list.sort(key=lambda s: (s[0].lower(), s[1].lower()))
                lst = num_list + lst
            return lst
        
        def create_tools_table(tools):
            tool_list = []
            rows = ""
            for t in tools:
                cnt = re.sub(r'\n>', '>', t[1])
                cnt = re.sub(r'\n', ' ', cnt)
                pn = check_match(re.search(re.compile(r'<(?:tool|std)nbr>(.*?)</(?:tool|std)nbr>', re.DOTALL), cnt)).strip()
                nom = check_match(re.search(re.compile(r'<(?:tool|std)name>(.*?)</(?:tool|std)name>', re.DOTALL), cnt)).strip()
                src = check_match(re.search(re.compile(r'<(?:tool|std)src>(.*?)</(?:tool|std)src>', re.DOTALL), cnt)).strip()
                desc = check_match(re.search(re.compile(r'<(?:tool|std)desc>(.*?)</(?:tool|std)desc>', re.DOTALL), cnt)).strip()
                if (pn, nom, desc, src) in tool_list or nom == "": #Check for duplicates
                    continue
                tool_list.append((pn, nom, desc, src))
            tool_list = sort_list(tool_list)
            for t in tool_list:
                rows += "<row>\n<entry><para>{}</para></entry>\n<entry><para>{}{}</para></entry>\n<entry><para>{}</para></entry>\n</row>".format(t[0], t[1], (" ({})".format(t[2])) if t[2] != "" else "", t[3])
            return rows

        def create_cons_table(cons):
            cons_list = []
            rows = ""
            for c in cons:
                cnt = re.sub(r'\n>', '>', c)
                cnt = re.sub(r'\n', ' ', cnt)
                pn = check_match(re.search(re.compile(r'<connbr>(.*?)</connbr>', re.DOTALL), cnt)).strip()
                nom = check_match(re.search(re.compile(r'<conname>(.*?)</conname>', re.DOTALL), cnt)).strip()
                src = check_match(re.search(re.compile(r'<consrc>(.*?)</consrc>', re.DOTALL), cnt)).strip()
                desc = check_match(re.search(re.compile(r'<condesc>(.*?)</condesc>', re.DOTALL), cnt)).strip()
                if (pn, nom, desc, src) in cons_list or nom == "": #Check for duplicates
                    continue
                    
                cons_list.append((pn, nom, desc, src))
            
            cons_list = sort_list(cons_list)
            for c in cons_list:	
                rows += "<row>\n<entry><para>{}</para></entry>\n<entry><para>{}{}</para></entry>\n<entry><para>{}</para></entry>\n</row>".format(c[0], c[1], (" ({})".format(c[2])) if c[2] != "" else "", c[3])
            return rows
            
        def replace_tables(match):
            nonlocal tab
            t = tool_table if tab == True else cons_table
            orig_table_rows = match.group(0).count('<row>')
            new_table_rows = t.count('<row>')
            if orig_table_rows > new_table_rows:
                return match.group(0)
            value = "<tbody>\n{}\n</tbody>".format(t) if t.strip() else match.group(0)
            tab = False
            return value	
        
        tools = pattern_tool2.findall(content)
        tool_table = create_tools_table(tools)
        cons = pattern_cons2.findall(content)
        cons_table = create_cons_table(cons)
        tab = True
        return re.sub(pattern_tbody, replace_tables, st_content, 2)
            
    def GetPMCData():
        nonlocal pub
        nonlocal vcl
        nonlocal intro_key
        try: #Prelim Info
            pm_text = '<pmEntry pmEntryType="pmt77">\n<pmEntryTitle>Proprietary Information</pmEntryTitle>'
            prelims = GetPrelim()
            for p in prelims:
                front_matter.append(p)
                pm_text += p.dmref
                
            pm_text += '</pmEntry>'
        
            #Add Transmittal Info
            pm_text += '<pmEntry pmEntryType="pmt52">\n<pmEntryTitle>TRANSMITTAL INFORMATION</pmEntryTitle>\n'
            transltr_dm = GetTransmittalInfo()
            front_matter.append(transltr_dm)
            pm_text += transltr_dm.dmref
            pm_text += '</pmEntry>'
            
            #Add Record of Revisions
            pm_text += '<pmEntry pmEntryType="pmt53">\n<pmEntryTitle>RECORD OF REVISIONS</pmEntryTitle>\n'
            ror_dm = GetRecordOfRevisions()
            front_matter.append(ror_dm)
            pm_text += ror_dm.dmref
            pm_text += '</pmEntry>'
            
            #Add Record of Temporary Revisions
            pm_text += '<pmEntry pmEntryType="pmt54">\n<pmEntryTitle>RECORD OF TEMPORARY REVISIONS</pmEntryTitle>\n'
            rotr_dm = GetTemporaryRevision()
            front_matter.append(rotr_dm)
            pm_text += rotr_dm.dmref
            pm_text += '</pmEntry>'
            
            #Add SBL
            pm_text += '<pmEntry pmEntryType="pmt55">\n<pmEntryTitle>SERVICE BULLETIN LIST</pmEntryTitle>\n'
            sbl_dm = GetServiceBulletin()
            front_matter.append(sbl_dm)
            pm_text += sbl_dm.dmref
            pm_text += '</pmEntry>'
            pm_text += """<pmEntry authorityDocument="{}" pmEntryType="pmt56">
    <pmEntryTitle>LIST OF EFFECTIVE PAGES</pmEntryTitle>
    <externalPubRef authorityDocument="LEP"><externalPubRefIdent></externalPubRefIdent></externalPubRef>
    </pmEntry>""".format(book_info['ata'])
        except:
            print("Failed while adding front matter")
            print(traceback.format_exc())
            
        intro = re.findall(re.compile(r'<intro(?:duc)?([^>]*)>(.*?)</intro(?:duc)?>', re.DOTALL), source)
        intro = intro[0]
        intro_key = check_match(re.search(r'key="([^"]+)"', intro[0]))
        
        if len(intro) > 0:
            pm_text += """<pmEntry authorityDocument="{}" pmEntryType="pmt58">
    <pmEntryTitle>INTRODUCTION</pmEntryTitle>""".format(book_info['ata'])
            list1 = re.findall(re.compile(r'<(l1item|prcitem1)>\n?<(para|title)>(.*?)</\2>(.*?)</\1>', re.DOTALL), intro[1])
            for l in list1:
                pm_text += """<pmEntry>
        <pmEntryTitle>{}</pmEntryTitle>""".format(l[2])
                list2 = re.findall(re.compile(r'<(l2item|prcitem2)>\n?<(para|title)>(.*?)</\2>(.*?)</\1>', re.DOTALL), l[3])
                for l2 in list2:
                    content = "<title>{}</title>\n{}".format(l2[2], l2[3])
                    if l[2] == "Acronyms and Abbreviations":
                        content = create_acronym_table(content)
                    dm_obj = DataModule(content, "", None, None, page_blocks["0"][2], "0", book_info, infoname="INTRODUCTION", pub=pub)
                    introduction.append(dm_obj)
                    pm_text += dm_obj.dmref + "\n"
                pm_text += "</pmEntry>"
            pm_text += "</pmEntry>"
        chapters = pattern_chapter.findall(source)
        for chap in chapters:
            chap_title = chap[1]
            chapnbr = check_match(pattern_chapnbr.search(chap[0]))
            pm_text += '<pmEntry authorityDocument="{}">\n<pmEntryTitle>{}</pmEntryTitle>\n'.format(chapnbr, chap_title)
            sections = pattern_section.findall(chap[2])
            for sect in sections:
                sect_title = re.sub(re.compile(r'<acronym>\n?<acronymTerm>([^<]+).*?</acronym>', re.DOTALL), r'\1', sect[1])
                chapnbr = check_match(pattern_chapnbr.search(sect[0]))
                sectnbr = check_match(pattern_sectnbr.search(sect[0]))
                pm_text += '<pmEntry authorityDocument="{}">\n<pmEntryTitle>{}</pmEntryTitle>\n'.format("{}-{}".format(chapnbr, sectnbr), sect_title)
                subjects = pattern_subject.findall(sect[2])
                for s in subjects:
                    DataModule.info_codes = {"018": ['A', 1], "040":  ['A', 1], "420":  ['A', 1], "913":  ['A', 1], "520":  ['A', 1], "720":  ['A', 1],  "500":  ['A', 1],  "200":  ['A', 1],  "700":  ['A', 1], "300":  ['A', 1], "600": ['A', 1], "910":  ['A', 1], "800":  ['A', 1]}
                    subj_title = re.sub(re.compile(r'<acronym>\n?<acronymTerm>([^<]+).*?</acronym>', re.DOTALL), r'\1', s[1])
                    chapnbr = check_match(pattern_chapnbr.search(s[0]))
                    sectnbr = check_match(pattern_sectnbr.search(s[0]))
                    subjnbr = check_match(pattern_subjnbr.search(s[0]))
                    subj_key = check_match(re.search(r'key="([^"]+)"', s[0]))
                    subj = "{}-{}-{}".format(chapnbr, sectnbr, subjnbr)
                    pm_text += '<pmEntry authorityDocument="{}">\n<pmEntryTitle>{}</pmEntryTitle>\n'.format(subj, subj_title)
                    pgblks = pattern_pgblk.findall(s[2])
                    for pb in pgblks:
                        pb_key = check_match(re.search(r'key="([^"]+)"', pb[1]))
                        pgblknbr = check_match(re.search(r'pgblknbr="([^"]+)"', pb[1]))
                        title =  pb[3].upper() if pb[3] != "" else page_blocks[pgblknbr][0]
                        confnbr = check_match(re.search(r'confnbr="([^"]+)"', pb[1]))
                        content = pb[4]
                        pm_text += '<pmEntry pmEntryType="{}">\n<pmEntryTitle>{}</pmEntryTitle>\n'.format(page_blocks[pgblknbr][1], title)
                        new_pgblk = True
                        source_tasks = pattern_tasks.findall(content)
                        if content != "<isempty/>" and content != "" and len(source_tasks) > 0:
                            if len(source_tasks) > 0:
                                for i,task in enumerate(source_tasks):
                                    task_title = task[1]
                                    task_title = re.sub(re.compile(r'<acronym>\n?<acronymTerm>([^<]+).*?</acronym>', re.DOTALL), r'\1', task_title)
                                    task_key = check_match(re.search('key="([^"]+)"', task[0]))
                                    task_func = check_match(re.search('func="([^"]+)"', task[0]))
                                    task_seq = check_match(re.search('seq="([^"]+)"', task[0]))
                                    task_authdoc = "{}-{}-{}-A01".format(subj, task_func, task_seq)
                                    task_obj = Task(pb_key, task_key, task_authdoc, pgblknbr, task_title, subj_key)
                                    task_content = task[2]
                                    pm_text += "<pmEntry{}>\n<pmEntryTitle>{}</pmEntryTitle>\n".format(' authorityDocument="{}"'.format(task_obj.authdoc) if task_obj.authdoc != "" else "", task_title)
                                    subtasks = pattern_subtask.findall(task_content)
                                    for j,st in enumerate(subtasks):
                                        st_key = check_match(re.search('key="([^"]+)"', st[0]))
                                        st_func = check_match(re.search('func="([^"]+)"', st[0]))
                                        st_seq = check_match(re.search('seq="([^"]+)"', st[0]))
                                        st_authdoc = "{}-{}-{}-A01".format(subj, st_func, st_seq)
                                        st_content = st[1]
                                        if "<title>Job Setup Data</title>" in st_content:
                                            st_content = fill_tools_and_cons(content, st_content)
                                        dm_obj = DataModule(st_content, st_key, st_authdoc, task_obj, page_blocks[pgblknbr][2], pgblknbr, book_info, authName="{}.{}{}".format(i + 1, chr((ord('A') - 1) + j // 26) if j >= 26 else "", chr((ord('A') + j % 26))), infoname=title, pub=pub)
                                        pm_text += dm_obj.dmref + "\n"
                                        task_obj.children.append(dm_obj)
                                    tasks.append(task_obj)
                                    pm_text += "</pmEntry>\n"
                                
                        else:
                            dm_obj = DataModule(None, None, None, None, book_info, page_blocks[pgblknbr][2], pgblknbr, True)
                            not_applicable.append(dm_obj)
                            pm_text += dm_obj.dmref
                            pm_text += '\n</pmEntry>'
                            continue
                        pm_text += "\n</pmEntry>"
                    pm_text += "\n</pmEntry>"
                pm_text += "\n</pmEntry>"
            pm_text += "\n</pmEntry>"
        return pm_text

    def find_duplicate_id(obj, list):
        for item in list:
            if item.id == obj.id:
                return True
        else:
            return False
        
    def process_data_module(dm, schema):
        local_tools = []
        local_cons = []
        local_warn_caut = []
        
        def get_warncaut_ref(id):
            return """<warningRef id="{0}" warningIdentNumber="{0}"><dmRef><dmRefIdent><dmCode assyCode="{1}" disassyCode="00" 
    disassyCodeVariant="A" infoCode="012" infoCodeVariant="{6}" itemLocationCode="D" modelIdentCode="HON{2}" 
    subSubSystemCode="{3}" subSystemCode="{4}" systemCode="{5}" systemDiffCode="EAA"/></dmRefIdent></dmRef></warningRef>\n""".format(id, book_info['ata'][6:8], book_info['cage'], book_info['ata'][4], book_info['ata'][3], book_info['ata'][0:2], "A" if "warn" in id else "B")

        def replace_tool(match):
            tool = re.sub(r'\n', ' ', match.group(0))
            pn = check_match(re.search(re.compile(r'<(?:tool|std)nbr>(.*?)</(?:tool|std)nbr>', re.DOTALL), tool))
            nom = check_match(re.search(re.compile(r'<(?:tool|std)name>(.*?)</(?:tool|std)name>', re.DOTALL), tool))
            src = check_match(re.search(re.compile(r'<(?:tool|std)src>(.*?)</(?:tool|std)src>', re.DOTALL), tool))
            desc = check_match(re.search(re.compile(r'<(?:tool|std)desc>(.*?)</(?:tool|std)desc>', re.DOTALL), tool))
            new_tool = ToolAndCons(pn, re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.DOTALL), r'\1', nom), src, re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.DOTALL), r'\1', desc), "tool", book_info) #Create new tool object. ID is generated on creation

            #If that object doesn't exist as a tool in this data module yet, add it
            if not find_duplicate_id(new_tool, local_tools):
                local_tools.append(new_tool)
        
            return '<internalRef internalRefId="{}" internalRefTargetType="irtt05"></internalRef>'.format(new_tool.id)
            
        def replace_cons(match):
            cons = re.sub(r'\n', ' ', match.group(0))
            pn = check_match(re.search(re.compile(r'<connbr>(.*?)</connbr>', re.DOTALL), cons))
            nom = check_match(re.search(re.compile(r'<conname>(.*?)</conname>', re.DOTALL), cons))
            src = check_match(re.search(re.compile(r'<consrc>(.*?)</consrc>', re.DOTALL), cons))
            desc = check_match(re.search(re.compile(r'<condesc>(.*?)</condesc>', re.DOTALL), cons))
            new_cons = ToolAndCons(pn, re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.DOTALL), r'\1', nom), src, re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.DOTALL), r'\1', desc), "cons", book_info) #Create new cons object. ID is generated on creation

            #If that object doesn't exist as a tool in this data module yet, add it
            if not find_duplicate_id(new_cons, local_cons):
                local_cons.append(new_cons)
        
            return '<internalRef internalRefId="{}" internalRefTargetType="irtt04"></internalRef>'.format(new_cons.id)
        
        def replace_table_warn_caut(match):
            return "<{0}>{1}</{0}>".format(match.group(1), re.sub(r'<(/?)para>', r'<\1warningAndCautionPara>', match.group(2)))
        
        def replace_warn_and_caut(match):
            caut_ref = ""
            warn_ref = ""
            all = re.sub(r'\n', ' ', match.group(0))
            all = re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.DOTALL), r'\1', all)
            all = re.sub(r'<(/)?randomList>', r'<\1attentionRandomList>', all)
            all = re.sub(r'<(/)?listItem>', r'<\1attentionRandomListItem>', all)
            all = re.sub(r'(?<=<attentionRandomListItem>)\n?<para>(.*?)</para>', r'\n<attentionListItemPara>\1</attentionListItemPara>', all)
            all = re.sub('<(/?)para>', r'<\1warningAndCautionPara>', all)
            warn_caut = re.findall(re.compile(r'<(warning|caution)>(.*?)</\1>', re.DOTALL), all)
            for wc in warn_caut:
                #wc_text = re.sub(re.compile(r'<acronym>.*?<acronymTerm>(.*?)</.*?</acronym>', re.DOTALL), r'\1', wc[1])
                if wc[0] == "warning":
                    for k,w in warn.items():
                        if w == wc[1]:
                            id = k
                            break
                    else:
                        id = 'warn-{}'.format(add_leading_zero(len(warn) + 1, 4))
                        warn[id] = wc[1]
                    warn_ref += "{} ".format(id)
                else:
                    for k,c in caut.items():
                        if c == wc[1]:
                            id = k
                            break
                    else:
                        id = 'caut-{}'.format(add_leading_zero(len(caut) + 1, 4))
                        caut[id] = wc[1]
                    caut_ref += "{} ".format(id)
                    
                if id not in local_warn_caut:
                    local_warn_caut.append(id)
                    
            caut_ref = caut_ref.strip()
            warn_ref = warn_ref.strip()
            
            step = match.group(2)
            if caut_ref != "":
                if "cautionRefs=" in step:
                    step = re.sub(r'cautionRefs="([^"]+)"', r'cautionRefs="\1 {}"'.format(caut_ref), step)
                else:
                    step += ' cautionRefs="{}"'.format(caut_ref)
                    
            if warn_ref != "":
                if "warningRefs=" in step:
                    step = re.sub(r'warningRefs="([^"]+)"', r'warningRefs="\1 {}"'.format(warn_ref), step)
                else:
                    step += ' warningRefs="{}"'.format(warn_ref)
                    
            return '{}\n<{}>'.format(match.group(1) if match.group(1) is not None else "", step)
                
        def link_graphic_references(match): #Return a cross reference to a figure
            gid = match.group(1)
            if gid in dm.graph_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt01"></internalRef>'.format(gid)
            elif gid in dm.sheet_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt09"></internalRef>'.format(gid)
            else:
                for t2 in tasks:
                    for dm2 in t2.children:
                        if dm2 != dm:
                            if gid in dm2.graph_refs:
                                return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(gid), dm2.dmref)
                            elif gid in dm2.sheet_refs:
                                return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[gid][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[gid][1])) #Will return a reference to the figure, not the sheet
                else:
                    for dm2 in introduction:
                        if gid in dm2.graph_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(gid), dm2.dmref)
                        elif gid in dm2.sheet_refs:
                            return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[gid][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[gid][1])) #Will return a reference to the figure, not the sheet
                    else:
                        return match.group(0)
                            
        def ref_pgblk(match, do_return=False): #Return the first child of the first task with a parent pgblk key matching the refid
            nonlocal intro_key
            for tsk in tasks:
                if len(tsk.children) > 0:
                    if tsk.parent_key == match.group(1):
                        return re.sub(' authorityDocument="[^"]+"', '', tsk.children[0].dmref)
                    elif tsk.subj_key == match.group(1):
                        return re.sub(' authorityDocument="[^"]+"', ' authorityName="{}"'.format(tsk.subj_num), tsk.children[0].dmref)
            else:
                if intro_key == match.group(1):
                    return re.sub(' authorityDocument="[^"]+"', '', introduction[0].dmref)
                elif not do_return:
                    return ref_int(match, True)
                else:
                    return match.group(0)
        
        def ref_int(match, do_return=False):
            refid = match.group(1)
            if refid in dm.graph_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt01"></internalRef>'.format(refid)
            elif refid in dm.table_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt02"></internalRef>'.format(refid)
            elif refid in dm.sheet_refs:
                return '<internalRef internalRefId="{}" internalRefTargetType="irtt09"></internalRef>'.format(refid)
            elif refid == dm.key:
                return re.sub(' authorityDocument="[^"]+"', (' authorityName="{}"'.format(dm.authname)) if dm.authname is not None else '', dm.dmref) 	
            
            for tsk in tasks:
                if tsk.key == refid:
                    authname = tsk.children[0].authname.split('.')[0] if tsk.children[0].authname is not None else ''
                    if authname != '':
                        authname = "{}. {}".format(authname, tsk.title)
                    return re.sub(' authorityDocument="[^"]+"', ' authorityName="{}"'.format(authname), tsk.children[0].dmref)
                for dm2 in tsk.children:
                    if dm == dm2:
                        continue
                    if dm2.key == refid:
                        return re.sub(' authorityDocument="[^"]+"', (' authorityName="{}"'.format(dm2.authname)) if dm2.authname is not None else '', dm2.dmref) 	
                    elif refid in dm2.graph_refs or refid in dm2.table_refs:
                        return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(refid), dm2.dmref)
                    elif refid in dm2.sheet_refs:
                        return re.sub('authorityDocument="[^"]+"', 'referredFragment="{}"'.format(dm2.sheet_refs[gid][0]), dm2.dmref + ', Sheet {}'.format(dm2.sheet_refs[gid][1]))
            else:
                if not do_return:
                    return ref_pgblk(match, True)
                else:
                    return match.group(0)
        
        def delegate_refint(match):
            if 'reftype="pgblk"' in match.group(0) or 'reftype="ipl"' in match.group(0):
                return ref_pgblk(match)
            else:
                return ref_int(match)
                
        dm.content = re.sub(pattern_grphcref, link_graphic_references, dm.content)
        dm.content = re.sub(pattern_refint, delegate_refint, dm.content)
        #dm.content = re.sub(re.compile(r'<entry>\n?(<dmRef.*?</dmRef>)', re.DOTALL), r'<entry>\n<para>\1</para>'
        #Replace tools, create header data
        dm.content = re.sub(pattern_tool, replace_tool, dm.content)
        for tool in local_tools:
            dm.tool_data += tool.sed
            
        #Replace tools, create header data
        dm.content = re.sub(pattern_cons, replace_cons, dm.content)
        for cons in local_cons:
            dm.cons_data += cons.sed
        #source = re.sub(r'<l(\d)item>\n?</l\1item>', '', source) #Found in OHM
        dm.content = re.sub(r'<(/?)prcitem\d>', r'<\1proceduralStep>' if schema != "descript" else r'<\1levelledPara>', dm.content)
        dm.content = re.sub(r'<(/?)l\ditem>', r'<\1proceduralStep>' if schema != "descript" else r'<\1levelledPara>', dm.content)
        while True:
            fix1 = re.subn(pattern_fixsteps, r'\2\n\1', dm.content)
            dm.content = fix1[0]
            fix2 = re.subn(pattern_fixsteps_list, r'\2\n\1', dm.content)
            dm.content = fix2[0]
            fix3 = re.subn(pattern_fixsteps_tabfig, r'\2\n\1', dm.content)
            dm.content = fix3[0]
            if fix1[1] + fix2[1] + fix3[1] == 0:
                break
        
        dm.content = re.sub(pattern_cautwarn, replace_warn_and_caut, dm.content)
        dm.content = re.sub(pattern_cautwarn2, replace_warn_and_caut, dm.content) #This one deals with warns/cauts at the top of the module 
        
        dm.content = re.sub(re.compile(r'<(warning|caution)>(.*?)</\1>', re.DOTALL), replace_table_warn_caut, dm.content)
        for id in local_warn_caut:
            dm.warn_and_caut_data += get_warncaut_ref(id)

    def create_acronym_table(dm_content):
        nonlocal acros_and_abbrs
        sorted_list = sorted(acros_and_abbrs, key=lambda s: s[0].lower())
        table = """<tgroup cols="2">
    <?PubTbl tgroup dispwid="605.00px"?>
    <colspec colname="col1" colwidth="0.39*"/><colspec colname="col2" colwidth="1.61*"/><tbody><row>
    <entry>
    <para>Term</para>
    </entry>
    <entry>
    <para>Full Term</para>
    </entry>
    </row>"""
        for a in sorted_list:
            table += """<row>
    <entry>
    <para>{}</para>
    </entry>
    <entry>
    <para>{}</para>
    </entry>
    </row>""".format(a[0], a[1])
        table += "</tbody></tgroup>"
        return re.sub(re.compile("<tgroup.*?</tgroup>", re.DOTALL), table, dm_content, 1)
        
    #String where we log all our log_print() messages so we can later write them to log file if desired.
    source_file = Path(file)
    dir = Path(directory)
    
    try:
            
        try: #Process the source data before reading it
            print("\nProcessing source data...", end='\r')
            source =  process_source(source_file)
        except:
            log_print("Processing source data... Error!")
            print(traceback.format_exc())
        else:
            log_print("Processing source data... Done!")
        
        try:
            print("\nGetting book info...", end='\r')
            book_info = GetBookInfo(source)
        except:
            log_print("Getting book info... Error!")
            print(traceback.format_exc())
        else:
            log_print("Getting book info... Done!")

        try:
            print("\nProcessing data modules...", end='\r')
            pmc = create_pmc(GetPMCData())
        except:
            log_print("Processing data modules... Error!")
            print(traceback.format_exc())
        else:
            log_print("Processing data modules... Done!")
        
        try:
            print("\nCreating data modules...", end='\r')
            #Create the Data Modules
            for t in tasks:
                for dm in t.children:		
                    try:
                        #Decide whether to create a Descriptive or Procedural module
                        if (page_blocks[dm.parent.pgblk][3] == "D" and dm.parent.title != "Planning Data") or dm.parent.pgblk == "0":
                            process_data_module(dm, "descript")
                            module = create_descript(dm)
                        else:
                            process_data_module(dm, "proced")
                            module = create_proced(dm)
                            
                        (dir /dm.filename).write_text(module, encoding='utf-8')	
                    except:
                        log_print("Failed to create {} data module.".format(dm.filename))
                        print(traceback.format_exc(), True)

            #Create Front Matter Data Modules
            for f in front_matter:
                try:
                    process_data_module(f, "descript")
                    (dir /f.filename).write_text(create_descript(f), encoding='utf-8')
                except:
                    log_print("Failed to create {} data module".format(f.filename))
                    
            for i in introduction:
                try:
                    process_data_module(i, "descript")
                    (dir /i.filename).write_text(create_descript(i), encoding='utf-8')
                except:
                    log_print("Failed to create {} data module".format(i.filename))
                    
            #Create N/A Data Modules
            for n in not_applicable:
                try:
                    (dir /n.filename).write_text(create_proced(n), encoding='utf-8')	
                except:
                    log_print("Failed to create {} data module".format(n.filename))
            try:
                if vcl is not None:
                    (dir /vcl.filename).write_text(create_descript(vcl, add_levelled_para=False), encoding='utf-8')
            except:
                log_print("Failed to create 018Z data module")
            
            #Create CIRs
            try:
                (dir /"DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "00NA")).write_text(create_cir("tool", ToolAndCons.tools), encoding='utf-8')
            except:
                log_print("Failed to create 00NA data module")
                
            try:
                (dir /"DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "00LA")).write_text(create_cir("cons", ToolAndCons.cons), encoding='utf-8')
            except:
                log_print("Failed to create 00LA data module")
                
            try:
                (dir /"DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "012A")).write_text(create_cir("warn", warn), encoding='utf-8')
            except:
                log_print("Failed to create 012A data module")
                
            try:
                (dir /"DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "012B")).write_text(create_cir("caut", caut), encoding='utf-8')
            except:
                log_print("Failed to create 012B data module")
            
            try:
                #Create 00P
                filename_00P = "DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "00PA")
                (dir /filename_00P).write_text(create_prdcrossreftable(filename_00P), encoding='utf-8')
            except:
                log_print("Failed to create 00PA data module")
                
            try:
                #Create 00W
                filename_00W = "DMC-HON{}-EAA-{}-00A-{}-D_sx-US.XML".format(book_info['cage'], book_info['ata'], "00WA")
                (dir /filename_00W).write_text(create_appliccrossreftable(filename_00W), encoding='utf-8')	
            except:
                log_print("Failed to create 00WA data module")
                
        except:
            log_print("Creating data modules... Error!")
            print(traceback.format_exc())
        else:
            log_print("Creating data modules... Done!")
            
        try:
            print("\nCreating PMC...", end='\r')
            pmc = re.sub(r'&([A-Z]+);', to_lower, pmc)
            (dir /'PMC-HON{0}-{0}-00001-01_sx-US.XML'.format(book_info['cage'])).write_text(pmc, encoding='utf-8')	
        except:
            log_print("Creating PMC... Error!")
            print(traceback.format_exc())
        else:
            log_print("Creating PMC... Done!")	
    except:
        log_print(traceback.format_exc())
        
    log_print("\nDone!")
    #If we reach this point, run the exit function
    return exit_handler()

if __name__ == "__main__":
    EMConvert(get_source(), Path('.'), None)