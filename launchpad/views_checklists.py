from launchpad.checklists.dpmo_export import dpmo_table_creation
import os
from PyPDF2 import PdfFileWriter, PdfFileReader
from PyPDF2.generic import BooleanObject, NameObject, IndirectObject
import json
from datetime import datetime
import calendar
from flask import (abort, flash, redirect, render_template, request,
                   url_for)
from flask_login import current_user, login_required
from launchpad.views import forbidden
from launchpad.checklists.table_index import TableCreation
from launchpad.checklists.view_tables import TableViews
from dateutil import tz
from launchpad.checklists.checklist_text import (dc_blocks, dc_length,
                                                 writer_blocks,
                                                 writer, writer_length,
                                                 illustration_blocks,
                                                 illustration, illustration_length,
                                                 editor_blocks, editor, editor_length,
                                                 qa_blocks_1, qa_blocks_2, qa_blocks_3,
                                                 qa_blocks, qa, qa_length, dl_blocks, dl,
                                                 dl_length, dpmo_blocks, dpmo, qa_note,
                                                 dpmo_length, dpmo_exclusion,
                                                 writer_A, writer_B, writer_C,
                                                 writer_D, writer_E, writer_F, writer_G,
                                                 illustration_A, editor_A, editor_B,
                                                 editor_C, editor_D, editor_E, editor_F,
                                                 editor_G, editor_H, editor_I, editor_J,
                                                 qa_A, qa_B, qa_C, qa_D, qa_E,
                                                 qa_G, qa_H, qa_I,
                                                 dl_A, dl_B, dl_C, dl_D, dl_E,
                                                 dl_F, dl_G, dl_H,
                                                 dpmo_A, dpmo_B, dpmo_C,
                                                 dpmo_D, dpmo_E, dpmo_F,
                                                 dpmo_G, dpmo_H, dpmo_I, dpmo_J, dpmo_K,
                                                 dpmo_L, dpmo_M, dpmo_N, dpmo_O, dpmo_P,
                                                 dc_1, dc_2, dc_3, dc_4, dc_5, dc_6, dc_7,
                                                 dc_8, dc_9, dc_10, trans_english, trans_french)
from launchpad import APP, DB, SOCKETIO
from launchpad.forms import (JobNo, JobCreation, QAUpdate,
                             ChecklistManager, TranslationUpdate,
                             DCUpdate, WriterUpdate,
                             IllustrationUpdate, EditorUpdate,
                             FinalDeliveryUpdate, DPMOUpdate,
                             JobUpdate, DPMOManager, ChecklistHistory)
from launchpad.models import User, Jobs, Checklists
import pandas
from pandas import DataFrame
import plotly
import plotly.graph_objects as go
import openpyxl
from openpyxl import load_workbook

# All checklist views

@APP.route('/checklists_home', methods=['GET', 'POST'])
@login_required
def checklists_home():
    return render_template('checklists_home.html')

# Future JIRA import
@APP.route('/job_search', methods=['GET', 'POST'])
@login_required
def job_search():
    form = JobNo()
    if request.method == 'POST':
        if form.validate_on_submit():
            job_no = request.form['job_no']
            job = Jobs.query.filter_by(job_no=job_no).first()
            if job is None:
                flash("Please input job information")
                return redirect(url_for('job_creation', job_no=job_no))
            else:
                return redirect(url_for('checklist_index', job_no=job_no))
        else:
            return render_template('job_search.html', form=form)
    else:
        return render_template('job_search.html', form=form)


# Future JIRA import
@APP.route('/job_update/<job_no>', methods=['GET', 'POST'])
@login_required
def job_update(job_no):
    form = JobUpdate()
    job = Jobs.query.filter_by(job_no=job_no).first()
    form.ata_number.data = job.ata_number
    form.d_number.data = job.d_number
    form.project.data = job.project
    form.type_of_publication.data = job.type_of_publication
    form.software.data = job.software
    form.type_of_job.data = job.type_of_job
    form.location_of_authoring.data = job.location_of_authoring
    if request.method == 'POST':
        if form.validate_on_submit():
            d_number_input = request.form['d_number']
            ata_number_input = request.form['ata_number']
            type_of_publication_input = request.form['type_of_publication']
            software_input = request.form['software']
            type_of_job_input = request.form['type_of_job']
            project_input = request.form['project']
            location_of_authoring_input = request.form['location_of_authoring']

            job.ata_number = ata_number_input
            job.d_number = d_number_input
            job.type_of_publication = type_of_publication_input
            job.software = software_input
            job.type_of_job = type_of_job_input
            job.project = project_input
            job.location_of_authoring = location_of_authoring_input
            DB.session.commit()
            flash('Job Information has been Updated')
            return redirect(url_for('checklist_index', job_no=job_no))
        else:
            return render_template('job_update.html',
                                   job_no=job_no, form=form)
    else:
        return render_template('job_update.html', job_no=job_no, form=form)

# Future JIRA import
@APP.route('/job_creation/<job_no>', methods=['GET', 'POST'])
@login_required
def job_creation(job_no):
    form = JobCreation()
    if request.method == 'POST':
        if form.validate_on_submit():
            d_number = request.form['d_number']
            ata_number = request.form['ata_number']
            type_of_publication = request.form['type_of_publication']
            software = request.form['software']
            type_of_job = request.form['type_of_job']
            project = request.form['project']
            location_of_authoring = request.form['location_of_authoring']
            new_job = Jobs(job_no=job_no, stage="",
                           ata_number=ata_number,
                           d_number=d_number,
                           project=project,
                           type_of_publication=type_of_publication,
                           software=software,
                           type_of_job=type_of_job,
                           location_of_authoring=location_of_authoring)
            DB.session.add(new_job)
            DB.session.commit()
            flash('A new job has been created!')
            return redirect(url_for('checklist_index', job_no=job_no))
        else:
            return render_template('job_creation.html',
                                   job_no=job_no, form=form)
    else:
        return render_template('job_creation.html', job_no=job_no, form=form)

# Future JIRA import
@APP.route('/checklist_index/<job_no>', methods=['GET', 'POST'])
@login_required
def checklist_index(job_no):
    job = Jobs.query.filter_by(job_no=job_no).first()
    project = job.project
    ata_number = job.ata_number
    d_number = job.d_number
    type_of_publication = job.type_of_publication
    software = job.software
    type_of_job = job.type_of_job
    location_of_authoring = job.location_of_authoring
    return render_template('checklist_index.html', job_no=job_no,
                           project=project, ata_number=ata_number, d_number=d_number,
                           type_of_publication=type_of_publication, software=software,
                           type_of_job=type_of_job, location_of_authoring=location_of_authoring)

# Helper Functions


def test_number(num):
    if num == "":
        return ""
    if int(num) < 0:
        return str(0)
    return str(int(num))


def delete_fn(job_no, delete_rev, checklist_type):
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type, revision=delete_rev).first()
    DB.session.delete(cl)
    DB.session.commit()
    # Fix all other revisions
    test = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter(Checklists.revision > delete_rev).all()
    if len(test) != 0:
        for checklist in test:
            checklist.revision -= 1
    DB.session.commit()
    flash('Checklist Deleted!')

# Data Conversion Checklists


@APP.route('/data_conversion_index/<job_no>', methods=['GET', 'POST'])
@login_required
def data_conversion_index(job_no):
    checklist_type = 'Data Conversion'
    checklist_list = Checklists.query.\
        filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    for x in checklist_list:
        rev.append(x.revision)
    max_revision = max(rev, default=0)
    row_content_list = []
    for r in rev:
        cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=r).first()
        checklist_data = json.loads(cl.data)
        stored_data = list(checklist_data.values())
        new_row = [cl.completed, cl.created_by]  # Completed, Created by
        day = cl.date_created.replace(tzinfo=tz.gettz('UTC'))
        new_row += [day.astimezone(tz.tzlocal()).strftime("%Y-%m-%d %H:%M:%S")]
        new_row += [cl.er]
        new_row += [stored_data[0]]  # Pages reviewed
        new_row += stored_data[1:5]  # Error totals and calculations
        new_row += stored_data[5:15]  # Error counts
        row_content_list += [new_row]
    table = TableCreation(checklist_type, max_revision,
                          row_content_list, job_no)
    return render_template('data_conversion_index.html',
                           table=table, job_no=job_no)

@APP.route('/data_conversion_index/data_conversion_update/<job_no>/<update_rev>',
           methods=['GET', 'POST'])
@login_required
def data_conversion_update(job_no, update_rev):
    checklist_type = 'Data Conversion'
    form = DCUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter_by(revision=update_rev).first()
    comment_name = 'Errors Count'
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    form.pages_reviewed.data = stored_data[0]
    form.errorcount1.data = stored_data[5]
    form.errorcount2.data = stored_data[6]
    form.errorcount3.data = stored_data[7]
    form.errorcount4.data = stored_data[8]
    form.errorcount5.data = stored_data[9]
    form.errorcount6.data = stored_data[10]
    form.errorcount7.data = stored_data[11]
    form.errorcount8.data = stored_data[12]
    form.errorcount9.data = stored_data[13]
    form.errorcount10.data = stored_data[14]
    form.er_level.data = cl.er
    form.op_manager.data = cl.op_manager

    if request.method == 'POST':
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        completion_test = True
        er = request.form['er_level']
        er = test_number(er)
        pages_input = request.form['pages_reviewed']
        pages_input = test_number(pages_input)
        if pages_input == "" or er == "":
            completion_test = False
        data = {"Pages Reviewed": pages_input, 'Total Major Errors': 0,
                'Total Minor Errors': 0,
                'Item 10 References as a Major PPM (<7000)': 0,
                'Item 10 References as a Minor PPM (<7000)': 0}
        error_major = 0
        error_minor = 0
        dc_di = []
        for x, y in dc_blocks.items():
            for k, v in y.items():
                dc_di += [k]
        error_list = [request.form['errorcount1'],
                      request.form['errorcount2'],
                      request.form['errorcount3'],
                      request.form['errorcount4'],
                      request.form['errorcount5'],
                      request.form['errorcount6'],
                      request.form['errorcount7'],
                      request.form['errorcount8'],
                      request.form['errorcount9'],
                      request.form['errorcount10']]
        for subject, di in zip(range(dc_length), dc_di):
            error = error_list[subject]
            error = test_number(error)
            data.update({subject: error})
            if error == "":
                completion_test = False
            if di == 'Major':
                if error == "":
                    error_major += 0
                else:
                    error_major += int(error)
            if di == 'Minor':
                if error == "":
                    error_minor += 0
                else:
                    error_minor += int(error)
        data.update({'Total Major Errors': error_major})
        data.update({'Total Minor Errors': error_minor})
        if pages_input == "":
            item_10_major = 0
            item_10_minor = 0
        else:
            if int(pages_input) == 0:
                item_10_major = 0
                item_10_minor = 0

            else:
                item_10_major = int((error_major/int(pages_input))*1000000)
                item_10_minor = int((error_minor/int(pages_input))*1000000)
        data.update({'Item 10 References as a Major PPM (<7000)':
                     item_10_major})
        data.update({'Item 10 References as a Minor PPM (<7000)':
                     item_10_minor})
        JSON_data = json.dumps(data)
        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('data_conversion_update.html',
                                        comment_name=comment_name,
                                        revision=update_rev, job_no=job_no,
                                        form=form, blocks=dc_blocks,
                                        dc_1=dc_1, dc_2=dc_2, dc_3=dc_3,
                                        dc_4=dc_4, dc_5=dc_5,
                                        dc_6=dc_6, dc_7=dc_7, dc_8=dc_8,
                                        dc_9=dc_9, dc_10=dc_10)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False

        history = cl.checklist_history
        parsed = json.loads(history)
        num_list = []
        for x in list(parsed.keys()):
            num_list += [int(x)]
        num = max(num_list)+1
        parsed.update({num: ['Edit', current_user.username, str(datetime.utcnow()), completed]})
        cl.checklist_history = json.dumps(parsed)
        cl.er = er
        cl.completed = completed
        cl.data = JSON_data
        cl.op_manager = op_manager
        DB.session.commit()
        flash('Data Conversion Checklist Updated!')
        return redirect(url_for('data_conversion_index', job_no=job_no))
    else:
        return render_template('data_conversion_update.html',
                               comment_name=comment_name,
                               revision=update_rev, job_no=job_no,
                               form=form, blocks=dc_blocks,
                               dc_1=dc_1, dc_2=dc_2, dc_3=dc_3,
                               dc_4=dc_4, dc_5=dc_5,
                               dc_6=dc_6, dc_7=dc_7, dc_8=dc_8,
                               dc_9=dc_9, dc_10=dc_10)


@APP.route('/data_conversion_add/<job_no>', methods=['GET', 'POST'])
@login_required
def data_conversion_add(job_no):
    checklist_type = 'Data Conversion'
    form = DCUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    eng_review = []
    for x in cl:
        rev.append(x.revision)
        eng_review.append(x.er)
    comment_name = 'Errors Count'
    revision = max(rev, default=0)+1
    eng_review_inter = [x for x in eng_review if x != '']
    eng_review = max(eng_review_inter, default=0)

    if len(cl) == 0 or len(eng_review_inter) == 0:
        form.er_level.data = 0
    else:
        form.er_level.data = int(eng_review) + 1

    if request.method == 'POST':
        completion_test = True
        job = Jobs.query.filter_by(job_no=job_no).first()
        pages_input = request.form['pages_reviewed']
        pages_input = test_number(pages_input)
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False

        er = request.form['er_level']
        er = test_number(er)
        if er == "" or pages_input == "":
            completion_test = False
        data = {"Pages Reviewed": pages_input, 'Total Major Errors': 0,
                'Total Minor Errors': 0,
                'Item 10 References as a Major PPM (<7000)': 0,
                'Item 10 References as a Minor PPM (<7000)': 0}
        error_major = 0
        error_minor = 0
        dc_di = []
        for x, y in dc_blocks.items():
            for k, v in y.items():
                dc_di += [k]
        error_list = [request.form['errorcount1'],
                      request.form['errorcount2'],
                      request.form['errorcount3'],
                      request.form['errorcount4'],
                      request.form['errorcount5'],
                      request.form['errorcount6'],
                      request.form['errorcount7'],
                      request.form['errorcount8'],
                      request.form['errorcount9'],
                      request.form['errorcount10']]
        for subject, di in zip(range(dc_length), dc_di):
            error = error_list[subject]
            error = test_number(error)
            data.update({subject: error})
            if error == "":
                completion_test = False
            if di == 'Major':
                if error == "":
                    error_major += 0
                else:
                    error_major += int(error)
            if di == 'Minor':
                if error == "":
                    error_minor += 0
                else:
                    error_minor += int(error)
        data.update({'Total Major Errors': error_major})
        data.update({'Total Minor Errors': error_minor})
        if pages_input == "":
            item_10_major = 0
            item_10_minor = 0
        else:
            if int(pages_input) == 0:
                item_10_major = 0
                item_10_minor = 0

            else:
                item_10_major = int((error_major/int(pages_input))*1000000)
                item_10_minor = int((error_minor/int(pages_input))*1000000)

        data.update({'Item 10 References as a Major PPM (<7000)':
                     item_10_major})
        data.update({'Item 10 References as a Minor PPM (<7000)':
                     item_10_minor})
        JSON_data = json.dumps(data)
        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('data_conversion_update.html', revision=revision,
                               job_no=job_no, form=form,
                               dc_1=dc_1, dc_2=dc_2, dc_3=dc_3,
                               dc_4=dc_4, dc_5=dc_5,
                               dc_6=dc_6, dc_7=dc_7,
                               dc_8=dc_8, dc_9=dc_9,
                               dc_10=dc_10, comment_name=comment_name)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        history = {1: ['Create', current_user.username, str(datetime.utcnow()), completed]}
        json_history = json.dumps(history)
        new_checklist = Checklists(checklist_type=checklist_type,
                                   completed=completed,
                                   created_by=current_user.username,
                                   date_created=datetime.utcnow(),
                                   er=er, op_manager=op_manager,
                                   revision=revision,
                                   data=JSON_data,
                                   checklist_history=json_history)
        job.checklists.append(new_checklist)
        DB.session.add(new_checklist)
        DB.session.commit()
        flash('New Data Conversion Checklist Created!')
        return redirect(url_for('data_conversion_index', job_no=job_no))
    else:
        return render_template('data_conversion_update.html', revision=revision,
                               job_no=job_no, form=form,
                               dc_1=dc_1, dc_2=dc_2, dc_3=dc_3,
                               dc_4=dc_4, dc_5=dc_5,
                               dc_6=dc_6, dc_7=dc_7,
                               dc_8=dc_8, dc_9=dc_9,
                               dc_10=dc_10, comment_name=comment_name)

@APP.route('/data_conversion_index/data_conversion_delete/<job_no>/<delete_rev>/<checklist_type>',
           methods=['GET', 'POST'])
@login_required
def data_conversion_delete(job_no, delete_rev, checklist_type):
    if not current_user.admin:
        flash('Only Admins can delete checklists')
        return redirect(url_for('data_conversion_index', job_no=job_no))
    else:
        delete_fn(job_no, delete_rev, checklist_type)
        return redirect(url_for('data_conversion_index', job_no=job_no))
# Standard Checklists

# Writer Checklists
@APP.route('/writer_index/<job_no>', methods=['GET', 'POST'])
@login_required
def writer_index(job_no):
    checklist_type = 'Writer'
    checklist_list = Checklists.query.\
        filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    for x in checklist_list:
        rev.append(x.revision)
    max_revision = max(rev, default=0)
    row_content_list = []
    for r in rev:
        cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=r).first()
        checklist_data = json.loads(cl.data)
        stored_data = list(checklist_data.values())
        new_row = [cl.completed, cl.created_by]

        day = cl.date_created.replace(tzinfo=tz.gettz('UTC'))
        new_row += [day.astimezone(tz.tzlocal()).strftime("%Y-%m-%d %H:%M:%S")]
        new_row += [cl.er]  # Engineering Review Level/ Customer Review
        new_row += [stored_data[1]]  # writer
        new_row += [stored_data[2]]  # date
        new_row += [stored_data[0]]  # changed_pages
        stored_2 = checklist_data.values()
        data_2 = []
        for x in list(stored_2)[3:]:
            data_2 += x.values()
        new_row += data_2
        row_content_list += [new_row]

    table = TableCreation(checklist_type, max_revision,
                          row_content_list, job_no)
    return render_template('writer_index.html', table=table, job_no=job_no)


@APP.route('/writer_add/<job_no>', methods=['GET', 'POST'])
@login_required
def writer_add(job_no):
    form = WriterUpdate()
    form.writer.data = current_user.fname + ' ' + current_user.lname
    form.date.data = datetime.now()
    checklist_type = 'Writer'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no))\
        .filter_by(checklist_type=checklist_type).all()
    rev = []
    eng_review = []
    for x in cl:
        rev.append(x.revision)
        eng_review.append(x.er)
    revision = max(rev, default=0)+1
    eng_review_inter = [x for x in eng_review if x != '']
    eng_review = max(eng_review_inter, default=0)
    if len(cl) == 0 or len(eng_review_inter) == 0:
        form.er_level.data = 0
        comment_name = 'Pre-Delivery Comments'
    else:
        form.er_level.data = int(eng_review) + 1
        comment_name = 'Engineering Review Comments #'+str(int(eng_review)+1)

    if request.method == "POST":
        completion_test = True
        er = request.form['er_level']
        er = test_number(er)
        pages_input = request.form['changed_pages_no']
        pages_input = test_number(pages_input)
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False

        if er == "" or pages_input == "":
            completion_test = False

        writer_input = request.form['writer']
        if writer_input == "":
            writer_input = current_user.fname + ' ' + current_user.lname
        date_input = request.form['date']
        data = {'Total Number of Change Pages': pages_input,
                'Writer': writer_input, 'Date': date_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30'],
                       request.form['input31'], request.form['input32'],
                       request.form['input33'], request.form['input34'],
                       request.form['input35'],
                       request.form['input36'], request.form['input37'],
                       request.form['input38'], request.form['input39'],
                       request.form['input40'],
                       request.form['input41'], request.form['input42'],
                       request.form['input43'], request.form['input44'],
                       request.form['input45'],
                       request.form['input46'], request.form['input47'],
                       request.form['input48'], request.form['input49'],
                       request.form['input50'],
                       request.form['input51'], request.form['input52'],
                       request.form['input53'], request.form['input54'],
                       request.form['input55']]
        for subject in range(len(inputs_list)):
            data.update({subject: {comment_name: inputs_list[subject]}})
            if inputs_list[subject] == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)
        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('writer_update.html',
                               revision=revision,
                               comment_name=comment_name,
                               job_no=job_no, form=form,
                               writer_A=writer_A, writer_B=writer_B,
                               writer_C=writer_C, writer_D=writer_D,
                               writer_E=writer_E,
                               writer_F=writer_F, writer_G=writer_G)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        history = {1: ['Create', current_user.username, str(datetime.utcnow()), completed]}
        json_history = json.dumps(history)
        new_checklist = Checklists(checklist_type=checklist_type,
                                    completed=completed, op_manager=op_manager,
                                   created_by=current_user.username,
                                   date_created=datetime.utcnow(),
                                   er=er, checklist_history=json_history,
                                   revision=revision,
                                   data=JSON_data)
        job = Jobs.query.filter_by(job_no=job_no).first()
        job.checklists.append(new_checklist)
        DB.session.add(new_checklist)
        DB.session.commit()
        flash('New Writer Checklist Created!')
        return redirect(url_for('writer_index', job_no=job_no))
    else:
        return render_template('writer_update.html',
                               revision=revision,
                               comment_name=comment_name,
                               job_no=job_no, form=form,
                               writer_A=writer_A, writer_B=writer_B,
                               writer_C=writer_C, writer_D=writer_D,
                               writer_E=writer_E,
                               writer_F=writer_F, writer_G=writer_G)


@APP.route('/writer_index/writer_update/<job_no>/<update_rev>',
           methods=['GET', 'POST'])
@login_required
def writer_update(job_no, update_rev):
    checklist_type = 'Writer'
    form = WriterUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter_by(revision=update_rev).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    if cl.created_by == 'JIRA WEBHOOK':
        cl.created_by = current_user.username
        cl.date_created = datetime.utcnow()
        form.writer.data = current_user.fname + ' ' + current_user.lname
        form.date.data = datetime.now().date()
    else:
        form.writer.data = stored_data[1]
        form.date.data = datetime.strptime(stored_data[2], '%Y-%m-%d')
    eng_review = cl.er
    form.er_level.data = eng_review
    form.changed_pages_no.data = stored_data[0]
    form.op_manager.data = cl.op_manager
    input_2 = [form.input1, form.input2, form.input3,
               form.input4, form.input5,
               form.input6, form.input7, form.input8,
               form.input9, form.input10,
               form.input11, form.input12, form.input13,
               form.input14, form.input15,
               form.input16, form.input17, form.input18,
               form.input19, form.input20,
               form.input21, form.input22, form.input23,
               form.input24, form.input25,
               form.input26, form.input27, form.input28,
               form.input29, form.input30,
               form.input31, form.input32, form.input33,
               form.input34, form.input35,
               form.input36, form.input37, form.input38,
               form.input39, form.input40,
               form.input41, form.input42, form.input43,
               form.input44, form.input45,
               form.input46, form.input47, form.input48,
               form.input49, form.input50,
               form.input51, form.input52, form.input53,
               form.input54, form.input55]
    for x in range(len(input_2)):
        input_2[x].data = str(list(stored_data[3+x].values())[0])
    if eng_review == "":
        comment_name = 'Comments'
    else:
        if int(eng_review) == 0:
            comment_name = 'Pre-Delivery Comments'
        else:
            comment_name = 'Engineering Review Comments #'+str(eng_review)

    if request.method == "POST":
        completion_test = True

        er = request.form['er_level']
        er = test_number(er)
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        pages_input = request.form['changed_pages_no']
        pages_input = test_number(pages_input)
        if er == "" or pages_input == "":
            completion_test = False
        writer_input = request.form['writer']
        if writer_input == "":
            writer_input = current_user.fname + ' ' + current_user.lname
        date_input = request.form['date']
        data = {'Total Number of Change Pages': pages_input,
                'Writer': writer_input, 'Date': date_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'], request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30'],
                       request.form['input31'], request.form['input32'],
                       request.form['input33'], request.form['input34'],
                       request.form['input35'],
                       request.form['input36'], request.form['input37'],
                       request.form['input38'], request.form['input39'],
                       request.form['input40'],
                       request.form['input41'], request.form['input42'],
                       request.form['input43'], request.form['input44'],
                       request.form['input45'],
                       request.form['input46'], request.form['input47'],
                       request.form['input48'], request.form['input49'],
                       request.form['input50'],
                       request.form['input51'], request.form['input52'],
                       request.form['input53'], request.form['input54'],
                       request.form['input55']]
        for subject in range(len(inputs_list)):
            data.update({subject: {comment_name: inputs_list[subject]}})
            if inputs_list[subject] == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)
        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('writer_update.html', comment_name=comment_name,
                               revision=update_rev, job_no=job_no,
                               form=form,
                               writer_A=writer_A, writer_B=writer_B,
                               writer_C=writer_C, writer_D=writer_D,
                               writer_E=writer_E,
                               writer_F=writer_F, writer_G=writer_G)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        cl.completed = completed
        cl.er = er
        cl.data = JSON_data
        cl.op_manager = op_manager

        history = cl.checklist_history
        parsed = json.loads(history)
        num_list = []
        for x in list(parsed.keys()):
            num_list += [int(x)]
        num = max(num_list)+1
        parsed.update({num: ['Edit', current_user.username, str(datetime.utcnow()), completed]})
        cl.checklist_history = json.dumps(parsed)

        DB.session.commit()
        flash('Writer Checklist Updated!')
        return redirect(url_for('writer_index', job_no=job_no))
    else:
        return render_template('writer_update.html', comment_name=comment_name,
                               revision=update_rev, job_no=job_no,
                               form=form,
                               writer_A=writer_A, writer_B=writer_B,
                               writer_C=writer_C, writer_D=writer_D,
                               writer_E=writer_E,
                               writer_F=writer_F, writer_G=writer_G)

@APP.route('/writer_index/writer_delete/<job_no>/<delete_rev>/<checklist_type>',
           methods=['GET', 'POST'])
@login_required
def writer_delete(job_no, delete_rev, checklist_type):
    if not current_user.admin:
        flash('Only Admins can delete checklists')
        return redirect(url_for('writer_index', job_no=job_no))
    else:
        delete_fn(job_no, delete_rev, checklist_type)
        return redirect(url_for('writer_index', job_no=job_no))

# Illustration Checklists


@APP.route('/illustration_index/<job_no>', methods=['GET', 'POST'])
@login_required
def illustration_index(job_no):
    checklist_type = 'Illustration'
    checklist_list = Checklists.query.\
        filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    for x in checklist_list:
        rev.append(x.revision)
    max_revision = max(rev, default=0)
    row_content_list = []
    for r in rev:
        cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=r).first()
        checklist_data = json.loads(cl.data)
        stored_data = list(checklist_data.values())
        new_row = [cl.completed, cl.created_by]
        day = cl.date_created.replace(tzinfo=tz.gettz('UTC'))
        new_row += [day.astimezone(tz.tzlocal()).strftime("%Y-%m-%d %H:%M:%S")]
        new_row += [cl.er] # Engineering Review Level/ Customer Review
        new_row += [stored_data[0]]  # illustrator
        new_row += [stored_data[1]]  # date
        new_row += [stored_data[2]]  # comments
        stored_2 = checklist_data.values()
        data_2 = []
        for x in list(stored_2)[3:]:
            data_2 += x.values()
        new_row += data_2
        row_content_list += [new_row]
    table = TableCreation(checklist_type, max_revision,
                          row_content_list, job_no)
    return render_template('illustration_index.html',
                           table=table, job_no=job_no)

@APP.route('/illustration_index/illustration_update/<job_no>/<update_rev>',
           methods=['GET', 'POST'])
@login_required
def illustration_update(job_no, update_rev):
    checklist_type = 'Illustration'
    form = IllustrationUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter_by(revision=update_rev).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    form.illustrator.data = stored_data[0]
    form.date.data = datetime.strptime(stored_data[1], '%Y-%m-%d')
    form.comments.data = stored_data[2]
    eng_review = cl.er
    form.op_manager.data = cl.op_manager
    form.er_level.data = eng_review
    input_2 = [form.input1, form.input2, form.input3, form.input4]
    for x in range(len(input_2)):
        input_2[x].data = str(list(stored_data[3+x].values())[0])
    if eng_review == "":
        comment_name = 'Comments'
    else:
        if int(eng_review) == 0:
            comment_name = 'Pre-Delivery Comments'
        else:
            comment_name = 'Engineering Review Comments #'+str(eng_review)

    if request.method == "POST":
        completion_test = True
        er = request.form['er_level']
        er = test_number(er)
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        if er == "":
            completion_test = False
        illustrator_input = request.form['illustrator']
        if illustrator_input == "":
            illustrator_input = current_user.fname + ' ' + current_user.lname
        date_input = request.form['date']
        data = {'Illustrator': illustrator_input, 'Date': date_input,
                'comments': request.form['comments']}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4']]
        for subject in range(len(inputs_list)):
            data.update({subject: {comment_name: inputs_list[subject]}})
            if inputs_list[subject] == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('illustration_update.html',
                               comment_name=comment_name,
                               revision=update_rev,
                               job_no=job_no, form=form,
                               illustration_A=illustration_A)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        cl.completed = completed
        cl.data = JSON_data
        cl.er = er
        cl.op_manager = op_manager
        history = cl.checklist_history
        parsed = json.loads(history)
        num_list = []
        for x in list(parsed.keys()):
            num_list += [int(x)]
        num = max(num_list)+1
        parsed.update({num: ['Edit', current_user.username, str(datetime.utcnow()), completed]})
        cl.checklist_history = json.dumps(parsed)
        DB.session.commit()
        flash('Illustration Checklist Updated!')
        return redirect(url_for('illustration_index', job_no=job_no))
    else:
        return render_template('illustration_update.html',
                               comment_name=comment_name,
                               revision=update_rev,
                               job_no=job_no, form=form,
                               illustration_A=illustration_A)


@APP.route('/illustration_index/illustration_add/<job_no>',
           methods=['GET', 'POST'])
@login_required
def illustration_add(job_no):
    checklist_type = 'Illustration'
    form = IllustrationUpdate()
    form.illustrator.data = current_user.fname + ' ' + current_user.lname
    form.date.data = datetime.now().date()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    eng_review = []
    for x in cl:
        rev.append(x.revision)
        eng_review.append(x.er)
    revision = max(rev, default=0)+1
    eng_review_inter = [x for x in eng_review if x != '']
    eng_review = max(eng_review_inter, default=0)

    if len(cl) == 0 or len(eng_review_inter) == 0:
        form.er_level.data = 0
        comment_name = 'Pre-Delivery Comments'
    else:
        form.er_level.data = int(eng_review) + 1
        comment_name = 'Engineering Review Comments #'+str(int(eng_review)+1)

    if request.method == "POST":
        completion_test = True
        er = request.form['er_level']
        er = test_number(er)
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        if er == "":
            completion_test = False
        illustrator_input = request.form['illustrator']
        if illustrator_input == "":
            illustrator_input = current_user.fname + ' ' + current_user.lname
        date_input = request.form['date']
        data = {'Illustrator': illustrator_input,
                'Date': date_input, 'comments': request.form['comments']}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4']]
        for subject in range(len(inputs_list)):
            data.update({subject: {comment_name: inputs_list[subject]}})
            if inputs_list[subject] == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('illustration_update.html',
                               revision=revision,
                               comment_name=comment_name,
                               job_no=job_no, form=form,
                               illustration_A=illustration_A)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        history = {1: ['Create', current_user.username, str(datetime.utcnow()), completed]}
        json_history = json.dumps(history)
        new_checklist = Checklists(checklist_type=checklist_type,
                                   completed=completed, op_manager=op_manager,
                                   created_by=current_user.username,
                                   date_created=datetime.utcnow(),
                                   revision=revision, checklist_history=json_history,
                                   er=er,
                                   data=JSON_data)
        job = Jobs.query.filter_by(job_no=job_no).first()
        job.checklists.append(new_checklist)
        DB.session.add(new_checklist)
        DB.session.commit()
        flash('New Illustration Checklist Created!')
        return redirect(url_for('illustration_index', job_no=job_no))
    else:
        return render_template('illustration_update.html',
                               revision=revision,
                               comment_name=comment_name,
                               job_no=job_no, form=form,
                               illustration_A=illustration_A)


@APP.route('/illustration_index/illustration_delete/<job_no>/<delete_rev>/<checklist_type>',
           methods=['GET', 'POST'])
@login_required
def illustration_delete(job_no, delete_rev, checklist_type):
    if not current_user.admin:
        flash('Only Admins can delete checklists')
        return redirect(url_for('illustration_index', job_no=job_no))
    else:
        delete_fn(job_no, delete_rev, checklist_type)
        return redirect(url_for('illustration_index', job_no=job_no))


# Editor Checklists
@APP.route('/editor_index/<job_no>', methods=['GET', 'POST'])
@login_required
def editor_index(job_no):
    checklist_type = 'Editor'
    checklist_list = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    for x in checklist_list:
        rev.append(x.revision)
    max_revision = max(rev, default=0)
    row_content_list = []
    for r in rev:
        cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=r).first()
        checklist_data = json.loads(cl.data)
        stored_data = list(checklist_data.values())
        new_row = [cl.completed, cl.created_by]
        day = cl.date_created.replace(tzinfo=tz.gettz('UTC'))
        new_row += [day.astimezone(tz.tzlocal()).strftime("%Y-%m-%d %H:%M:%S")]
        new_row += [cl.er]
        new_row += [stored_data[0]]  # editor
        new_row += [stored_data[1]]  # date
        stored_2 = checklist_data.values()
        data_2 = []
        for x in list(stored_2)[2:]:
            data_2 += x.values()
        new_row += data_2
        row_content_list += [new_row]
    table = TableCreation(checklist_type, max_revision,
                          row_content_list, job_no)
    return render_template('editor_index.html', table=table, job_no=job_no)


@APP.route('/editor_index/editor_add/<job_no>', methods=['GET', 'POST'])
@login_required
def editor_add(job_no):
    checklist_type = 'Editor'
    form = EditorUpdate()
    form.editor.data = current_user.fname + ' ' + current_user.lname
    form.date.data = datetime.now().date()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()

    rev = []
    eng_review = []
    for x in cl:
        rev.append(x.revision)
        eng_review.append(x.er)
    revision = max(rev, default=0)+1
    eng_review_inter = [x for x in eng_review if x != '']
    eng_review = max(eng_review_inter, default=0)

    if len(cl) == 0 or len(eng_review_inter) == 0:
        form.er_level.data = 0
        comment_name = 'Pre-Delivery Comments'
    else:
        form.er_level.data = int(eng_review) + 1
        comment_name = 'Engineering Review Comments #'+str(int(eng_review)+1)

    if request.method == 'POST':
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        er = request.form['er_level']
        er = test_number(er)
        if er == "":
            completion_test = False
        editor_input = request.form['editor']
        if editor_input == "":
            editor_input = current_user.fname + ' ' + current_user.lname
        date_input = request.form['date']
        data = {'Editor': editor_input, 'Date': date_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30'],
                       request.form['input31'], request.form['input32'],
                       request.form['input33'], request.form['input34'],
                       request.form['input35'],
                       request.form['input36'], request.form['input37'],
                       request.form['input38'], request.form['input39'],
                       request.form['input40'],
                       request.form['input41'], request.form['input42'],
                       request.form['input43'], request.form['input44'],
                       request.form['input45'],
                       request.form['input46'], request.form['input47'],
                       request.form['input48'], request.form['input49'],
                       request.form['input50'],
                       request.form['input51']]
        for subject in range(len(inputs_list)):
            input_valid = test_number(inputs_list[subject])
            data.update({subject: {comment_name: input_valid}})
            if input_valid == "":
                completion_test = False

        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('editor_update.html', comment_name=comment_name,
                               revision=revision, job_no=job_no, form=form,
                               editor_A=editor_A, editor_B=editor_B,
                               editor_C=editor_C, editor_D=editor_D,
                               editor_E=editor_E,
                               editor_F=editor_F, editor_G=editor_G,
                               editor_H=editor_H, editor_I=editor_I,
                               editor_J=editor_J)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False

        history = {1: ['Create', current_user.username, str(datetime.utcnow()), completed]}
        json_history = json.dumps(history)
        new_checklist = Checklists(checklist_type=checklist_type,
                                   created_by=current_user.username,
                                   completed=completed, op_manager=op_manager,
                                   date_created=datetime.utcnow(),
                                   revision=revision,
                                   er=er, checklist_history=json_history,
                                   data=JSON_data)
        job = Jobs.query.filter_by(job_no=job_no).first()
        job.checklists.append(new_checklist)
        DB.session.add(new_checklist)
        DB.session.commit()
        flash('New Editor Checklist Created!')
        return redirect(url_for('editor_index', job_no=job_no))
    else:
        return render_template('editor_update.html', comment_name=comment_name,
                               revision=revision, job_no=job_no, form=form,
                               editor_A=editor_A, editor_B=editor_B,
                               editor_C=editor_C, editor_D=editor_D,
                               editor_E=editor_E,
                               editor_F=editor_F, editor_G=editor_G,
                               editor_H=editor_H, editor_I=editor_I,
                               editor_J=editor_J)


@APP.route('/editor_index/editor_update/<job_no>/<update_rev>',
           methods=['GET', 'POST'])
@login_required
def editor_update(job_no, update_rev):
    checklist_type = 'Editor'
    form = EditorUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter_by(revision=update_rev).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    form.op_manager.data = cl.op_manager
    if cl.created_by == 'JIRA WEBHOOK':
        cl.created_by = current_user.username
        cl.date_created = datetime.utcnow()
        form.editor.data = current_user.fname + ' ' + current_user.lname
        form.date.data = datetime.now().date()
    else:
        form.editor.data = stored_data[0]
        form.date.data = datetime.strptime(stored_data[1], '%Y-%m-%d')
    eng_review = cl.er
    form.er_level.data = eng_review
    input_2 = [form.input1, form.input2, form.input3,
               form.input4, form.input5,
               form.input6, form.input7, form.input8,
               form.input9, form.input10,
               form.input11, form.input12, form.input13,
               form.input14, form.input15,
               form.input16, form.input17, form.input18,
               form.input19, form.input20,
               form.input21, form.input22, form.input23,
               form.input24, form.input25,
               form.input26, form.input27, form.input28,
               form.input29, form.input30,
               form.input31, form.input32, form.input33,
               form.input34, form.input35,
               form.input36, form.input37, form.input38,
               form.input39, form.input40,
               form.input41, form.input42, form.input43,
               form.input44, form.input45,
               form.input46, form.input47, form.input48,
               form.input49, form.input50,
               form.input51]
    for x in range(len(input_2)):
        if list(stored_data[2+x].values())[0] == "":
            input_2[x].data = ""
        else:
            input_2[x].data = int(list(stored_data[2+x].values())[0])
    if eng_review == "":
        comment_name = 'Comments'
    else:
        if int(eng_review) == 0:
            comment_name = 'Pre-Delivery Comments'
        else:
            comment_name = 'Engineering Review Comments #'+str(eng_review)

    if request.method == "POST":
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        er = request.form['er_level']
        er = test_number(er)
        if er == "":
            completion_test = False
        editor_input = request.form['editor']
        if editor_input == "":
            editor_input = current_user.fname + ' ' + current_user.lname
        date_input = request.form['date']
        data = {'Editor': editor_input, 'Date': date_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30'],
                       request.form['input31'], request.form['input32'],
                       request.form['input33'], request.form['input34'],
                       request.form['input35'],
                       request.form['input36'], request.form['input37'],
                       request.form['input38'], request.form['input39'],
                       request.form['input40'],
                       request.form['input41'], request.form['input42'],
                       request.form['input43'], request.form['input44'],
                       request.form['input45'],
                       request.form['input46'], request.form['input47'],
                       request.form['input48'], request.form['input49'],
                       request.form['input50'],
                       request.form['input51']]
        for subject in range(len(inputs_list)):
            input_valid = test_number(inputs_list[subject])
            data.update({subject: {comment_name: input_valid}})
            if input_valid == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('editor_update.html', comment_name=comment_name,
                               revision=update_rev, job_no=job_no, form=form,
                               editor_A=editor_A, editor_B=editor_B,
                               editor_C=editor_C, editor_D=editor_D,
                               editor_E=editor_E,
                               editor_F=editor_F, editor_G=editor_G,
                               editor_H=editor_H, editor_I=editor_I,
                               editor_J=editor_J)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        cl.completed=completed
        cl.er = er
        cl.op_manager = op_manager
        cl.data = JSON_data

        history = cl.checklist_history
        parsed = json.loads(history)
        num_list = []
        for x in list(parsed.keys()):
            num_list += [int(x)]
        num = max(num_list)+1
        parsed.update({num: ['Edit', current_user.username, str(datetime.utcnow()), completed]})
        cl.checklist_history = json.dumps(parsed)

        DB.session.commit()
        flash('Editor Checklist Updated!')
        return redirect(url_for('editor_index', job_no=job_no))
    else:
        return render_template('editor_update.html', comment_name=comment_name,
                               revision=update_rev, job_no=job_no, form=form,
                               editor_A=editor_A, editor_B=editor_B,
                               editor_C=editor_C, editor_D=editor_D,
                               editor_E=editor_E,
                               editor_F=editor_F, editor_G=editor_G,
                               editor_H=editor_H, editor_I=editor_I,
                               editor_J=editor_J)


@APP.route('/editor_index/editor_delete/<job_no>/<delete_rev>/<checklist_type>',
           methods=['GET', 'POST'])
@login_required
def editor_delete(job_no, delete_rev, checklist_type):
    if not current_user.admin:
        flash('Only Admins can delete checklists')
        return redirect(url_for('editor_index', job_no=job_no))
    else:
        delete_fn(job_no, delete_rev, checklist_type)
        return redirect(url_for('editor_index', job_no=job_no))

# QA Checklists


@APP.route('/qa_index/<job_no>', methods=['GET', 'POST'])
@login_required
def qa_index(job_no):
    checklist_type = 'QA'
    checklist_list = Checklists.query.\
        filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    for x in checklist_list:
        rev.append(x.revision)
    max_revision = max(rev, default=0)
    row_content_list = []
    for r in rev:
        cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=r).first()
        checklist_data = json.loads(cl.data)
        stored_data = list(checklist_data.values())
        new_row = [cl.completed, cl.created_by]
        day = cl.date_created.replace(tzinfo=tz.gettz('UTC'))
        new_row += [day.astimezone(tz.tzlocal()).strftime("%Y-%m-%d %H:%M:%S")]
        new_row += [cl.er]
        new_row += [stored_data[0]]  # qa 1
        new_row += [stored_data[1]]  # date 1
        new_row += [stored_data[2]]  # qa 2
        new_row += [stored_data[3]]  # date 2
        new_row += [stored_data[4]]  # qa 3
        new_row += [stored_data[5]]  # date 3
        stored_2 = checklist_data.values()
        data_2 = []
        for x in list(stored_2)[6:]:
            data_2 += x.values()
        new_row += data_2
        row_content_list += [new_row]
    table = TableCreation(checklist_type, max_revision,
                          row_content_list, job_no)
    return render_template('qa_index.html', table=table, job_no=job_no)


@APP.route('/qa_add/<job_no>', methods=['GET', 'POST'])
@login_required
def qa_add(job_no):
    checklist_type = 'QA'
    form = QAUpdate()
    form.qa_1.data = current_user.fname + ' ' + current_user.lname
    form.date_1.data = datetime.now().date()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()

    rev = []
    eng_review = []
    for x in cl:
        rev.append(x.revision)
        eng_review.append(x.er)

    revision = max(rev, default=0)+1
    eng_review_inter = [x for x in eng_review if x != '']
    eng_review = max(eng_review_inter, default=0)

    if len(cl) == 0 or len(eng_review_inter) == 0:
        form.er_level.data = 0
        comment_name = 'Pre-Delivery Comments'
    else:
        form.er_level.data = eng_review + 1
        comment_name = 'Engineering Review Comments #'+str(int(eng_review)+1)
        
    if request.method == 'POST':
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        er = request.form['er_level']
        er = test_number(er)
        if er == "":
            completion_test = False
        qa_1_input = request.form['qa_1']
        if qa_1_input == "":
            qa_1_input = current_user.fname + ' ' + current_user.lname
        date_1_input = request.form['date_1']
        qa_2_input = request.form['qa_2']
        date_2_input = request.form['date_2']
        qa_3_input = request.form['qa_3']
        date_3_input = request.form['date_3']
        completion_check = [qa_2_input, date_2_input, qa_3_input, date_3_input]
        for x in completion_check:
            if x == "":
                completion_test = False
        data = {'QA 1': qa_1_input, 'Date 1': date_1_input,
                'QA 2': qa_2_input, 'Date 2': date_2_input,
                'QA 3': qa_3_input, 'Date 3': date_3_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30'],
                       request.form['input31'], request.form['input32']]
        for subject in range(len(inputs_list)):
            input_valid = test_number(inputs_list[subject])
            data.update({subject: {comment_name: input_valid}})
            # if subject in range(21):
            if input_valid == "":
                completion_test = False

        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date 1'] == "":
            date_data['Date 1'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('qa_update.html', comment_name=comment_name,
                               revision=revision,
                               job_no=job_no, form=form,
                               qa_A=qa_A, qa_B=qa_B, qa_C=qa_C,
                               qa_D=qa_D, qa_E=qa_E,
                               qa_G=qa_G, qa_H=qa_H, qa_I=qa_I,
                               qa_note=qa_note)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        history = {1: ['Create', current_user.username, str(datetime.utcnow()), completed]}
        json_history = json.dumps(history)
        new_checklist = Checklists(checklist_type=checklist_type,
                                   created_by=current_user.username,
                                   completed=completed,
                                   date_created=datetime.utcnow(),
                                   revision=revision, op_manager=op_manager,
                                   er=er, checklist_history=json_history,
                                   data=JSON_data)
        job = Jobs.query.filter_by(job_no=job_no).first()
        job.checklists.append(new_checklist)
        DB.session.add(new_checklist)
        DB.session.commit()
        flash('New QA Checklist Created!')
        return redirect(url_for('qa_index', job_no=job_no))
    else:
        return render_template('qa_update.html', comment_name=comment_name,
                               revision=revision,
                               job_no=job_no, form=form,
                               qa_A=qa_A, qa_B=qa_B, qa_C=qa_C,
                               qa_D=qa_D, qa_E=qa_E,
                               qa_G=qa_G, qa_H=qa_H, qa_I=qa_I,
                               qa_note=qa_note)


@APP.route('/qa_index/qa_update/<job_no>/<update_rev>', methods=['GET', 'POST'])
@login_required
def qa_update(job_no, update_rev):
    checklist_type = 'QA'
    form = QAUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter_by(revision=update_rev).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    if cl.created_by == 'JIRA WEBHOOK':
        cl.created_by = current_user.username
        cl.date_created = datetime.utcnow()
        form.qa_1.data = current_user.fname + ' ' + current_user.lname
        form.date_1.data = datetime.now().date()
        form.qa_2.data = ""
        form.date_2.data = ""
        form.qa_3.data = ""
        form.date_3.data = ""
    else:
        form.qa_1.data = stored_data[0]
        form.date_1.data = datetime.strptime(stored_data[1], '%Y-%m-%d')
        form.qa_2.data = stored_data[2]
        if stored_data[3] == "":
            form.date_2.data = ""
        else:
            form.date_2.data = datetime.strptime(stored_data[3], '%Y-%m-%d')
        form.qa_3.data = stored_data[4]
        if stored_data[5] == "":
            form.date_3.data = ""
        else:
            form.date_3.data = datetime.strptime(stored_data[5], '%Y-%m-%d')
    input_2 = [form.input1, form.input2, form.input3,
               form.input4, form.input5,
               form.input6, form.input7, form.input8,
               form.input9, form.input10,
               form.input11, form.input12, form.input13,
               form.input14, form.input15,
               form.input16, form.input17, form.input18,
               form.input19, form.input20,
               form.input21, form.input22, form.input23,
               form.input24, form.input25,
               form.input26, form.input27, form.input28,
               form.input29, form.input30,
               form.input31, form.input32]
    for x in range(len(input_2)):
        if list(stored_data[6+x].values())[0] == "":
            input_2[x].data = ""
        else:
            input_2[x].data = int(list(stored_data[6+x].values())[0])
    eng_review = cl.er
    form.er_level.data = eng_review
    form.op_manager.data = cl.op_manager
    if eng_review == "":
        comment_name = 'Comments'
    else:
        if int(eng_review) == 0:
            comment_name = 'Pre-Delivery Comments'
        else:
            comment_name = 'Engineering Review Comments #'+str(eng_review)
    if request.method == "POST":
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        er = request.form['er_level']
        er = test_number(er)
        if er == "":
            completion_test = False
        qa_1_input = request.form['qa_1']
        if qa_1_input == "":
            qa_1_input = current_user.fname + ' ' + current_user.lname
        date_1_input = request.form['date_1']
        qa_2_input = request.form['qa_2']
        date_2_input = request.form['date_2']
        qa_3_input = request.form['qa_3']
        date_3_input = request.form['date_3']
        completion_check = [qa_2_input, date_2_input, qa_3_input, date_3_input]
        for x in completion_check:
            if x == "":
                completion_test = False
        data = {'QA 1': qa_1_input, 'Date 1': date_1_input,
                'QA 2': qa_2_input, 'Date 2': date_2_input,
                'QA 3': qa_3_input, 'Date 3': date_3_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30'],
                       request.form['input31'], request.form['input32']]
        for subject in range(len(inputs_list)):
            input_valid = test_number(inputs_list[subject])
            data.update({subject: {comment_name: input_valid}})
            if input_valid == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date 1'] == "":
            date_data['Date 1'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('qa_update.html', comment_name=comment_name,
                               revision=update_rev,
                               job_no=job_no, form=form,
                               qa_A=qa_A, qa_B=qa_B, qa_C=qa_C,
                               qa_D=qa_D, qa_E=qa_E,
                               qa_G=qa_G, qa_H=qa_H, qa_I=qa_I,
                               qa_note=qa_note)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        cl.completed=completed
        cl.data = JSON_data
        cl.er = er
        cl.op_manager = op_manager

        history = cl.checklist_history
        parsed = json.loads(history)
        num_list = []
        for x in list(parsed.keys()):
            num_list += [int(x)]
        num = max(num_list)+1
        parsed.update({num: ['Edit', current_user.username, str(datetime.utcnow()), completed]})
        cl.checklist_history = json.dumps(parsed)
        DB.session.commit()
        flash('QA Checklist Updated!')
        return redirect(url_for('qa_index', job_no=job_no))
    else:
        return render_template('qa_update.html', comment_name=comment_name,
                               revision=update_rev,
                               job_no=job_no, form=form,
                               qa_A=qa_A, qa_B=qa_B, qa_C=qa_C,
                               qa_D=qa_D, qa_E=qa_E,
                               qa_G=qa_G, qa_H=qa_H, qa_I=qa_I,
                               qa_note=qa_note)


@APP.route('/qa_index/qa_delete/<job_no>/<delete_rev>/<checklist_type>',
           methods=['GET', 'POST'])
@login_required
def qa_delete(job_no, delete_rev, checklist_type):
    if not current_user.admin:
        flash('Only Admins can delete checklists')
        return redirect(url_for('qa_index', job_no=job_no))
    else:
        delete_fn(job_no, delete_rev, checklist_type)
        return redirect(url_for('qa_index', job_no=job_no))
# Final Delivery Checklists


@APP.route('/final_delivery_index/<job_no>', methods=['GET', 'POST'])
@login_required
def final_delivery_index(job_no):
    checklist_type = 'Final Delivery'
    checklist_list = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    for x in checklist_list:
        rev.append(x.revision)
    max_revision = max(rev, default=0)
    row_content_list = []
    for r in rev:
        cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=r).first()
        checklist_data = json.loads(cl.data)
        stored_data = list(checklist_data.values())
        new_row = [cl.completed, cl.created_by]
        day = cl.date_created.replace(tzinfo=tz.gettz('UTC'))
        new_row += [day.astimezone(tz.tzlocal()).strftime("%Y-%m-%d %H:%M:%S")]
        new_row += [cl.er]
        new_row += [stored_data[0]]  # name
        new_row += [stored_data[1]]  # date
        stored_2 = checklist_data.values()
        data_2 = []
        for x in list(stored_2)[2:]:
            data_2 += x.values()
        new_row += data_2
        row_content_list += [new_row]
    table = TableCreation(checklist_type, max_revision,
                          row_content_list, job_no)
    return render_template('final_delivery_index.html',
                           table=table, job_no=job_no)


@APP.route('/final_delivery_add/<job_no>', methods=['GET', 'POST'])
@login_required
def final_delivery_add(job_no):
    checklist_type = 'Final Delivery'
    form = FinalDeliveryUpdate()
    form.name_fd.data = current_user.fname + ' ' + current_user.lname
    form.date.data = datetime.now().date()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    eng_review = []
    for x in cl:
        rev.append(x.revision)
        eng_review.append(x.er)
    revision = max(rev, default=0)+1
    eng_review_inter = [x for x in eng_review if x != '']
    eng_review = max(eng_review_inter, default=0)

    if len(cl) == 0 or len(eng_review_inter) == 0:
        form.er_level.data = 0
    else:
        form.er_level.data = int(eng_review) + 1
    comment_name = 'Completed'
    if request.method == 'POST':
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        er = request.form['er_level']
        er = test_number(er)
        if er == "":
            completion_test = False
        name_input = request.form['name_fd']
        if name_input == "":
            name_input = current_user.fname + ' ' + current_user.lname
        date_input = request.form['date']
        data = {'Name': name_input, 'Date': date_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30']]
        for subject in range(len(inputs_list)):
            data.update({subject: {comment_name: inputs_list[subject]}})
            if inputs_list[subject] == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('final_delivery_update.html',
                               comment_name=comment_name,
                               revision=revision,
                               job_no=job_no, form=form,
                               dl_A=dl_A, dl_B=dl_B, dl_C=dl_C,
                               dl_D=dl_D, dl_E=dl_E,
                               dl_F=dl_F, dl_G=dl_G, dl_H=dl_H)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        history = {1: ['Create', current_user.username, str(datetime.utcnow()), completed]}
        json_history = json.dumps(history)
        new_checklist = Checklists(checklist_type=checklist_type,
                                   created_by=current_user.username,
                                   completed=completed, op_manager=op_manager,
                                   date_created=datetime.utcnow(),
                                   revision=revision,
                                   er=er, checklist_history=json_history,
                                   data=JSON_data)
        job = Jobs.query.filter_by(job_no=job_no).first()
        job.checklists.append(new_checklist)
        DB.session.add(new_checklist)
        DB.session.commit()
        flash('New Final Delivery Checklist Created!')
        return redirect(url_for('final_delivery_index', job_no=job_no))
    else:
        return render_template('final_delivery_update.html',
                               comment_name=comment_name,
                               revision=revision,
                               job_no=job_no, form=form,
                               dl_A=dl_A, dl_B=dl_B, dl_C=dl_C,
                               dl_D=dl_D, dl_E=dl_E,
                               dl_F=dl_F, dl_G=dl_G, dl_H=dl_H)


@APP.route('/final_delivery_index/final_delivery_update/<job_no>/<update_rev>', methods=['GET', 'POST'])
@login_required
def final_delivery_update(job_no, update_rev):
    checklist_type = 'Final Delivery'
    form = FinalDeliveryUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter_by(revision=update_rev).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    if cl.created_by == 'JIRA WEBHOOK':
        cl.created_by = current_user.username
        cl.date_created = datetime.utcnow()
        form.name_fd.data = current_user.fname + ' ' + current_user.lname
        form.date.data = datetime.now().date()
    else:
        form.name_fd.data = stored_data[0]
        form.date.data = datetime.strptime(stored_data[1], '%Y-%m-%d')

    form.er_level.data = cl.er
    form.op_manager.data = cl.op_manager

    input_2 = [form.input1, form.input2, form.input3,
               form.input4, form.input5,
               form.input6, form.input7, form.input8,
               form.input9, form.input10,
               form.input11, form.input12, form.input13,
               form.input14, form.input15,
               form.input16, form.input17, form.input18,
               form.input19, form.input20,
               form.input21, form.input22, form.input23,
               form.input24, form.input25,
               form.input26, form.input27, form.input28,
               form.input29, form.input30]
    for x in range(len(input_2)):
        input_2[x].data = str(list(stored_data[2+x].values())[0])
    comment_name = 'Completed'
    if request.method == "POST":
        completion_test = True
        er = request.form['er_level']
        er = test_number(er)
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        if er == "":
            completion_test == False
        name_fd_input = request.form['name_fd']
        if name_fd_input == "":
            name_fd_input = current_user.fname + ' ' + current_user.lname
        date_input = request.form['date']
        data = {'Name': name_fd_input, 'Date': date_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30']]
        for subject in range(len(inputs_list)):
            data.update({subject: {comment_name: inputs_list[subject]}})
            if inputs_list[subject] == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('final_delivery_update.html',
                               comment_name=comment_name,
                               revision=update_rev,
                               job_no=job_no, form=form,
                               dl_A=dl_A, dl_B=dl_B, dl_C=dl_C,
                               dl_D=dl_D, dl_E=dl_E,
                               dl_F=dl_F, dl_G=dl_G, dl_H=dl_H)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        cl.completed=completed
        cl.data = JSON_data
        cl.er = er
        cl.op_manager = op_manager
        history = cl.checklist_history
        parsed = json.loads(history)
        num_list = []
        for x in list(parsed.keys()):
            num_list += [int(x)]
        num = max(num_list)+1
        parsed.update({num: ['Edit', current_user.username, str(datetime.utcnow()), completed]})
        cl.checklist_history = json.dumps(parsed)
        DB.session.commit()
        flash('Final Delivery Checklist Updated!')
        return redirect(url_for('final_delivery_index', job_no=job_no))
    else:
        return render_template('final_delivery_update.html',
                               comment_name=comment_name,
                               revision=update_rev,
                               job_no=job_no, form=form,
                               dl_A=dl_A, dl_B=dl_B, dl_C=dl_C,
                               dl_D=dl_D, dl_E=dl_E,
                               dl_F=dl_F, dl_G=dl_G, dl_H=dl_H)

@APP.route('/final_delivery_index/final_delivery_delete/<job_no>/<delete_rev>/<checklist_type>',
           methods=['GET', 'POST'])
@login_required
def final_delivery_delete(job_no, delete_rev, checklist_type):
    if not current_user.admin:
        flash('Only Admins can delete checklists')
        return redirect(url_for('final_delivery_index', job_no=job_no))
    else:
        delete_fn(job_no, delete_rev, checklist_type)
        return redirect(url_for('final_delivery_index', job_no=job_no))
# DPMO Checklists


@APP.route('/dpmo_index/<job_no>', methods=['GET', 'POST'])
@login_required
def dpmo_index(job_no):
    checklist_type = 'DPMO'
    checklist_list = Checklists.query.\
        filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    for x in checklist_list:
        rev.append(x.revision)
    max_revision = max(rev, default=0)
    row_content_list = []
    for r in rev:
        cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=r).first()
        checklist_data = json.loads(cl.data)
        stored_data = list(checklist_data.values())
        new_row = [cl.completed, cl.created_by]
        day = cl.date_created.replace(tzinfo=tz.gettz('UTC'))
        new_row += [day.astimezone(tz.tzlocal()).strftime("%Y-%m-%d %H:%M:%S")]
        new_row += [cl.er]
        new_row += [stored_data[0]]  # date
        new_row += [stored_data[1]]  # total_pages
        new_row += [stored_data[2]]  # DPMO score
        stored_2 = checklist_data.values()
        data_2 = []
        for x in list(stored_2)[3:]:
            data_2 += x.values()
        new_row += data_2
        row_content_list += [new_row]
    table = TableCreation(checklist_type, max_revision,
                          row_content_list, job_no)
    return render_template('dpmo_index.html', table=table, job_no=job_no)

@APP.route('/dpmo_index/dpmo_update/<job_no>/<update_rev>',
           methods=['GET', 'POST'])
@login_required
def dpmo_update(job_no, update_rev):
    checklist_type = 'DPMO'
    form = DPMOUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter_by(revision=update_rev).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    form.er_level.data = cl.er
    form.op_manager.data = cl.op_manager
    if cl.created_by == 'JIRA WEBHOOK':
        cl.created_by = current_user.username
        cl.date_created = datetime.utcnow()
        form.date.data = datetime.now().date()
    else:
        form.date.data = datetime.strptime(stored_data[0], '%Y-%m-%d')
    form.total_pages.data = stored_data[1]
    input_2 = [form.input1, form.input2, form.input3,
               form.input4, form.input5,
               form.input6, form.input7, form.input8,
               form.input9, form.input10,
               form.input11, form.input12, form.input13,
               form.input14, form.input15,
               form.input16, form.input17, form.input18,
               form.input19, form.input20,
               form.input21, form.input22, form.input23,
               form.input24, form.input25,
               form.input26, form.input27, form.input28,
               form.input29, form.input30,
               form.input31, form.input32, form.input33,
               form.input34, form.input35,
               form.input36, form.input37, form.input38,
               form.input39, form.input40,
               form.input41, form.input42]
    for x in range(len(input_2)):
        if list(stored_data[3+x].values())[0] == "":
            input_2[x].data = ""
        else:
            input_2[x].data = int(list(stored_data[3+x].values())[0])

    comment_name = 'Number of Errors'
    if request.method == "POST":
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        er = request.form['er_level']
        er = test_number(er)
        total_pages_input = request.form['total_pages']
        test_number(total_pages_input)
        if total_pages_input == "" or er == "":
            completion_test = False
        date_input = request.form['date']
        data = {'Date': date_input, 'Total Pages': total_pages_input, 'DPMO Score': 0}
        error_total = 0
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30'],
                       request.form['input31'], request.form['input32'],
                       request.form['input33'], request.form['input34'],
                       request.form['input35'],
                       request.form['input36'], request.form['input37'],
                       request.form['input38'], request.form['input39'],
                       request.form['input40'],
                       request.form['input41'], request.form['input42']]
        for subject in range(len(inputs_list)):
            error = test_number(inputs_list[subject])
            if inputs_list[subject] == "":
                completion_test = False
            data.update({subject: {'Number of errors': error}})
            if subject not in dpmo_exclusion:
                if error == "":
                    error_total += 0
                else:
                    error_total += int(error)
        if total_pages_input == "":
            score = 0
        else:
            if int(total_pages_input) == 0:
                score = 0
            else:
                score = int((error_total/int(total_pages_input))*1000000)
        data.update({'DPMO Score': score})
        JSON_data = json.dumps(data)
        # Replaces an empty date field with the default date
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('dpmo_update.html', comment_name=comment_name,
                               revision=update_rev, job_no=job_no, form=form,
                               dpmo_A=dpmo_A, dpmo_B=dpmo_B,
                               dpmo_C=dpmo_C,
                               dpmo_D=dpmo_D, dpmo_E=dpmo_E,
                               dpmo_F=dpmo_F, dpmo_G=dpmo_G,
                               dpmo_H=dpmo_H,
                               dpmo_I=dpmo_I, dpmo_J=dpmo_J,
                               dpmo_K=dpmo_K,
                               dpmo_L=dpmo_L, dpmo_M=dpmo_M,
                               dpmo_N=dpmo_N,
                               dpmo_O=dpmo_O, dpmo_P=dpmo_P)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        cl.completed=completed
        cl.data = JSON_data
        cl.er = er
        cl.op_manager = op_manager

        history = cl.checklist_history
        parsed = json.loads(history)
        num_list = []
        for x in list(parsed.keys()):
            num_list += [int(x)]
        num = max(num_list)+1
        parsed.update({num: ['Edit', current_user.username, str(datetime.utcnow()), completed]})
        cl.checklist_history = json.dumps(parsed)

        DB.session.commit()
        flash('DPMO Checklist Updated!')
        return redirect(url_for('dpmo_index', job_no=job_no))
    else:
        return render_template('dpmo_update.html', comment_name=comment_name,
                               revision=update_rev, job_no=job_no, form=form,
                               dpmo_A=dpmo_A, dpmo_B=dpmo_B,
                               dpmo_C=dpmo_C,
                               dpmo_D=dpmo_D, dpmo_E=dpmo_E,
                               dpmo_F=dpmo_F, dpmo_G=dpmo_G,
                               dpmo_H=dpmo_H,
                               dpmo_I=dpmo_I, dpmo_J=dpmo_J,
                               dpmo_K=dpmo_K,
                               dpmo_L=dpmo_L, dpmo_M=dpmo_M,
                               dpmo_N=dpmo_N,
                               dpmo_O=dpmo_O, dpmo_P=dpmo_P)


@APP.route('/dpmo_add/<job_no>', methods=['GET', 'POST'])
@login_required
def dpmo_add(job_no):
    checklist_type = 'DPMO'
    form = DPMOUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    form.date.data = datetime.now().date()
    rev = []
    eng_review = []
    for x in cl:
        rev.append(x.revision)
        eng_review.append(x.er)
    eng_review_inter = [x for x in eng_review if x != '']
    eng_review = max(eng_review_inter, default=0)
    revision = max(rev, default=0)+1
    comment_name = 'Number of Errors'
    if len(cl) == 0 or len(eng_review_inter) == 0:
        form.er_level.data = 0
    else:
        form.er_level.data = int(eng_review) + 1
    if request.method == 'POST':
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        pages_input = request.form['total_pages']
        date_input = request.form['date']
        pages_input = test_number(pages_input)
        er = request.form['er_level']
        er = test_number(er)
        if er == "" or pages_input == "":
            completion_test = False
        data = {'Date': date_input, 'Total Pages': pages_input,
                'DPMO Score': 0}
        error_total = 0
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'], request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12'],
                       request.form['input13'], request.form['input14'],
                       request.form['input15'],
                       request.form['input16'], request.form['input17'],
                       request.form['input18'], request.form['input19'],
                       request.form['input20'],
                       request.form['input21'], request.form['input22'],
                       request.form['input23'], request.form['input24'],
                       request.form['input25'],
                       request.form['input26'], request.form['input27'],
                       request.form['input28'], request.form['input29'],
                       request.form['input30'],
                       request.form['input31'], request.form['input32'],
                       request.form['input33'], request.form['input34'],
                       request.form['input35'],
                       request.form['input36'], request.form['input37'],
                       request.form['input38'], request.form['input39'],
                       request.form['input40'],
                       request.form['input41'], request.form['input42']]
        for subject in range(len(inputs_list)):
            error = test_number(inputs_list[subject])
            if error == "":
                completion_test = False
            data.update({subject: {'Number of errors': error}})
            if subject not in dpmo_exclusion:
                if error == "":
                    error_total += 0
                else:
                    error_total += int(error)
        if pages_input == "":
            score = 0
        else:
            if int(pages_input) == 0:
                score = 0
            else:
                score = int((error_total/int(pages_input))*1000000)
        data.update({'DPMO Score': score})
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date'] == "":
            date_data['Date'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('dpmo_update.html', comment_name=comment_name,
                               revision=revision, job_no=job_no, form=form,
                               dpmo_A=dpmo_A, dpmo_B=dpmo_B,
                               dpmo_C=dpmo_C,
                               dpmo_D=dpmo_D, dpmo_E=dpmo_E,
                               dpmo_F=dpmo_F, dpmo_G=dpmo_G,
                               dpmo_H=dpmo_H,
                               dpmo_I=dpmo_I, dpmo_J=dpmo_J,
                               dpmo_K=dpmo_K,
                               dpmo_L=dpmo_L, dpmo_M=dpmo_M,
                               dpmo_N=dpmo_N,
                               dpmo_O=dpmo_O, dpmo_P=dpmo_P)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        history = {1: ['Create', current_user.username, str(datetime.utcnow()), completed]}
        json_history = json.dumps(history)
        new_checklist = Checklists(checklist_type=checklist_type,
                                   created_by=current_user.username,
                                   completed=completed,
                                   date_created=datetime.utcnow(),
                                   revision=revision, op_manager=op_manager,
                                   er=er, checklist_history=json_history,
                                   data=JSON_data)
        job = Jobs.query.filter_by(job_no=job_no).first()
        job.checklists.append(new_checklist)
        DB.session.add(new_checklist)
        DB.session.commit()
        flash('New DPMO Checklist Created!')
        return redirect(url_for('dpmo_index', job_no=job_no))
    else:
        return render_template('dpmo_update.html', comment_name=comment_name,
                               revision=revision, job_no=job_no, form=form,
                               dpmo_A=dpmo_A, dpmo_B=dpmo_B,
                               dpmo_C=dpmo_C,
                               dpmo_D=dpmo_D, dpmo_E=dpmo_E,
                               dpmo_F=dpmo_F, dpmo_G=dpmo_G,
                               dpmo_H=dpmo_H,
                               dpmo_I=dpmo_I, dpmo_J=dpmo_J,
                               dpmo_K=dpmo_K,
                               dpmo_L=dpmo_L, dpmo_M=dpmo_M,
                               dpmo_N=dpmo_N,
                               dpmo_O=dpmo_O, dpmo_P=dpmo_P)


@APP.route('/dpmo_index/dpmo_delete/<job_no>/<delete_rev>/<checklist_type>',
           methods=['GET', 'POST'])
@login_required
def dpmo_delete(job_no, delete_rev, checklist_type):
    if not current_user.admin:
        flash('Only Admins can delete checklists')
        return redirect(url_for('dpmo_index', job_no=job_no))
    else:
        delete_fn(job_no, delete_rev, checklist_type)
        return redirect(url_for('dpmo_index', job_no=job_no))


# Translation Checklists
@APP.route('/translation_index/<job_no>', methods=['GET', 'POST'])
@login_required
def translation_index(job_no):
    checklist_type = 'Translation'
    checklist_list = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()
    rev = []
    for x in checklist_list:
        rev.append(x.revision)
    max_revision = max(rev, default=0)
    row_content_list = []
    for r in rev:
        cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=r).first()
        checklist_data = json.loads(cl.data)
        stored_data = list(checklist_data.values())
        new_row = [cl.completed, cl.created_by]
        day = cl.date_created.replace(tzinfo=tz.gettz('UTC'))
        new_row += [day.astimezone(tz.tzlocal()).strftime("%Y-%m-%d %H:%M:%S")]
        new_row += [cl.er]
        new_row += [stored_data[0]]  # Date 1
        new_row += [stored_data[1]]  # Client
        new_row += [stored_data[2]]  # Translator
        new_row += [stored_data[3]]  # Date 2
        new_row += [stored_data[4]]  # English Memory
        new_row += [stored_data[5]]  # French Memory
        stored_2 = checklist_data.values()
        data_2 = []
        for x in list(stored_2)[6:]:
            data_2 += x.values()
        new_row += data_2
        row_content_list += [new_row]
    table = TableCreation(checklist_type, max_revision,
                          row_content_list, job_no)
    return render_template('translation_index.html', table=table, job_no=job_no)


@APP.route('/translation_index/translation_add/<job_no>', methods=['GET', 'POST'])
@login_required
def translation_add(job_no):
    checklist_type = 'Translation'
    form = TranslationUpdate()
    form.translator.data = current_user.fname + ' ' + current_user.lname
    form.date1.data = datetime.now().date()
    form.date2.data = datetime.now().date()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).all()

    rev = []
    eng_review = []
    for x in cl:
        rev.append(x.revision)
        eng_review.append(x.er)
    revision = max(rev, default=0)+1
    eng_review_inter = [x for x in eng_review if x != '']
    eng_review = max(eng_review_inter, default=0)
    if len(cl) == 0 or len(eng_review_inter) == 0:
        form.er_level.data = 0
    else:
        form.er_level.data = int(eng_review) + 1
    if request.method == 'POST':
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        er = request.form['er_level']
        er = test_number(er)
        if er == "":
            completion_test = False
        translator_input = request.form['translator']
        if translator_input == "":
            translator_input = current_user.fname + ' ' + current_user.lname
        date1_input = request.form['date1']
        date2_input = request.form['date2']
        client_input = request.form['client']
        if client_input == "":
            completion_test = False
        memory_english_input = request.form['memory_english']
        if memory_english_input == "":
            completion_test = False
        memory_french_input = request.form['memory_french']
        if memory_french_input == "":
            completion_test = False
        data = {'Date 1': date1_input, 'Client': client_input, 'Translator': translator_input,
                'Date 2': date2_input, 'English Memory': memory_english_input, 'French Memory': memory_french_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'], request.form['input6'],
                       request.form['input7'], request.form['input8'],
                       request.form['input9'], request.form['input10'],
                       request.form['input11'], request.form['input12']]
        for subject in range(len(inputs_list)):
            input_valid = inputs_list[subject]
            data.update({subject: {'Verified': input_valid}})
            if input_valid == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date 1'] == "":
            date_data['Date 1'] = str(datetime.now().date())
        if date_data['Date 2'] == "":
            date_data['Date 2'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('translation_update.html',
                               revision=revision, job_no=job_no, form=form,
                               trans_english=trans_english, trans_french=trans_french)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        history = {1: ['Create', current_user.username, str(datetime.utcnow()), completed]}
        json_history = json.dumps(history)
        new_checklist = Checklists(checklist_type=checklist_type,
                                   created_by=current_user.username,
                                   completed=completed, op_manager=op_manager,
                                   date_created=datetime.utcnow(),
                                   revision=revision,
                                   er=er, checklist_history=json_history,
                                   data=JSON_data)
        job = Jobs.query.filter_by(job_no=job_no).first()
        job.checklists.append(new_checklist)
        DB.session.add(new_checklist)
        DB.session.commit()
        flash('New Translation Checklist Created!')
        return redirect(url_for('translation_index', job_no=job_no))
    else:
        return render_template('translation_update.html',
                               revision=revision, job_no=job_no, form=form,
                               trans_english=trans_english, trans_french=trans_french)


@APP.route('/translation_index/translation_update/<job_no>/<update_rev>',
           methods=['GET', 'POST'])
@login_required
def translation_update(job_no, update_rev):
    checklist_type = 'Translation'
    form = TranslationUpdate()
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
        filter_by(checklist_type=checklist_type).\
        filter_by(revision=update_rev).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    form.op_manager.data = cl.op_manager
    if cl.created_by == 'JIRA WEBHOOK':
        cl.created_by = current_user.username
        cl.date_created = datetime.utcnow()
        form.translator.data = current_user.fname + ' ' + current_user.lname
        form.date1.data = datetime.now().date()
        form.date2.data = datetime.now().date()
    else:
        form.translator.data = stored_data[2]
        form.date1.data = datetime.strptime(stored_data[0], '%Y-%m-%d')
        form.date2.data = datetime.strptime(stored_data[3], '%Y-%m-%d')
        form.client.data = stored_data[1]
        form.memory_english.data = stored_data[4]
        form.memory_french.data = stored_data[5]
    eng_review = cl.er
    form.er_level.data = eng_review
    input_2 = [form.input1, form.input2, form.input3,
               form.input4, form.input5,
               form.input6, form.input7, form.input8,
               form.input9, form.input10,
               form.input11, form.input12]
    for x in range(len(input_2)):
        if list(stored_data[6+x].values())[0] == "":
            input_2[x].data = ""
        else:
            input_2[x].data = list(stored_data[6+x].values())[0]

    if request.method == "POST":
        completion_test = True
        op_manager = request.form['op_manager']
        if op_manager == "":
            completion_test = False
        er = request.form['er_level']
        er = test_number(er)
        if er == "":
            completion_test = False
        translator_input = request.form['translator']
        if translator_input == "":
            translator_input = current_user.fname + ' ' + current_user.lname
        client_input = request.form['client']
        if client_input == "":
            completion_test = False
        date1_input = request.form['date1']
        date2_input = request.form['date2']
        memory_english_input = request.form['memory_english']
        if memory_english_input == "":
            completion_test = False
        memory_french_input = request.form['memory_french']
        if memory_french_input == "":
            completion_test = False
        data = {'Date 1': date1_input, 'Client': client_input, 'Translator': translator_input,
                'Date 2': date2_input, 'English Memory': memory_english_input, 'French Memory': memory_french_input}
        inputs_list = [request.form['input1'], request.form['input2'],
                       request.form['input3'], request.form['input4'],
                       request.form['input5'],
                       request.form['input6'], request.form['input7'],
                       request.form['input8'], request.form['input9'],
                       request.form['input10'],
                       request.form['input11'], request.form['input12']]
        for subject in range(len(inputs_list)):
            input_valid = inputs_list[subject]
            data.update({subject: {'Verified': input_valid}})
            if input_valid == "":
                completion_test = False
        JSON_data = json.dumps(data)
        date_data = json.loads(JSON_data)
        if date_data['Date 1'] == "":
            date_data['Date 1'] = str(datetime.now().date())
        if date_data['Date 2'] == "":
            date_data['Date 2'] = str(datetime.now().date())
        JSON_data = json.dumps(date_data)

        if 'submit_complete' in request.form:
            if completion_test == False:
                flash('All fields need to be filled in to save this checklist as complete.')
                return render_template('translation_update.html',
                               revision=update_rev, job_no=job_no, form=form,
                               trans_english=trans_english, trans_french=trans_french)
            else:
                completed = True
        if 'submit_incomplete' in request.form:
            completed = False
        cl.completed=completed
        cl.er = er
        cl.op_manager = op_manager
        cl.data = JSON_data

        history = cl.checklist_history
        parsed = json.loads(history)
        num_list = []
        for x in list(parsed.keys()):
            num_list += [int(x)]
        num = max(num_list)+1
        parsed.update({num: ['Edit', current_user.username, str(datetime.utcnow()), completed]})
        cl.checklist_history = json.dumps(parsed)
        DB.session.commit()
        flash('Translation Checklist Updated!')
        return redirect(url_for('translation_index', job_no=job_no))
    else:
        return render_template('translation_update.html',
                               revision=update_rev, job_no=job_no, form=form,
                               trans_english=trans_english, trans_french=trans_french)


@APP.route('/translation_index/translation_delete/<job_no>/<delete_rev>/<checklist_type>',
           methods=['GET', 'POST'])
@login_required
def translation_delete(job_no, delete_rev, checklist_type):
    if not current_user.admin:
        flash('Only Admins can delete checklists')
        return redirect(url_for('translation_index', job_no=job_no))
    else:
        delete_fn(job_no, delete_rev, checklist_type)
        return redirect(url_for('translation_index', job_no=job_no))

# Excel/PDF Exports
@APP.route('/translation_index/translation_export_pdf/<job_no>/<revision>', methods=['GET', 'POST'])
@login_required
def translation_export_pdf(job_no, revision):
    checklist_type = 'Translation'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=revision).first()
    checklist_data = json.loads(cl.data)
    job_no_input = str(job_no)
    stored_data = list(checklist_data.values())
    date1 = str(datetime.strptime(str(stored_data[0]), '%Y-%m-%d').strftime('%y-%m-%d'))
    client = str(stored_data[1])
    translator = str(stored_data[2])
    date2 = str(datetime.strptime(str(stored_data[3]), '%Y-%m-%d').strftime('%y-%m-%d'))
    memory_english = str(stored_data[4])
    memory_french = str(stored_data[5])

    list_of_data = []
    stored_2 = checklist_data.values()
    for x in list(stored_2)[6:]:
        list_of_data += x.values()
    checkbox_values = []
    for x in list_of_data:
        if str(x) == 'Verified':
            checkbox_values += ['/Yes']
        else:
            checkbox_values += ['/Off']

    loadpath = os.path.join(os.getcwd(), APP.config['CHECKLISTEXCEL_FOLDER'])
    infile = os.path.join(loadpath, 'translation_checklist.pdf')
    filename = '{}_checklist_{}.pdf'.format(checklist_type,
                                    datetime.now()
                                    .strftime("%Y-%m-%d_%H-%M-%S"))
    path = os.path.expanduser('~/Downloads')
    outfile = os.path.join(path, filename)

    ANNOT_KEY = '/Annots'
    ANNOT_FIELD_KEY = '/T'
    ANNOT_VAL_KEY = '/V'
    ANNOT_RECT_KEY = '/Rect'
    SUBTYPE_KEY = '/Subtype'
    WIDGET_SUBTYPE_KEY = '/Widget'

    def set_need_appearances_writer(writer):
        try:
            catalog = writer._root_object
            if "/AcroForm" not in catalog:
                writer._root_object.update({
                    NameObject("/AcroForm"): IndirectObject(len(writer._objects), 0, writer)})
            need_appearances = NameObject("/NeedAppearances")
            writer._root_object["/AcroForm"][need_appearances] = BooleanObject(True)
            return writer
        except Exception as e:
            return writer

    inputStream = open(infile, "rb")
    pdf_reader = PdfFileReader(inputStream, strict=False)
    if "/AcroForm" in pdf_reader.trailer["/Root"]:
        pdf_reader.trailer["/Root"]["/AcroForm"].update(
            {NameObject("/NeedAppearances"): BooleanObject(True)})

    pdf_writer = PdfFileWriter()
    set_need_appearances_writer(pdf_writer)
    if "/AcroForm" in pdf_writer._root_object:
        pdf_writer._root_object["/AcroForm"].update(
            {NameObject("/NeedAppearances"): BooleanObject(True)})
    field_dictionary = {'Text1': job_no_input, 'Text2': date1, 'Text3': client, 'Text4':memory_english,
                        'Text5': memory_french, 'Text6': translator, 'Text7': date2}
    fields = {'CheckBox1':checkbox_values[0], 'CheckBox2':checkbox_values[1], 'CheckBox3':checkbox_values[2],
               'CheckBox4':checkbox_values[3], 'CheckBox5':checkbox_values[4], 'CheckBox6':checkbox_values[5],
               'CheckBox7':checkbox_values[6], 'CheckBox8':checkbox_values[7], 'CheckBox9':checkbox_values[8],
               'CheckBox10':checkbox_values[9], 'CheckBox11':checkbox_values[10], 'CheckBox12':checkbox_values[11]}

    def updateCheckboxValues(page, fields):
        for j in range(0, len(page['/Annots'])):
            writer_annot = page['/Annots'][j].getObject()
            for field in fields:
                if writer_annot.get('/T') == field:
                    writer_annot.update({
                        NameObject("/V"): NameObject(fields[field]),
                        NameObject("/AS"): NameObject(fields[field])
                    })
    pdf_writer.addPage(pdf_reader.getPage(0))
    pdf_writer.updatePageFormFieldValues(pdf_writer.getPage(0), field_dictionary)
    updateCheckboxValues(pdf_writer.getPage(0), fields)
    outputStream = open(outfile, "wb")
    pdf_writer.write(outputStream)
    inputStream.close()
    outputStream.close()

    flash('{} downloaded'.format(filename))
    
    return redirect(url_for('translation_index', job_no=job_no))

@APP.route('/dpmo_index/dpmo_export_excel/<job_no>/<revision>', methods=['GET', 'POST'])
@login_required
def dpmo_export_excel(job_no, revision):
    checklist_type = 'DPMO'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=revision).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    date = stored_data[0]
    total_pages = stored_data[1]
    list_of_data = [] # list of 'number of errors'
    stored_2 = checklist_data.values()
    for x in list(stored_2)[3:]:
        list_of_data += x.values()
    list_of_rows = [12,13,14,16,18,19,20,21,22,23,25,26,27,29,31,
                    33,35,37,38,39,41,42,43,45,46,47,49,51,52,54,55,
                    57,58,59,60,61,62,63,64,65,67,68]

    loadpath = os.path.join(os.getcwd(), APP.config['CHECKLISTEXCEL_FOLDER'])
    loadpath = os.path.join(loadpath, '813648.xlsx')
    wb = load_workbook(loadpath)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet1.cell(row=3, column=3).value=str(date)
    Sheet1.cell(row=4, column=3).value=str(total_pages)

    for row_num, data in zip(list_of_rows, list_of_data):
        Sheet1.cell(row=row_num, column=4).value=str(data)

    filename = '{}_checklist_{}.xlsx'.format(checklist_type,
                                    datetime.now()
                                    .strftime("%Y-%m-%d_%H-%M-%S"))
    path = os.path.expanduser('~/Downloads')
    output_file = os.path.join(path, filename)
    wb.save(output_file)
    flash('{} downloaded'.format(filename))
    return redirect(url_for('dpmo_index', job_no=job_no))


@APP.route('/final_delivery_index/final_delivery_export_excel/<job_no>/<revision>', methods=['GET', 'POST'])
@login_required
def final_delivery_export_excel(job_no, revision):
    checklist_type = 'Final Delivery'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=revision).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    name = stored_data[0]
    date = stored_data[1]
    list_of_data = [] # list of 'number of errors'
    stored_2 = checklist_data.values()
    for x in list(stored_2)[2:]:
        list_of_data += x.values()
    list_of_rows = [6,7,8,10,11,12,13,14,16,17,18,19,21,22,23,25,26,
                    28,29,30,31,32,33,34,36,37,38,39,41,42]

    loadpath = os.path.join(os.getcwd(), APP.config['CHECKLISTEXCEL_FOLDER'])
    loadpath = os.path.join(loadpath, 'final_delivery.xlsx')
    wb = load_workbook(loadpath)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet1.cell(row=44, column=4).value=str(name)
    Sheet1.cell(row=46, column=4).value=str(date)

    for row_num, data in zip(list_of_rows, list_of_data):
        Sheet1.cell(row=row_num, column=5).value=str(data)

    filename = '{}_checklist_{}.xlsx'.format(checklist_type,
                                    datetime.now()
                                    .strftime("%Y-%m-%d_%H-%M-%S"))
    path = os.path.expanduser('~/Downloads')
    output_file = os.path.join(path, filename)
    wb.save(output_file)
    flash('{} downloaded'.format(filename))
    return redirect(url_for('final_delivery_index', job_no=job_no))


@APP.route('/data_conversion_index/data_conversion_export_excel/<job_no>/<revision>', methods=['GET', 'POST'])
@login_required
def data_conversion_export_excel(job_no, revision):
    checklist_type = 'Data Conversion'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=revision).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    pages = stored_data[0]
    loadpath = os.path.join(os.getcwd(), APP.config['CHECKLISTEXCEL_FOLDER'])
    loadpath = os.path.join(loadpath, 'data_conversion.xlsx')
    wb = load_workbook(loadpath)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet1.cell(row=3, column=4).value=str(pages)
    Sheet1.cell(row=2, column=4).value=str(job_no)
    for row_num, data in zip((range(len(stored_data[5:]))), stored_data[5:]):
        Sheet1.cell(row=row_num+6, column=4).value=str(data)

    filename = '{}_checklist_{}.xlsx'.format(checklist_type,
                                    datetime.now()
                                    .strftime("%Y-%m-%d_%H-%M-%S"))
    path = os.path.expanduser('~/Downloads')
    output_file = os.path.join(path, filename)
    wb.save(output_file)
    flash('{} downloaded'.format(filename))
    return redirect(url_for('data_conversion_index', job_no=job_no))


@APP.route('/writer_index/writer_export_excel/<job_no>/<revision>', methods=['GET', 'POST'])
@login_required
def writer_export_excel(job_no, revision):
    checklist_type = 'Writer'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=revision).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    if cl.er == '':
        subject = 'Comments'
    else:
        if int(cl.er) == 0:
            subject = 'Pre-Delivery Comments'
        else:
            subject = 'Engineering Review Comments #'+str(cl.er)
    pages = stored_data[0]
    writer = stored_data[1]
    date = stored_data[2]

    list_of_data = []
    stored_2 = checklist_data.values()
    for x in list(stored_2)[3:]:
        list_of_data += x.values()
    list_of_rows = list(range(6, 67))
    list_of_rows.remove(9)
    list_of_rows.remove(14)
    list_of_rows.remove(17)
    list_of_rows.remove(50)
    list_of_rows.remove(61)
    list_of_rows.remove(64)
    loadpath = os.path.join(os.getcwd(), APP.config['CHECKLISTEXCEL_FOLDER'])
    loadpath = os.path.join(loadpath, 'writer.xlsx')
    wb = load_workbook(loadpath)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet1.cell(row=68, column=5).value=str(pages)
    Sheet1.cell(row=70, column=5).value=str(writer)
    Sheet1.cell(row=72, column=5).value=str(date)
    Sheet1.cell(row=4, column=5).value=str(subject)

    for row_num, data in zip(list_of_rows, list_of_data):
        Sheet1.cell(row=row_num, column=5).value=str(data)
    filename = '{}_checklist_{}.xlsx'.format(checklist_type,
                                    datetime.now()
                                    .strftime("%Y-%m-%d_%H-%M-%S"))
    path = os.path.expanduser('~/Downloads')
    output_file = os.path.join(path, filename)
    wb.save(output_file)
    flash('{} downloaded'.format(filename))
    return redirect(url_for('writer_index', job_no=job_no))

@APP.route('/illustration_index/illustration_export_excel/<job_no>/<revision>', methods=['GET', 'POST'])
@login_required
def illustration_export_excel(job_no, revision):
    checklist_type = 'Illustration'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=revision).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    if cl.er == '':
        subject = 'Comments'
    else:
        if int(cl.er) == 0:
            subject = 'Pre-Delivery Comments'
        else:
            subject = 'Engineering Review Comments #'+str(cl.er)
    illustrator = stored_data[0]
    date = stored_data[1]
    comments = stored_data[2]
    list_of_data = []
    stored_2 = checklist_data.values()
    for x in list(stored_2)[3:]:
        list_of_data += x.values()
    list_of_rows = [6,7,8,9]
    loadpath = os.path.join(os.getcwd(), APP.config['CHECKLISTEXCEL_FOLDER'])
    loadpath = os.path.join(loadpath, 'illustration.xlsx')
    wb = load_workbook(loadpath)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet1.cell(row=11, column=5).value=str(illustrator)
    Sheet1.cell(row=13, column=5).value=str(date)
    Sheet1.cell(row=15, column=5).value=str(comments)
    Sheet1.cell(row=4, column=5).value=str(subject)

    for row_num, data in zip(list_of_rows, list_of_data):
        Sheet1.cell(row=row_num, column=5).value=str(data)
    filename = '{}_checklist_{}.xlsx'.format(checklist_type,
                                    datetime.now()
                                    .strftime("%Y-%m-%d_%H-%M-%S"))
    path = os.path.expanduser('~/Downloads')
    output_file = os.path.join(path, filename)
    wb.save(output_file)
    flash('{} downloaded'.format(filename))
    return redirect(url_for('illustration_index', job_no=job_no))


@APP.route('/qa_index/qa_export_excel/<job_no>/<revision>', methods=['GET', 'POST'])
@login_required
def qa_export_excel(job_no, revision):
    checklist_type = 'QA'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=revision).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    if cl.er == '':
        subject = 'Comments'
    else:
        if int(cl.er) == 0:
            subject = 'Pre-Delivery Comments'
        else:
            subject = 'Engineering Review Comments #'+str(cl.er)
    qa1 = stored_data[0]
    date1= stored_data[1]
    qa2 = stored_data[2]
    date2 =stored_data[3]
    qa3 =stored_data[4]
    date3 = stored_data[5]
    list_of_data = []
    stored_2 = checklist_data.values()
    for x in list(stored_2)[6:]:
        list_of_data += x.values()
    list_of_rows = [6,7,9,10,11,13,14,15,16,17,18,19,20,22,23,24,25,27,28,30,31,39,40,41,42,43,44,45]
    loadpath = os.path.join(os.getcwd(), APP.config['CHECKLISTEXCEL_FOLDER'])
    loadpath = os.path.join(loadpath, 'qa.xlsx')
    wb = load_workbook(loadpath)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet1.cell(row=33, column=5).value=str(qa1)
    Sheet1.cell(row=35, column=5).value=str(date1)
    Sheet1.cell(row=47, column=5).value=str(qa2)
    Sheet1.cell(row=49, column=5).value=str(date2)
    Sheet1.cell(row=57, column=5).value=str(qa3)
    Sheet1.cell(row=59, column=5).value=str(date3)
    Sheet1.cell(row=4, column=5).value=str(subject)

    for row_num, data in zip(list_of_rows, list_of_data):
        Sheet1.cell(row=row_num, column=5).value=str(data)
    filename = '{}_checklist_{}.xlsx'.format(checklist_type,
                                    datetime.now()
                                    .strftime("%Y-%m-%d_%H-%M-%S"))
    path = os.path.expanduser('~/Downloads')
    output_file = os.path.join(path, filename)
    wb.save(output_file)
    flash('{} downloaded'.format(filename))
    return redirect(url_for('qa_index', job_no=job_no))

@APP.route('/editor_index/editor_export_excel/<job_no>/<revision>', methods=['GET', 'POST'])
@login_required
def editor_export_excel(job_no, revision):
    checklist_type = 'Editor'
    cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
            filter_by(checklist_type=checklist_type).\
            filter_by(revision=revision).first()
    checklist_data = json.loads(cl.data)
    stored_data = list(checklist_data.values())
    if cl.er == '':
        subject = 'Comments'
    else:
        if int(cl.er) == 0:
            subject = 'Pre-Delivery Comments'
        else:
            subject = 'Engineering Review Comments #'+str(cl.er)
    editor = stored_data[0]
    date= stored_data[1]
    list_of_data = []
    stored_2 = checklist_data.values()
    for x in list(stored_2)[2:]:
        list_of_data += x.values()
    list_of_rows = [6,8,9,11,12,13,14]+list(range(16,37))+\
                    list(range(38,44))+list(range(45,49))+\
                    [51,52,53,55,56,57,58,60,61,62,63,65]
    loadpath = os.path.join(os.getcwd(), APP.config['CHECKLISTEXCEL_FOLDER'])
    loadpath = os.path.join(loadpath, 'editor.xlsx')
    wb = load_workbook(loadpath)
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet1.cell(row=67, column=5).value=str(editor)
    Sheet1.cell(row=69, column=5).value=str(date)
    Sheet1.cell(row=4, column=5).value=str(subject)

    for row_num, data in zip(list_of_rows, list_of_data):
        Sheet1.cell(row=row_num, column=5).value=str(data)
    filename = '{}_checklist_{}.xlsx'.format(checklist_type,
                                    datetime.now()
                                    .strftime("%Y-%m-%d_%H-%M-%S"))
    path = os.path.expanduser('~/Downloads')
    output_file = os.path.join(path, filename)
    wb.save(output_file)
    flash('{} downloaded'.format(filename))
    return redirect(url_for('editor_index', job_no=job_no))

# Exporting/Viewing a checklist

@APP.route('/dpmo_manager', methods=['GET', 'POST'])
@login_required
def dpmo_manager():
    if not current_user.admin:
        abort(403)
    form = DPMOManager()
    filters = form.filters()
    if request.method == "POST":
        query = Checklists.query.join(Checklists.jobs)

        def export(df, form=DPMOManager()):
            filename = '{}_{}.csv'.format('DPMO_metrics',
                                          datetime.now()
                                          .strftime("%Y-%m-%d_%H-%M-%S"))
            path = os.path.expanduser('~/Downloads')
            output_file = os.path.join(path, filename)
            export_csv = df.to_csv(output_file, index=None, header=False)
            flash('{} created!'.format(filename))

        if form.validate_on_submit():
            filters = form.filters()
            display_type = request.form['display_type']

            # MONTHLY METRICS TABLES
            if display_type == 'Monthly Metrics':
                month1 = request.form['month1']
                if month1 == "":
                    flash('Please Input a Month')
                    return render_template('dpmo_manager.html', form=form)
                else:
                    year_input = int(month1[0:4])
                    month_input = int(month1[-2:])
                    month_start = datetime(year_input, month_input, 1)
                    month_end_day = calendar.monthrange(year_input, month_input)[1]
                    month_end = datetime(year_input, month_input, month_end_day)
                    after = datetime(year_input, month_input, 1, 0, 0, 0)
                    after = after.replace(tzinfo=tz.tzlocal())
                    after = after.astimezone(tz.gettz('UTC'))
                    before = datetime(year_input, month_input, month_end_day, 23, 59, 59)
                    before = before.replace(tzinfo=tz.tzlocal())
                    before = before.astimezone(tz.gettz('UTC'))

                    query = Checklists.query.join(Checklists.jobs)
                    query = query.filter(Checklists.checklist_type == 'DPMO')
                    jn_list = []
                    if "Job_no" in filters:
                        jobs = filters["Job_no"].split(',')
                        for j in jobs:
                            job = Jobs.query.filter_by(job_no=j.strip()).first()
                            if job:
                                jn_list.append(job.job_no)
                        query = query.filter(Checklists.jobs.any(Jobs.job_no.in_(jn_list)))

                    if "User" in filters:
                        user_ids = []
                        users = filters["User"].split(',')
                        for u in users:
                            user = User.query.filter_by(username=u.strip()).first()
                            if user:
                                user_ids.append(user.username)
                        query = query.filter(Checklists.jobs.any
                                            (Checklists.created_by.in_(user_ids)))
                    else:
                        users = []

                    if "OM" in filters:
                        om_inter = []
                        om_list = filters["OM"].data.split(',')
                        for om in om_list:
                            om_inter.append(om)
                        om_inter = [x for x in om_inter if x != '']
                        if len(om_inter) != 0:
                            query = query.filter(Checklists.jobs.any
                                                    (Checklists.op_manager.in_(om_inter)))

                    query = query.filter(Checklists.date_created <= before)
                    query = query.filter(Checklists.date_created >= after)

                    if "Project" in filters:
                        project = filters["Project"]
                        if project != 'All':
                            query = query.filter(Jobs.project == project)
                    # if "Complete" in filters:
                    #     complete = filters['Complete']
                    #     if complete == 'Completed':
                    query = query.filter(Checklists.completed == True)
                    #     if complete == 'Incomplete':
                    #         query = query.filter(Checklists.completed == False)
                    if "ER" in filters:
                        eng_review = filters["ER"]
                        if eng_review == 'Pre-Delivery (ER Level 0)':
                            query = query.filter(Checklists.er == 0)
                        if eng_review == 'Engineering Review (ER Level 1+)':
                            query = query.filter(Checklists.er != 0)
                            query = query.filter(Checklists.er != "")
                        if eng_review == 'Latest':
                            if jn_list == []:
                                for value in Jobs.query.distinct(Jobs.job_no):
                                    jn_list.append(value.job_no)
                            qlist = []
                            for jn in jn_list:
                                qlist.append(query.filter(Jobs.job_no == jn).
                                            order_by(Checklists.er.desc()).first())
                            qlist = list(filter(None, qlist))

                    def order_qlist(elem):
                        return elem.date_created

                    if "Order" in filters:
                        order = filters["Order"]
                        if eng_review == 'Pre-Delivery (ER Level 0)':
                            if order == 'Ascending Order':
                                query = query.order_by(Checklists.date_created.asc())
                            else:
                                query = query.order_by(Checklists.date_created.desc())
                        if eng_review == 'Engineering Review (ER Level 1+)' or eng_review == 'All':
                            if order == 'Ascending Order':
                                query = query.order_by(Jobs.job_no.asc(),
                                                    Checklists.date_created.asc())
                            else:
                                query = query.order_by(Jobs.job_no.asc(),
                                                    Checklists.date_created.desc())
                        if eng_review == 'Latest':
                            if order == 'Ascending Order':
                                qlist.sort(key=order_qlist)
                            else:
                                qlist.sort(reverse=True, key=order_qlist)
                            results = qlist
                        else:
                            results = query.all()
                    if len(results) == 0:
                        flash('There are no checklists matching the current filters')
                        return render_template('dpmo_manager.html', form=form)
                    else:
                        table_list_monthly = dpmo_table_creation(results, before, after, year_input, month_input, 'Monthly Metrics')
                        if form.export.data:
                            export(table_list_monthly[11])
                        return render_template('dpmo_manager.html', form=form, table1=table_list_monthly[0], table2=table_list_monthly[1],
                                        table3=table_list_monthly[2], table4=table_list_monthly[3], table5=table_list_monthly[4], table6=table_list_monthly[5],
                                        table7=table_list_monthly[6], table8=table_list_monthly[7], fig_site=table_list_monthly[8],
                                        fig_pareto=table_list_monthly[9], fig_runchart=table_list_monthly[10])

            if display_type == 'Date Range':
                query = Checklists.query.join(Checklists.jobs)
                query = query.filter(Checklists.checklist_type == 'DPMO')
                jn_list = []
                if "Job_no" in filters:
                    jobs = filters["Job_no"].split(',')
                    for j in jobs:
                        job = Jobs.query.filter_by(job_no=j.strip()).first()
                        if job:
                            jn_list.append(job.job_no)
                    query = query.filter(Checklists.jobs.any(Jobs.job_no.in_(jn_list)))

                if "User" in filters:
                    user_ids = []
                    users = filters["User"].split(',')
                    for u in users:
                        user = User.query.filter_by(username=u.strip()).first()
                        if user:
                            user_ids.append(user.username)
                    query = query.filter(Checklists.jobs.any
                                        (Checklists.created_by.in_(user_ids)))
                else:
                    users = []

                if "OM" in filters:
                    om_inter = []
                    om_list = filters["OM"].data.split(',')
                    for om in om_list:
                        om_inter.append(om)
                    om_inter = [x for x in om_inter if x != '']
                    if len(om_inter) != 0:
                        query = query.filter(Checklists.jobs.any
                                                (Checklists.op_manager.in_(om_inter)))
                if "After" in filters:
                    after = filters["After"]
                    after = after.replace(tzinfo=tz.tzlocal())
                    after = after.astimezone(tz.gettz('UTC'))
                    query = query.filter(Checklists.date_created >= after)
                if "Before" in filters:
                    before = filters["Before"]
                    before = before.replace(tzinfo=tz.tzlocal())
                    before = before.astimezone(tz.gettz('UTC'))
                    query = query.filter(Checklists.date_created <= before)

                if "Project" in filters:
                    project = filters["Project"]
                    if project != 'All':
                        query = query.filter(Jobs.project == project)
                # if "Complete" in filters:
                #     complete = filters['Complete']
                #     if complete == 'Completed':
                query = query.filter(Checklists.completed == True)
                #     if complete == 'Incomplete':
                #         query = query.filter(Checklists.completed == False)
                if "ER" in filters:
                    eng_review = filters["ER"]
                    if eng_review == 'Pre-Delivery (ER Level 0)':
                        query = query.filter(Checklists.er == 0)
                    if eng_review == 'Engineering Review (ER Level 1+)':
                        query = query.filter(Checklists.er != 0)
                        query = query.filter(Checklists.er != "")
                    if eng_review == 'Latest':
                        if jn_list == []:
                            for value in Jobs.query.distinct(Jobs.job_no):
                                jn_list.append(value.job_no)
                        qlist = []
                        for jn in jn_list:
                            qlist.append(query.filter(Jobs.job_no == jn).
                                        order_by(Checklists.er.desc()).first())
                        qlist = list(filter(None, qlist))

                def order_qlist(elem):
                    return elem.date_created

                if "Order" in filters:
                    order = filters["Order"]
                    if eng_review == 'Pre-Delivery (ER Level 0)':
                        if order == 'Ascending Order':
                            query = query.order_by(Checklists.date_created.asc())
                        else:
                            query = query.order_by(Checklists.date_created.desc())
                    if eng_review == 'Engineering Review (ER Level 1+)' or eng_review == 'All':
                        if order == 'Ascending Order':
                            query = query.order_by(Jobs.job_no.asc(),
                                                Checklists.date_created.asc())
                        else:
                            query = query.order_by(Jobs.job_no.asc(),
                                                Checklists.date_created.desc())
                    if eng_review == 'Latest':
                        if order == 'Ascending Order':
                            qlist.sort(key=order_qlist)
                        else:
                            qlist.sort(reverse=True, key=order_qlist)
                        results = qlist
                    else:
                        results = query.all()
                if len(results) == 0:
                    flash('There are no checklists matching the current filters')
                    return render_template('dpmo_manager.html', form=form)
                else:
                    if "After" in filters:
                        after = filters["After"]
                        after = after.replace(tzinfo=tz.tzlocal())
                        after = after.astimezone(tz.gettz('UTC'))
                    else:
                        after = None
                    if "Before" in filters:
                        before = filters["Before"]
                        before = before.replace(tzinfo=tz.tzlocal())
                        before = before.astimezone(tz.gettz('UTC'))
                    else:
                        before = None
                    table_list_range = dpmo_table_creation(results, before, after, None, None, 'Date Range')
                    if form.export.data:
                        export(table_list_range[4])
                    return render_template('dpmo_manager.html', form=form, table1=table_list_range[0], table2=table_list_range[1],
                                table3=table_list_range[2], table4=table_list_range[3], export_df=table_list_range[4])
        else:
            return render_template('dpmo_manager.html', form=form)
    else:
        return render_template('dpmo_manager.html', form=form)


# Searches for the correct checklist(s)
@APP.route('/checklist_manager', methods=['GET', 'POST'])
@login_required
def checklist_manager():
    if not current_user.admin:
        abort(403)
    form = ChecklistManager()
    filters = form.filters()

    # Query
    query = Checklists.query.join(Checklists.jobs)
    if form.validate_on_submit():
        filters = form.filters()
        if "Checklist_type" in filters:
            checklist_type = filters['Checklist_type']
            query = query.filter(Checklists.checklist_type
                                     == checklist_type)
        jn_list = []
        if "Job_no" in filters:
            jobs = filters["Job_no"].split(',')
            for j in jobs:
                job = Jobs.query.filter_by(job_no=j.strip()).first()
                if job:
                    jn_list.append(job.job_no)
            query = query.filter(Checklists.jobs.any(Jobs.job_no.in_(jn_list)))

        if "User" in filters:
            user_ids = []
            users = filters["User"].split(',')
            for u in users:
                user = User.query.filter_by(username=u.strip()).first()
                if user:
                    user_ids.append(user.username)
            query = query.filter(Checklists.jobs.any
                                 (Checklists.created_by.in_(user_ids)))
        else:
            users = []
        if "OM" in filters:
            om_inter = []
            om_list = filters["OM"].data.split(',')
            for om in om_list:
                om_inter.append(om)
            om_inter = [x for x in om_inter if x != '']
            if len(om_inter) != 0:
                query = query.filter(Checklists.jobs.any
                                        (Checklists.op_manager.in_(om_inter)))
        if "After" in filters:
            after = filters["After"]
            after = after.replace(tzinfo=tz.tzlocal())
            after = after.astimezone(tz.gettz('UTC'))
            query = query.filter(Checklists.date_created >= after)
        if "Before" in filters:
            before = filters["Before"]
            before = before.replace(tzinfo=tz.tzlocal())
            before = before.astimezone(tz.gettz('UTC'))
            query = query.filter(Checklists.date_created <= before)
        if "Project" in filters:
            project = filters["Project"]
            if project != 'All':
                query = query.filter(Jobs.project == project)
        if "Complete" in filters:
            complete = filters['Complete']
            if complete == 'Completed':
                query = query.filter(Checklists.completed == True)
            if complete == 'Incomplete':
                query = query.filter(Checklists.completed == False)
        if "ER" in filters:
            eng_review = filters["ER"]
            if eng_review == 'Pre-Delivery (ER Level 0)':
                query = query.filter(Checklists.er == 0)
            if eng_review == 'Engineering Review (ER Level 1+)':
                query = query.filter(Checklists.er != 0)
                query = query.filter(Checklists.er != "")
            if eng_review == 'Latest':
                if jn_list == []:
                    for value in Jobs.query.distinct(Jobs.job_no):
                        jn_list.append(value.job_no)
                qlist = []
                for jn in jn_list:
                    qlist.append(query.filter(Jobs.job_no == jn).
                                 order_by(Checklists.er.desc()).first())
                qlist = list(filter(None, qlist))

        def order_qlist(elem):
            return elem.date_created

        if "Order" in filters:
            order = filters["Order"]
            if eng_review == 'Pre-Delivery (ER Level 0)':
                if order == 'Ascending Order':
                    query = query.order_by(Checklists.date_created.asc())
                else:
                    query = query.order_by(Checklists.date_created.desc())
            if eng_review == 'Engineering Review (ER Level 1+)' or eng_review == 'All':
                if order == 'Ascending Order':
                    query = query.order_by(Jobs.job_no.asc(),
                                           Checklists.date_created.asc())
                else:
                    query = query.order_by(Jobs.job_no.asc(),
                                           Checklists.date_created.desc())
            if eng_review == 'Latest':
                if order == 'Ascending Order':
                    qlist.sort(key=order_qlist)
                else:
                    qlist.sort(reverse=True, key=order_qlist)
                results = qlist
            else:
                results = query.all()
        if len(results) == 0:
            flash('There are no checklists matching the current filters')
            return render_template('checklist_manager.html', form=form)

        # Exports the dataframe as a csv
        def export(df, form=ChecklistManager()):
            filename = '{}_{}.csv'.format(checklist_type,
                                          datetime.now()
                                          .strftime("%Y-%m-%d_%H-%M-%S"))
            path = os.path.expanduser('~/Downloads')
            output_file = os.path.join(path, filename)
            export_csv = df.to_csv(output_file, index=None, header=False)
            flash('{} created!'.format(filename))

        # Creates a dataframe from the checklist(s)
        # All checklists except for 'Data Conversion''
        def default_export(checklist_type, type_blocks, all_sections,
                           subject_list):
            new_col = []
            export_letter = []
            export_titles = []
            export_number = []
            export_descript = []
            for letter, dict_title in type_blocks.items():
                export_letter += [letter]
                for title, dict_descript in dict_title.items():
                    export_titles += [title]
                    for number, descript in dict_descript.items():
                        export_number += [number]
                        export_descript += [descript]
            d = ['']*(10+len(subject_list))
            numbers = ['']*(10+len(subject_list))
            subjects = ['Job Number', 'Revision', 'ER Level', 'Type of Checklist', 'Date Created (YYYY-mm-dd hh:mm:ss)',
                        'Created By', 'Last Edited By', 'Date Last Edited', 'Status'] +\
                subject_list + ['Subject']
            combined = ['Job Number', 'Revision', 'ER Level', 'Type of Checklist', 'Date Created (YYYY-mm-dd hh:mm:ss)',
                        'Created By', 'Last Edited By', 'Date Last Edited', 'Status'] +\
                subject_list + ['Subject']
            for x, a, b in zip(all_sections, export_titles, export_letter):
                numbers += [a] + list(x.keys())
                subjects += [''] + list(x.values())
                d += [b] + ['']*len(x)
                combined += [a] + list(x.values())
            if checklist_type == 'DPMO':
                error_labels = ['']*(10+len(subject_list)) + ['', 'Major',  'Major', 'Minor', '', 'Major', '', 'Major', 'Major', 'Major', 
                'Major', 'Major', 'Minor', '', 'Major', 'Major', 'Major',  '', 'Major', '', 'Major', '', 'Major', '', 'Major', '', 
                'Major', 'Major', 'Minor', '', 'Major', 'Major', 'Major',  '', 'Major', 'Minor', 'Minor', '', 'Major', '', 'Major', 'Major', '',
                'Major', 'Minor', '', 'Major', 'Major', 'Major', 'Major', 'Major', 'Major', 'Major',  'Minor', 'Minor', '', 'Minor', 'Minor']

                df = DataFrame(data=error_labels)
                df['Subject'] = combined
            else:
                df = DataFrame(data=d)
                df['Section'] = numbers
                df['Subject'] = subjects
            for log in results:
                col_2 = []
                job = log.jobs
                current_jn = job[0].job_no
                checklist_data = json.loads(log.data)
                stored_data = checklist_data.values()
                data_2 = []
                for x in list(stored_data)[len(subject_list):]:
                    data_2 += x.values()
                if checklist_type == 'Final Delivery':
                    comment_type = 'Completed'
                # if checklist_type == 'Translation':
                #     comment_type = 'Verified'
                if checklist_type == 'DPMO':
                    comment_type = 'Number of Errors'
                else:
                    if log.er == '':
                        comment_type = 'Comments'
                    else:
                        if int(log.er) == 0:
                            comment_type = 'Pre-Delivery Comments'
                        else:
                            comment_type = 'Engineering Review Comments #' + str(log.er)
                if str(log.completed) == 'True':
                    status = 'Complete'
                else:
                    status = 'Incomplete'
                parsed = json.loads(log.checklist_history)
                num_list = []
                for x in list(parsed.keys()):
                    num_list += [int(x)]
                num = max(num_list)
                edit_by = parsed.get(str(num))[1]
                edit_date = datetime.strptime(parsed.get(str(num))[2], '%Y-%m-%d %H:%M:%S.%f')
                date_c = log.date_created.replace(tzinfo=tz.gettz('UTC'))
                date_c = (date_c.astimezone(tz.tzlocal())).strftime("%Y-%m-%d %H:%M:%S")
                date_e =edit_date.replace(tzinfo=tz.gettz('UTC'))
                date_e = (date_e.astimezone(tz.tzlocal())).strftime("%Y-%m-%d %H:%M:%S")
                data_initial = [current_jn, log.revision, log.er,
                                log.checklist_type,
                                str(date_c),
                                str(log.created_by), str(edit_by),
                                str(date_e), status] +\
                    list(stored_data)[0:len(subject_list)] + [comment_type]
                col_1 = [current_jn, log.revision, log.er,
                         log.checklist_type,
                         str(date_c),
                         str(log.created_by), str(edit_by),
                         str(date_e), status] +\
                    list(stored_data)[0:len(subject_list)]            
                temp = 0
                for x in all_sections:
                    data_initial += [''] + data_2[temp:(temp+len(x))]
                    col_2 += data_2[temp:(temp+len(x))]
                    temp += len(x)
                new_col += [col_1 + col_2]
                df['Job No: ' + str(current_jn) +
                   ' Rev: ' + str(log.revision)]\
                    = data_initial
            return [df, new_col]

        # Turning the Data Conversion checklist into a dataframe
        if checklist_type == 'Data Conversion':
            dc_export_di = []
            dc_export_descript = []
            dc_export_code = []
            for x, y in dc_blocks.items():
                for k, v in y.items():
                    dc_export_code += [x]
                    dc_export_di += [k]
                    dc_export_descript += [v]
            d = ['']*14 + ['Defect Intensity'] + dc_export_di
            df = DataFrame(data=d)
            df['Section'] = (['']*14 + ['Codes'] + dc_export_code)
            df['Subject'] = (['Job Number', 'Revision', 'ER Level', 'Type of Checklist', 'Date Created (YYYY-mm-dd hh:mm:ss)',
                            'Created By', 'Last Edited By', 'Date Last Edited', 'Status', 'Pages Reviewed',
                              'Total Major Errors',
                              'Total Minor Errors',
                              'Item 10 References as a Major PPM (<7000)',
                              'Item 10 References as a Minor PPM (<7000)']
                             + ['Description']
                             + dc_export_descript)
            new_col = []
            for log in results:
                job = log.jobs
                current_jn = job[0].job_no
                checklist_data = json.loads(log.data)
                if str(log.completed) == 'True':
                    status = 'Complete'
                else:
                    status = 'Incomplete'
                stored_data = list(checklist_data.values())
                parsed = json.loads(log.checklist_history)
                num_list = []
                for x in list(parsed.keys()):
                    num_list += [int(x)]
                num = max(num_list)
                edit_by = parsed.get(str(num))[1]
                edit_date = datetime.strptime(parsed.get(str(num))[2], '%Y-%m-%d %H:%M:%S.%f')
                date_c = log.date_created.replace(tzinfo=tz.gettz('UTC'))
                date_c = (date_c.astimezone(tz.tzlocal())).strftime("%Y-%m-%d %H:%M:%S")
                date_e =edit_date.replace(tzinfo=tz.gettz('UTC'))
                date_e = (date_e.astimezone(tz.tzlocal())).strftime("%Y-%m-%d %H:%M:%S")
                data_initial2 = [current_jn, log.revision, log.er,
                                log.checklist_type, str(date_c),
                                str(log.created_by), str(edit_by), str(date_e), status]+\
                    list(stored_data)[0:5]
                new_column2 = data_initial2 + stored_data[5:]
                new_col += [new_column2]
                df['Job No: ' + str(current_jn) + ' Rev: ' +
                    str(log.revision)] = data_initial2 + ['Number of Errors'] + stored_data[5:]
            col2 = new_col
            table_dc = TableViews('Data Conversion', col2)

            if form.export.data:
                export(df)
                return render_template('checklist_manager.html', form=form)
            return render_template('checklist_manager.html',
                                   form=form, graph=table_dc)

        if checklist_type == 'Translation':
            trans_subjects = []
            for eng, fr in zip(trans_english.values(), trans_french.values()):
                trans_subjects += [eng + '/' + fr]

            d = ['Job Number', 'Revision', 'ER Level', 'Type of Checklist', 'Date Created (YYYY-mm-dd hh:mm:ss)',
                            'Created By', 'Last Edited By', 'Date Last Edited', 'Status', 'Date', 'Client', 'Translator',
                            '(Signature) Date', 'Memory Used', 'Mémoire utilisée'] + ['Subject'] + trans_subjects
            df = DataFrame(data=d)
            new_col = []
            for log in results:
                job = log.jobs
                current_jn = job[0].job_no
                checklist_data = json.loads(log.data)
                stored_data = checklist_data.values()
                data_2 = []
                for x in list(stored_data)[6:]:
                    data_2 += x.values()

                job = log.jobs
                current_jn = job[0].job_no
                checklist_data = json.loads(log.data)
                if str(log.completed) == 'True':
                    status = 'Complete'
                else:
                    status = 'Incomplete'
                parsed = json.loads(log.checklist_history)
                num_list = []
                for x in list(parsed.keys()):
                    num_list += [int(x)]
                num = max(num_list)
                edit_by = parsed.get(str(num))[1]
                edit_date = datetime.strptime(parsed.get(str(num))[2], '%Y-%m-%d %H:%M:%S.%f')

                date_c = log.date_created.replace(tzinfo=tz.gettz('UTC'))
                date_c = (date_c.astimezone(tz.tzlocal())).strftime("%Y-%m-%d %H:%M:%S")
                date_e =edit_date.replace(tzinfo=tz.gettz('UTC'))
                date_e = (date_e.astimezone(tz.tzlocal())).strftime("%Y-%m-%d %H:%M:%S")
                data_initial2 = [current_jn, log.revision, log.er,
                                log.checklist_type, str(date_c),
                                str(log.created_by), str(edit_by), str(date_e), status]+\
                    list(stored_data)[0:6]

                new_column2 = data_initial2 + data_2
                new_col += [new_column2]
                df['Job No: ' + str(current_jn) + ' Rev: ' +
                    str(log.revision)] = data_initial2 + ['Verified'] + data_2
            col2 = new_col
            table_translation = TableViews('Translation', col2)

            if form.export.data:
                export(df)
                return render_template('checklist_manager.html', form=form)
            return render_template('checklist_manager.html',
                                   form=form, graph=table_translation)

        # Creating dataframes from each checklist type
        if checklist_type == 'Writer':
            df = default_export(checklist_type, writer_blocks, writer,
                                ['Total Pages', 'Writer', 'Date'])[0]
            cols = default_export(checklist_type, writer_blocks, writer,
                                ['Total Pages', 'Writer', 'Date'])[1]
            table = TableViews(checklist_type, cols)
            if form.export.data:
                export(df)
                return render_template('checklist_manager.html', form=form)
            return render_template('checklist_manager.html',
                                   form=form, graph=table)

        if checklist_type == 'Illustration':
            df = default_export(checklist_type, illustration_blocks,
                                illustration,
                                ['Illustrator', 'Date', 'Comments'])[0]
            cols = default_export(checklist_type, illustration_blocks,
                                illustration,
                                ['Illustrator', 'Date', 'Comments'])[1]
            table = TableViews(checklist_type, cols)
            if form.export.data:
                export(df)
                return render_template('checklist_manager.html', form=form)
            return render_template('checklist_manager.html',
                                   form=form, graph=table)

        if checklist_type == 'Editor':
            df = default_export(checklist_type, editor_blocks, editor,
                                ['Editor', 'Date'])[0]
            cols = default_export(checklist_type, editor_blocks, editor,
                                ['Editor', 'Date'])[1]
            table = TableViews(checklist_type, cols)
            if form.export.data:
                export(df)
                return render_template('checklist_manager.html', form=form)
            return render_template('checklist_manager.html',
                                   form=form, graph=table)

        if checklist_type == 'QA':
            df = default_export(checklist_type, qa_blocks, qa,
                                ['QA 1', 'Date 1', 'QA 2', 'Date 2',
                                 'QA 3', 'Date 3'])[0]
            cols = default_export(checklist_type, qa_blocks, qa,
                                ['QA 1', 'Date 1', 'QA 2', 'Date 2',
                                 'QA 3', 'Date 3'])[1]
            table = TableViews(checklist_type, cols)
            if form.export.data:
                export(df)
                return render_template('checklist_manager.html', form=form)
            return render_template('checklist_manager.html',
                                   form=form, graph=table)

        if checklist_type == 'Final Delivery':
            df = default_export(checklist_type, dl_blocks, dl,
                                ['Name', 'Date'])[0]
            cols = default_export(checklist_type, dl_blocks, dl,
                                ['Name', 'Date'])[1]
            table = TableViews(checklist_type, cols)
            if form.export.data:
                export(df)
                return render_template('checklist_manager.html', form=form)
            return render_template('checklist_manager.html',
                                   form=form, graph=table)

        if checklist_type == 'DPMO':
            df = default_export(checklist_type, dpmo_blocks, dpmo,
                                ['Date', 'Total Pages', 'DPMO Score'])[0]
            cols = default_export(checklist_type, dpmo_blocks, dpmo,
                                ['Date', 'Total Pages', 'DPMO Score'])[1]
            table = TableViews(checklist_type, cols)
            if form.export.data:
                export(df)
                return render_template('checklist_manager.html', form=form)
            return render_template('checklist_manager.html',
                                   form=form, graph=table)

    else:
        return render_template('checklist_manager.html', form=form)

@APP.route('/checklist_logs', methods=['GET', 'POST'])
@login_required
def checklist_logs():
    if not current_user.admin:
        abort(403)
    form = ChecklistHistory()

    def export(df, job_no, form=ChecklistHistory()):
        filename = '{}_{}_{}.csv'.format(str(job_no), 'checklist_log',
                                        datetime.now()
                                        .strftime("%Y-%m-%d_%H-%M-%S"))
        path = os.path.expanduser('~/Downloads')
        output_file = os.path.join(path, filename)
        export_csv = df.to_csv(output_file, index=None, header=False)
        flash('{} downloaded!'.format(filename))

    if request.method == "POST":
        job_no = request.form['job_no']
        figure_list = []
        df = DataFrame(data=[('','','','','', '')], columns=[0,1,2,3,4,5])
        for checklist_type in ['Data Conversion', 'Writer', 'Illustration', 'Editor', 'QA', 'Final Delivery', 'DPMO', 'Translation']:
            
            cl = Checklists.query.filter(Checklists.jobs.any(job_no=job_no)).\
                    filter_by(checklist_type=checklist_type).all()
            df = df.append(pandas.Series(['', '', '', '', '', ''], index=df.columns), ignore_index=True)
            df = df.append(pandas.Series([checklist_type, '', '', '', '', ''], index=df.columns), ignore_index=True)
            if len(cl) == 0:
                df = df.append(pandas.Series(["No " + checklist_type + " Checklists exists for this Job Number", '', '', '', '', ''], index=df.columns), ignore_index=True)
                table = "<table class='bottomBorder'><thead><tr><th colspan=6><b>"+ checklist_type +" Checklists</b></th><tr></thead>"
                table += "<tbody><tr><td colspan=6>No " + checklist_type + " Checklists exists for this Job Number</td></tr></tbody></table>"
                figure_list += [table]

            else:
                df = df.append(pandas.Series(['Revision', 'Count', 'Action', 'User', 'Date', 'Status'], index=df.columns), ignore_index=True)
                table = "<table class='bottomBorder'><thead><tr><th colspan=6><b>"+ checklist_type +" Checklists</b></th><tr>"
                table += "<tr><th><b>Revision</b></th><th><b>Count</b></th><th><b>Action</b></th><th><b>User</b></th><th><b>Date</b></th><th><b>Status</b></th></tr></thead>"
                table += "<tbody>"
                for log in cl:
                    checklist_history = json.loads(log.checklist_history)
                    for key in list(checklist_history.keys()):
                        date = datetime.strptime(checklist_history.get(key)[2], '%Y-%m-%d %H:%M:%S.%f')
                        date = date.replace(tzinfo=tz.gettz('UTC'))
                        date = date.astimezone(tz.tzlocal()).strftime('%Y-%m-%d %H:%M:%S')
                        if str(checklist_history.get(key)[3]) == 'True':
                            status = 'Complete'
                        else:
                            status = 'Incomplete'
                        table += "<tr><td>"+str(log.revision)+"</td><td>"+str(key)+"</td><td>"+str(checklist_history.get(key)[0])+"</td><td>"+\
                                        str(checklist_history.get(key)[1])+"</td><td>"+str(date)+"</td><td>"+status+"</td></tr>"
                        df = df.append(pandas.Series([log.revision, key, checklist_history.get(key)[0], checklist_history.get(key)[1], date, status], index=df.columns), ignore_index=True)
                table += "</tbody></table>"
                figure_list += [table]
        if form.export.data:
            export(df, job_no)
        return render_template('checklist_logs.html', form=form, data_conversion_logs=figure_list[0],
                                writer_logs=figure_list[1], illustration_logs=figure_list[2],
                                editor_logs=figure_list[3], qa_logs=figure_list[4],
                                final_delivery_logs=figure_list[5], dpmo_logs=figure_list[6], translation_logs=figure_list[7])
    else:
        return render_template('checklist_logs.html', form=form)
